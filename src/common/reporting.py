from __future__ import annotations

from pathlib import Path
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

def md_table(df: pd.DataFrame, max_rows: int | None = None) -> str:
    """Format a pandas DataFrame as a GitHub-style markdown table."""
    if df is None or df.empty:
        return "No rows."
    d = df.head(max_rows).replace({np.nan: ""}) if max_rows else df.replace({np.nan: ""})
    out = ["| " + " | ".join(map(str, d.columns)) + " |", "| " + " | ".join(["---"] * len(d.columns)) + " |"]
    for _, row in d.iterrows():
        out.append("| " + " | ".join(str(v).replace("\n", " ").replace("|", "/") for v in row.tolist()) + " |")
    return "\n".join(out)

def autosize_xlsx(path: Path) -> None:
    """Style Excel sheets, freeze panes, style headers, and auto-fit columns."""
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        
        # Style header row
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            
        # Style columns
        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            # Scan top 200 rows to determine max text length
            width = max(len("" if c.value is None else str(c.value)) for c in col[:200])
            ws.column_dimensions[letter].width = min(max(width + 2, 10), 45)
            
    wb.save(path)
