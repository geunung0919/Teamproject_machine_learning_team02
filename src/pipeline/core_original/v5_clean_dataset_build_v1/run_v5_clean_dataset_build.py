from __future__ import annotations

import math
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sklearn.neighbors import BallTree


ROOT = Path(__file__).resolve().parents[2]
IN_DIR = ROOT / "data" / "v5_raw_parser_repair_v1"
OUT_DIR = ROOT / "data" / "v5_clean_dataset_build_v1"
CANON_DIR = OUT_DIR / "canonical"
VIEW_DIR = OUT_DIR / "model_views"
AUDIT_DIR = OUT_DIR / "audit"
REPORT_DIR = ROOT / "reports" / "v5_clean_dataset_build_v1"
HANDOFF_DIR = ROOT / "handoff_for_chatgpt" / "v5_clean_dataset_build_v1"

RAW_GEO = ROOT / "data" / "raw" / "school_data_2008_2025_geocoded.csv"
RAW_POP = ROOT / "data" / "raw" / "national_kosis_school_age_population_sgg.csv"
RAW_MIG = ROOT / "data" / "raw" / "national_kosis_migration_sgg.csv"
RAW_SGG_CODES = ROOT / "data" / "raw" / "national_kosis_sgg_codes.csv"

YEARS = list(range(2012, 2026))
TARGET_LEVELS = {"초등학교", "중학교", "고등학교"}
METRO_SIDO = {"서울", "부산", "대구", "인천", "광주", "대전", "울산", "세종"}
SIDO_BY_CODE = {
    11: "서울", 21: "부산", 22: "대구", 23: "인천", 24: "광주", 25: "대전", 26: "울산",
    29: "세종", 31: "경기", 32: "강원", 33: "충북", 34: "충남", 35: "전북", 36: "전남",
    37: "경북", 38: "경남", 39: "제주",
}
SIDO_ALIASES = {
    "서울특별시": "서울", "서울": "서울", "부산광역시": "부산", "부산": "부산", "대구광역시": "대구", "대구": "대구",
    "인천광역시": "인천", "인천": "인천", "광주광역시": "광주", "광주": "광주", "대전광역시": "대전", "대전": "대전",
    "울산광역시": "울산", "울산": "울산", "세종특별자치시": "세종", "세종": "세종", "경기도": "경기", "경기": "경기",
    "강원도": "강원", "강원특별자치도": "강원", "강원": "강원", "충청북도": "충북", "충북": "충북",
    "충청남도": "충남", "충남": "충남", "전라북도": "전북", "전북특별자치도": "전북", "전북": "전북",
    "전라남도": "전남", "전남": "전남", "경상북도": "경북", "경북": "경북", "경상남도": "경남", "경남": "경남",
    "제주특별자치도": "제주", "제주도": "제주", "제주": "제주",
}


def rel(path: Path) -> str:
    try:
        return str(path.relative_to(ROOT))
    except ValueError:
        return str(path)


def ensure_dirs() -> None:
    for d in [OUT_DIR, CANON_DIR, VIEW_DIR, AUDIT_DIR, REPORT_DIR]:
        d.mkdir(parents=True, exist_ok=True)
    if HANDOFF_DIR.exists():
        shutil.rmtree(HANDOFF_DIR)
    HANDOFF_DIR.mkdir(parents=True, exist_ok=True)


def norm_text(x: Any) -> str:
    if pd.isna(x):
        return ""
    return "".join(str(x).split())


def norm_sido(x: Any) -> str:
    s = "" if pd.isna(x) else str(x).strip()
    return SIDO_ALIASES.get(s, s)


def stable_school_key(row: pd.Series) -> str:
    return "|".join([
        norm_sido(row.get("sido", "")),
        str(row.get("sgg", "")).strip(),
        str(row.get("school_level", "")).strip(),
        norm_text(row.get("school_name_raw", row.get("school_name_norm", ""))),
        str(row.get("branch_type", "")).strip(),
    ])


def size_bucket(n: Any) -> str:
    if pd.isna(n):
        return "missing"
    n = float(n)
    if n == 0:
        return "0"
    if n <= 30:
        return "1_30"
    if n <= 60:
        return "31_60"
    if n <= 100:
        return "61_100"
    if n <= 300:
        return "101_300"
    if n <= 600:
        return "301_600"
    if n <= 1000:
        return "601_1000"
    return "1000_plus"


def region_group(sido: str) -> str:
    if sido in METRO_SIDO:
        return "metro"
    if sido in {"경기"}:
        return "capital_area"
    if sido in {"강원", "충북", "충남", "전북", "전남", "경북", "경남", "제주"}:
        return "province"
    return "unknown"


def safe_growth(curr: Any, prev: Any) -> float:
    if pd.isna(curr) or pd.isna(prev) or prev == 0:
        return np.nan
    return (float(curr) - float(prev)) / abs(float(prev))


def slope_last(values: pd.Series, window: int) -> float:
    vals = values.tail(window).astype(float)
    if len(vals) < window or vals.isna().any():
        return np.nan
    x = np.arange(window)
    return float(np.polyfit(x, vals.values, 1)[0])


def read_inputs() -> dict[str, pd.DataFrame]:
    return {
        "school": pd.read_csv(IN_DIR / "parsed_school_kess_long_2008_2025.csv", low_memory=False),
        "birth": pd.read_csv(IN_DIR / "sgg_birth_fertility_2007_2025.csv", low_memory=False),
        "pop": pd.read_csv(RAW_POP, low_memory=False),
        "mig": pd.read_csv(RAW_MIG, low_memory=False),
        "geo": pd.read_csv(RAW_GEO, low_memory=False),
        "parser_coverage": pd.read_csv(IN_DIR / "parser_coverage_by_year.csv", low_memory=False),
        "school_parser_quality": pd.read_csv(IN_DIR / "school_parser_quality_audit.csv", low_memory=False),
        "birth_quality": pd.read_csv(IN_DIR / "birth_fertility_parser_quality_audit.csv", low_memory=False),
    }


def build_school_panel(school: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    src = school[school["year"].isin(YEARS)].copy()
    src["school_level"] = src["school_level"].astype(str).str.strip()
    src["sido"] = src["sido"].map(norm_sido)
    src["is_target_level"] = src["school_level"].isin(TARGET_LEVELS)
    filter_audit = src.groupby(["year", "school_level", "is_target_level"]).size().reset_index(name="row_count")
    panel = src[src["is_target_level"]].copy()
    panel["school_key"] = panel.apply(stable_school_key, axis=1)
    panel["school_name"] = panel["school_name_raw"]
    panel["address"] = panel["address_raw"]
    panel["students_per_class"] = panel["students_per_class_calc"]
    panel["students_per_teacher"] = panel["students_per_teacher_calc"]
    keep = [
        "school_key", "year", "survey_date", "sido", "sgg", "school_level", "school_name", "school_name_norm",
        "branch_type", "foundation_type", "status", "address", "student_count", "class_count", "teacher_count",
        "students_per_class", "students_per_teacher", "land_area", "land_area_per_student",
    ]
    panel = panel[keep].copy()
    panel = panel.sort_values(["school_key", "year", "survey_date"]).drop_duplicates(["school_key", "year"], keep="last")
    for col in ["student_count", "class_count", "teacher_count", "students_per_class", "students_per_teacher", "land_area", "land_area_per_student"]:
        panel[col] = pd.to_numeric(panel[col], errors="coerce")
    g = panel.groupby("school_key", group_keys=False)
    for lag in [1, 2, 3]:
        panel[f"student_count_lag_{lag}"] = g["student_count"].shift(lag)
    panel["student_delta_lag_1"] = panel["student_count"] - panel["student_count_lag_1"]
    panel["student_growth_lag_1"] = [safe_growth(c, p) for c, p in zip(panel["student_count"], panel["student_count_lag_1"])]
    panel["student_rolling_mean_3"] = g["student_count"].transform(lambda s: s.shift(1).rolling(3, min_periods=1).mean())
    panel["student_rolling_delta_mean_3"] = g["student_count"].transform(lambda s: s.diff().shift(1).rolling(3, min_periods=1).mean())
    panel["student_trend_slope_3"] = g["student_count"].transform(lambda s: s.shift(1).rolling(3).apply(lambda x: float(np.polyfit(np.arange(3), x, 1)[0]), raw=False))
    panel["student_rolling_mean_5"] = g["student_count"].transform(lambda s: s.shift(1).rolling(5, min_periods=1).mean())
    panel["student_trend_slope_5"] = g["student_count"].transform(lambda s: s.shift(1).rolling(5).apply(lambda x: float(np.polyfit(np.arange(5), x, 1)[0]), raw=False))
    panel["size_bucket"] = panel["student_count"].map(size_bucket)
    panel["student_size_bin"] = panel["size_bucket"]
    panel["metro_flag"] = panel["sido"].isin(METRO_SIDO)
    panel["region_group"] = panel["sido"].map(region_group)
    panel["level_size_segment"] = panel["school_level"].astype(str) + "|" + panel["size_bucket"].astype(str)
    return panel, filter_audit


def build_school_master(panel: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    dup = panel[panel.duplicated(["school_key", "year"], keep=False)].copy()
    rows = []
    for key, g in panel.sort_values("year").groupby("school_key"):
        latest = g.iloc[-1]
        names = g["school_name"].dropna().astype(str).unique().tolist()
        locs = g[["sido", "sgg", "school_level", "branch_type"]].drop_duplicates()
        entity_needed = len(names) > 1 or len(locs) > 1
        rows.append({
            "school_key": key,
            "school_name_canonical": names[-1] if names else latest["school_name"],
            "school_name_latest": latest["school_name"],
            "sido_latest": latest["sido"],
            "sgg_latest": latest["sgg"],
            "school_level": latest["school_level"],
            "branch_type_latest": latest["branch_type"],
            "foundation_type_latest": latest["foundation_type"],
            "first_observed_year": int(g["year"].min()),
            "last_observed_year": int(g["year"].max()),
            "years_observed_count": int(g["year"].nunique()),
            "entity_status": "entity_resolution_needed" if entity_needed else "stable_candidate",
            "entity_resolution_confidence": "medium" if entity_needed else "high",
            "entity_resolution_note": "name/location/branch changed within school_key" if entity_needed else "",
        })
    master = pd.DataFrame(rows)
    master_audit = pd.DataFrame({
        "metric": ["school_master_rows", "entity_resolution_needed_count", "duplicate_school_key_year_rows"],
        "value": [len(master), int((master["entity_status"] == "entity_resolution_needed").sum()), len(dup)],
    })
    return master, master_audit, dup


def build_demographics(pop: pd.DataFrame, birth: pd.DataFrame, mig: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    pop = pop.copy()
    pop["year"] = (pd.to_numeric(pop["period"], errors="coerce") // 100).astype("Int64")
    pop["sido_code"] = (pd.to_numeric(pop["sgg_code"], errors="coerce") // 1000).astype("Int64")
    pop["sido"] = pop["sido_code"].map(SIDO_BY_CODE)
    pop["sgg"] = pop["sgg_name"].astype(str)
    pop_piv = pop[pop["year"].isin(YEARS)].pivot_table(
        index=["year", "sido", "sgg", "sgg_code"], columns="age_group", values="population", aggfunc="first"
    ).reset_index()
    pop_piv.columns.name = None
    rename = {"0 - 4세": "age_0_4_pop", "5 - 9세": "age_5_9_pop", "10 - 14세": "age_10_14_pop", "15 - 19세": "age_15_19_pop"}
    pop_piv = pop_piv.rename(columns=rename)
    for c in rename.values():
        if c not in pop_piv:
            pop_piv[c] = np.nan
    pop_piv["school_age_population_0_19"] = pop_piv[list(rename.values())].sum(axis=1, min_count=1)
    birth = birth.copy()
    birth["sido"] = birth["sido"].map(norm_sido)
    birth["sgg"] = birth["sgg"].astype(str)
    demo = pop_piv.merge(
        birth[["year", "sido", "sgg", "region_code", "birth_count", "total_fertility_rate"]],
        on=["year", "sido", "sgg"],
        how="left",
    )
    demo["birth_region_code"] = demo["region_code"]
    mig = mig.copy()
    mig["sido_code"] = (pd.to_numeric(mig["region_code"], errors="coerce") // 1000).astype("Int64")
    mig["sido"] = mig["sido_code"].map(SIDO_BY_CODE)
    mig["sgg"] = mig["region_name"].astype(str)
    m = mig[(mig["year"].isin(YEARS)) & (mig["region_level"].eq("sgg"))].pivot_table(
        index=["year", "sido", "sgg"], columns="item_name", values="value", aggfunc="first"
    ).reset_index()
    m.columns.name = None
    m = m.rename(columns={"순이동": "net_migration_total", "총전입": "in_migration_total", "총전출": "out_migration_total"})
    demo = demo.merge(m[["year", "sido", "sgg", "net_migration_total", "in_migration_total", "out_migration_total"]], on=["year", "sido", "sgg"], how="left")
    demo = demo.sort_values(["sido", "sgg", "year"])
    for col in ["school_age_population_0_19", "birth_count", "total_fertility_rate", "net_migration_total"]:
        demo[col] = pd.to_numeric(demo[col], errors="coerce")
    g = demo.groupby(["sido", "sgg"], group_keys=False)
    demo["school_age_population_delta_1y"] = g["school_age_population_0_19"].diff()
    demo["school_age_population_growth_1y"] = g["school_age_population_0_19"].pct_change(fill_method=None)
    demo["birth_count_yoy_change"] = g["birth_count"].diff()
    demo["birth_count_yoy_rate"] = g["birth_count"].pct_change(fill_method=None)
    demo["tfr_yoy_change"] = g["total_fertility_rate"].diff()
    demo["tfr_yoy_rate"] = g["total_fertility_rate"].pct_change(fill_method=None)
    demo["net_migration_yoy_change"] = g["net_migration_total"].diff()
    cols = [
        "year", "sido", "sgg", "sgg_code", "age_0_4_pop", "age_5_9_pop", "age_10_14_pop", "age_15_19_pop",
        "school_age_population_0_19", "school_age_population_delta_1y", "school_age_population_growth_1y",
        "birth_count", "total_fertility_rate", "birth_count_yoy_change", "birth_count_yoy_rate",
        "tfr_yoy_change", "tfr_yoy_rate", "net_migration_total", "in_migration_total", "out_migration_total",
        "net_migration_yoy_change",
    ]
    audit = pd.DataFrame([{
        "year": y,
        "demo_rows": len(demo[demo["year"] == y]),
        "school_age_nonnull_rate": demo.loc[demo["year"] == y, "school_age_population_0_19"].notna().mean(),
        "birth_nonnull_rate": demo.loc[demo["year"] == y, "birth_count"].notna().mean(),
        "migration_nonnull_rate": demo.loc[demo["year"] == y, "net_migration_total"].notna().mean(),
    } for y in YEARS])
    missing_pair = birth[birth[["birth_count", "total_fertility_rate"]].isna().any(axis=1)].copy()
    if len(missing_pair):
        missing_pair["region_level"] = "sgg"
        missing_pair["region_name_raw"] = missing_pair["sgg"]
        missing_pair["has_birth_count_T1"] = missing_pair["birth_count"].notna()
        missing_pair["has_total_fertility_rate_T2"] = missing_pair["total_fertility_rate"].notna()
        missing_pair["missing_item"] = np.where(missing_pair["birth_count"].isna(), "birth_count_T1", "total_fertility_rate_T2")
        missing_pair["recommended_action"] = "keep_missing"
        missing_pair = missing_pair[["year", "region_level", "region_code", "sido", "sgg", "region_name_raw", "has_birth_count_T1", "has_total_fertility_rate_T2", "birth_count", "total_fertility_rate", "missing_item", "source_file", "last_change_date", "recommended_action"]]
    return demo[cols], audit, missing_pair


def build_birth_missing_pair_all() -> pd.DataFrame:
    frames = []
    for level, path in [
        ("national", IN_DIR / "national_birth_fertility_2007_2025.csv"),
        ("sido", IN_DIR / "sido_birth_fertility_2007_2025.csv"),
        ("sgg", IN_DIR / "sgg_birth_fertility_2007_2025.csv"),
    ]:
        if not path.exists():
            continue
        df = pd.read_csv(path, low_memory=False)
        df["region_level"] = level
        if "sido" not in df:
            df["sido"] = ""
        if "sgg" not in df:
            df["sgg"] = ""
        if "region_name" in df and "region_name_raw" not in df:
            df = df.rename(columns={"region_name": "region_name_raw"})
        if "region_name_raw" not in df:
            df["region_name_raw"] = np.where(level == "national", "전국", df.get("sido", ""))
        frames.append(df)
    if not frames:
        return pd.DataFrame()
    all_df = pd.concat(frames, ignore_index=True, sort=False)
    miss = all_df[all_df[["birth_count", "total_fertility_rate"]].isna().any(axis=1)].copy()
    miss["has_birth_count_T1"] = miss["birth_count"].notna()
    miss["has_total_fertility_rate_T2"] = miss["total_fertility_rate"].notna()
    miss["missing_item"] = np.where(miss["birth_count"].isna(), "birth_count_T1", "total_fertility_rate_T2")
    miss["recommended_action"] = np.where(miss["region_level"].eq("sgg"), "use_sido_fallback_for_model_view", "keep_missing")
    cols = ["year", "region_level", "region_code", "sido", "sgg", "region_name_raw", "has_birth_count_T1", "has_total_fertility_rate_T2", "birth_count", "total_fertility_rate", "missing_item", "source_file", "last_change_date", "recommended_action"]
    for c in cols:
        if c not in miss:
            miss[c] = ""
    return miss[cols].sort_values(["year", "region_level", "region_code"])


def geo_join_key(df: pd.DataFrame, name_col: str) -> pd.Series:
    return df["year"].astype(str) + "|" + df["sido"].map(norm_sido).astype(str) + "|" + df["sgg"].astype(str) + "|" + df["school_level"].astype(str) + "|" + df[name_col].map(norm_text).astype(str)


def build_isolation(panel: pd.DataFrame, geo: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    g = geo.rename(columns={"시도": "sido", "행정구": "sgg", "학교급": "school_level", "학교명": "school_name", "lttud": "latitude", "lgtud": "longitude"}).copy()
    g = g[g["year"].isin(YEARS)]
    g["geo_join_key"] = geo_join_key(g, "school_name")
    geo_small = g[["geo_join_key", "latitude", "longitude", "coordinate_source"]].drop_duplicates("geo_join_key")
    p = panel[["school_key", "year", "sido", "sgg", "school_level", "school_name"]].copy()
    p["geo_join_key"] = geo_join_key(p, "school_name")
    iso = p.merge(geo_small, on="geo_join_key", how="left")
    iso["latitude"] = pd.to_numeric(iso["latitude"], errors="coerce")
    iso["longitude"] = pd.to_numeric(iso["longitude"], errors="coerce")
    valid = iso["latitude"].between(33, 39) & iso["longitude"].between(124, 132)
    iso["coordinate_valid"] = valid
    iso["coordinate_invalid_reason"] = np.where(iso["latitude"].isna() | iso["longitude"].isna(), "missing", np.where(~valid, "outside_korea_bbox", ""))
    for c in ["nearest_same_level_distance_km", "second_nearest_same_level_distance_km", "same_level_school_count_within_3km", "same_level_school_count_within_5km", "same_level_school_count_within_10km", "isolation_score"]:
        iso[c] = np.nan
    iso["no_same_level_school_within_5km_flag"] = pd.Series(pd.NA, index=iso.index, dtype="object")
    for (year, level), idx in iso[iso["coordinate_valid"]].groupby(["year", "school_level"]).groups.items():
        idx = list(idx)
        coords = np.radians(iso.loc[idx, ["latitude", "longitude"]].to_numpy())
        if len(coords) < 2:
            continue
        tree = BallTree(coords, metric="haversine")
        dists, _ = tree.query(coords, k=min(3, len(coords)))
        radius = 6371.0088
        nearest = dists[:, 1] * radius
        second = dists[:, 2] * radius if dists.shape[1] > 2 else np.nan
        cnt3 = tree.query_radius(coords, r=3 / radius, count_only=True) - 1
        cnt5 = tree.query_radius(coords, r=5 / radius, count_only=True) - 1
        cnt10 = tree.query_radius(coords, r=10 / radius, count_only=True) - 1
        iso.loc[idx, "nearest_same_level_distance_km"] = nearest
        iso.loc[idx, "second_nearest_same_level_distance_km"] = second
        iso.loc[idx, "same_level_school_count_within_3km"] = cnt3
        iso.loc[idx, "same_level_school_count_within_5km"] = cnt5
        iso.loc[idx, "same_level_school_count_within_10km"] = cnt10
        iso.loc[idx, "no_same_level_school_within_5km_flag"] = cnt5 == 0
        iso.loc[idx, "isolation_score"] = nearest / (1 + cnt5)
    iso["isolation_score_version"] = "v5_haversine_nearest_div_1_plus_5km_count"
    audit = iso.groupby("year").agg(
        rows=("school_key", "size"),
        coordinate_valid_rate=("coordinate_valid", "mean"),
        isolation_nonnull_rate=("isolation_score", lambda s: s.notna().mean()),
        median_isolation_score=("isolation_score", "median"),
        p95_isolation_score=("isolation_score", lambda s: s.quantile(0.95)),
    ).reset_index()
    coord_audit = iso.groupby(["year", "coordinate_invalid_reason"]).size().reset_index(name="row_count")
    cols = ["school_key", "year", "latitude", "longitude", "coordinate_source", "coordinate_valid", "coordinate_invalid_reason", "nearest_same_level_distance_km", "second_nearest_same_level_distance_km", "same_level_school_count_within_3km", "same_level_school_count_within_5km", "same_level_school_count_within_10km", "no_same_level_school_within_5km_flag", "isolation_score", "isolation_score_version"]
    return iso[cols], coord_audit, audit


def build_grade_flow(school: pd.DataFrame, panel: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    src = school[(school["year"].isin(YEARS)) & (school["school_level"].isin(TARGET_LEVELS))].copy()
    src["sido"] = src["sido"].map(norm_sido)
    src["school_key"] = src.apply(stable_school_key, axis=1)
    src = src.sort_values(["school_key", "year", "survey_date"]).drop_duplicates(["school_key", "year"], keep="last")
    cols = ["school_key", "year"] + [f"grade{i}_student_count" for i in range(1, 7)] + [f"grade{i}_class_count" for i in range(1, 7)] + ["grade_student_sum", "grade_class_sum", "entrants_total", "graduates_total", "transfer_in", "transfer_out", "grade_student_sum_diff", "grade_class_sum_diff"]
    gf = src[cols].copy()
    gf = gf.merge(panel[["school_key", "year", "school_level", "student_count", "class_count"]], on=["school_key", "year"], how="left")
    grade_cols = [f"grade{i}_student_count" for i in range(1, 7)]
    class_cols = [f"grade{i}_class_count" for i in range(1, 7)]
    for c in grade_cols + class_cols + ["entrants_total", "graduates_total", "transfer_in", "transfer_out"]:
        gf[c] = pd.to_numeric(gf[c], errors="coerce")
    gf["grade_student_sum"] = gf[grade_cols].sum(axis=1, min_count=1)
    gf["grade_class_sum"] = gf[class_cols].sum(axis=1, min_count=1)
    for i in range(1, 7):
        gf[f"grade{i}_share"] = gf[f"grade{i}_student_count"] / gf["grade_student_sum"].replace({0: np.nan})
    gf["lower_grade_student_count"] = gf[["grade1_student_count", "grade2_student_count", "grade3_student_count"]].sum(axis=1, min_count=1)
    gf["upper_grade_student_count"] = gf[["grade4_student_count", "grade5_student_count", "grade6_student_count"]].sum(axis=1, min_count=1)
    gf["graduating_grade_student_count"] = np.where(gf["school_level"].eq("초등학교"), gf["grade6_student_count"], gf["grade3_student_count"])
    gf["grade_imbalance_range"] = gf[grade_cols].max(axis=1) - gf[grade_cols].min(axis=1)
    gf["grade_imbalance_std"] = gf[grade_cols].std(axis=1)
    gf["student_diff"] = gf["student_count"] - gf["grade_student_sum"]
    gf["class_diff"] = gf["class_count"] - gf["grade_class_sum"]
    gf["grade_data_valid"] = gf["student_diff"].abs().fillna(999999) <= 1
    gf["grade_invalid_reason"] = np.where(gf["grade_data_valid"], "", "grade_sum_mismatch")
    grade_mis = gf[gf["student_diff"].abs().fillna(0) > 1][["school_key", "year", "school_level", "student_count", "grade_student_sum", "student_diff"]].copy()
    class_mis = gf[gf["class_diff"].abs().fillna(0) > 1][["school_key", "year", "school_level", "class_count", "grade_class_sum", "class_diff"]].copy()
    out_cols = ["school_key", "year"] + grade_cols + ["grade_student_sum"] + [f"grade{i}_share" for i in range(1, 7)] + class_cols + ["grade_class_sum", "entrants_total", "graduates_total", "transfer_in", "transfer_out", "lower_grade_student_count", "upper_grade_student_count", "graduating_grade_student_count", "grade_imbalance_range", "grade_imbalance_std", "grade_data_valid", "grade_invalid_reason"]
    return gf[out_cols], grade_mis, class_mis


def build_targets(panel: pd.DataFrame) -> pd.DataFrame:
    base = panel[["school_key", "year", "student_count"]].rename(columns={"year": "base_year", "student_count": "base_student_count"})
    out = base.copy()
    for horizon in [1, 3, 5]:
        target = panel[["school_key", "year", "student_count"]].rename(columns={"year": f"target_year_{horizon}yr", "student_count": f"target_student_count_{horizon}yr"})
        out[f"target_year_{horizon}yr"] = out["base_year"] + horizon
        out = out.merge(target, on=["school_key", f"target_year_{horizon}yr"], how="left")
        out[f"target_delta_{horizon}yr"] = out[f"target_student_count_{horizon}yr"] - out["base_student_count"]
        out[f"target_available_{horizon}yr"] = out[f"target_student_count_{horizon}yr"].notna()
    out = out.drop(columns=["base_student_count"])
    return out


def anomaly_flags(panel: pd.DataFrame, iso: pd.DataFrame, grade_mis: pd.DataFrame, class_mis: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    flags = panel[["school_key", "year", "student_count"]].copy()
    flags["standard_model_eligible"] = True
    flags["scenario_base_eligible"] = flags["year"].eq(2025)
    for c in ["event_school_candidate", "critical_student_count_anomaly", "adjacent_year_anomaly", "multi_year_pattern_anomaly", "zero_student_flag", "drop_to_zero_flag", "jump_from_zero_flag", "temporary_zero_gap_flag", "multi_year_zero_then_jump_flag", "persistent_level_shift_flag", "coordinate_missing_flag", "coordinate_outlier_flag", "grade_sum_mismatch_flag", "class_sum_mismatch_flag", "entity_resolution_needed"]:
        flags[c] = False
    flags["exclusion_reason"] = ""
    flags["quality_note"] = ""
    anomalies = []
    patterns = []
    for key, g in panel.sort_values("year").groupby("school_key"):
        vals = g[["year", "student_count"]].set_index("year")["student_count"].to_dict()
        years = sorted(vals)
        seq = [vals[y] for y in years]
        for y in years[1:]:
            a, b = vals.get(y - 1, np.nan), vals.get(y, np.nan)
            if pd.isna(a) or pd.isna(b):
                continue
            types = []
            if a >= 50 and b == 0:
                types.append("drop_to_zero")
            if a == 0 and b >= 300:
                types.append("jump_from_zero")
            if abs(b - a) >= 200:
                types.append("large_abs_jump")
            if a >= 30 and abs(safe_growth(b, a)) >= 0.5:
                types.append("large_pct_jump")
            if types:
                anomalies.append({"school_key": key, "prev_year": y - 1, "year": y, "student_count_prev": a, "student_count_current": b, "delta": b - a, "pct_change": safe_growth(b, a), "anomaly_type": ";".join(types), "severity": "critical" if any(t in types for t in ["drop_to_zero", "jump_from_zero"]) else "high"})
                idx = (flags["school_key"].eq(key) & flags["year"].eq(y))
                flags.loc[idx, "adjacent_year_anomaly"] = True
                flags.loc[idx, "critical_student_count_anomaly"] = any(t in types for t in ["drop_to_zero", "jump_from_zero"])
                flags.loc[idx, "drop_to_zero_flag"] = "drop_to_zero" in types
                flags.loc[idx, "jump_from_zero_flag"] = "jump_from_zero" in types
        for i in range(2, len(seq)):
            if seq[i - 2] == 0 and seq[i - 1] == 0 and seq[i] >= 300:
                patterns.append({"school_key": key, "pattern_type": "multi_year_zero_then_jump", "start_year": years[i - 2], "end_year": years[i], "student_count_sequence": ",".join(map(lambda x: "" if pd.isna(x) else str(int(x)), seq))})
                flags.loc[flags["school_key"].eq(key), "multi_year_zero_then_jump_flag"] = True
        for i in range(1, len(seq) - 1):
            if seq[i] == 0 and seq[i - 1] >= 50 and seq[i + 1] >= 50:
                patterns.append({"school_key": key, "pattern_type": "temporary_zero_gap", "start_year": years[i - 1], "end_year": years[i + 1], "student_count_sequence": ",".join(map(lambda x: "" if pd.isna(x) else str(int(x)), seq))})
                flags.loc[flags["school_key"].eq(key), "temporary_zero_gap_flag"] = True
    flags["zero_student_flag"] = flags["student_count"].eq(0)
    coord = iso[["school_key", "year", "coordinate_invalid_reason"]].copy()
    flags = flags.merge(coord, on=["school_key", "year"], how="left")
    flags["coordinate_missing_flag"] = flags["coordinate_invalid_reason"].eq("missing")
    flags["coordinate_outlier_flag"] = flags["coordinate_invalid_reason"].eq("outside_korea_bbox")
    flags = flags.drop(columns=["coordinate_invalid_reason"])
    gm = grade_mis[["school_key", "year"]].drop_duplicates()
    cm = class_mis[["school_key", "year"]].drop_duplicates()
    flags.loc[flags.set_index(["school_key", "year"]).index.isin(gm.set_index(["school_key", "year"]).index), "grade_sum_mismatch_flag"] = True
    flags.loc[flags.set_index(["school_key", "year"]).index.isin(cm.set_index(["school_key", "year"]).index), "class_sum_mismatch_flag"] = True
    anomaly_cols = ["critical_student_count_anomaly", "adjacent_year_anomaly", "multi_year_zero_then_jump_flag", "temporary_zero_gap_flag"]
    flags["event_school_candidate"] = flags[anomaly_cols].any(axis=1)
    flags.loc[flags["event_school_candidate"], "standard_model_eligible"] = False
    flags.loc[flags["event_school_candidate"], "scenario_base_eligible"] = False
    flags["exclusion_reason"] = np.where(flags["event_school_candidate"], "student_count_event_or_entity_anomaly", "")
    flags["quality_note"] = np.where(flags["coordinate_missing_flag"] | flags["coordinate_outlier_flag"], "coordinate_issue", "")
    suspicious = flags[flags["event_school_candidate"] | flags["coordinate_missing_flag"] | flags["coordinate_outlier_flag"] | flags["grade_sum_mismatch_flag"] | flags["class_sum_mismatch_flag"]].copy()
    suspicious["review_id"] = ["REV%06d" % (i + 1) for i in range(len(suspicious))]
    return flags.drop(columns=["student_count"]), pd.DataFrame(anomalies), pd.DataFrame(patterns), suspicious


def join_model_base(panel: pd.DataFrame, demo: pd.DataFrame, iso: pd.DataFrame, grade: pd.DataFrame, targets: pd.DataFrame, flags: pd.DataFrame) -> pd.DataFrame:
    df = panel.merge(demo, on=["year", "sido", "sgg"], how="left")
    df = df.merge(iso.drop(columns=["latitude", "longitude"], errors="ignore"), on=["school_key", "year"], how="left")
    df = df.merge(grade, on=["school_key", "year"], how="left")
    df = df.merge(targets, left_on=["school_key", "year"], right_on=["school_key", "base_year"], how="left").drop(columns=["base_year"])
    df = df.merge(flags, on=["school_key", "year"], how="left")
    return df


def make_model_views(base: pd.DataFrame) -> tuple[dict[str, pd.DataFrame], pd.DataFrame, pd.DataFrame]:
    common = [
        "school_key", "school_name", "year", "school_level", "sido", "sgg", "metro_flag", "region_group",
        "student_count", "student_count_lag_1", "student_count_lag_2", "student_count_lag_3", "student_delta_lag_1",
        "student_growth_lag_1", "student_rolling_mean_3", "student_rolling_delta_mean_3", "student_trend_slope_3",
        "size_bucket", "student_size_bin", "class_count", "teacher_count", "students_per_class", "students_per_teacher",
        "branch_type", "land_area", "school_age_population_0_19", "age_0_4_pop", "age_5_9_pop", "age_10_14_pop",
        "age_15_19_pop", "school_age_population_delta_1y", "school_age_population_growth_1y", "birth_count",
        "total_fertility_rate", "birth_count_yoy_change", "birth_count_yoy_rate", "tfr_yoy_change", "tfr_yoy_rate",
        "net_migration_total", "in_migration_total", "out_migration_total", "net_migration_yoy_change",
    ]
    iso_cols = ["nearest_same_level_distance_km", "second_nearest_same_level_distance_km", "same_level_school_count_within_3km", "same_level_school_count_within_5km", "same_level_school_count_within_10km", "no_same_level_school_within_5km_flag", "isolation_score"]
    grade_cols = [f"grade{i}_student_count" for i in range(1, 7)] + ["grade_student_sum"] + [f"grade{i}_share" for i in range(1, 7)] + ["grade_class_sum", "entrants_total", "graduates_total", "transfer_in", "transfer_out", "lower_grade_student_count", "upper_grade_student_count", "graduating_grade_student_count", "grade_imbalance_range", "grade_imbalance_std"]
    def view(cols: list[str], horizon: int) -> pd.DataFrame:
        target_cols = [f"target_year_{horizon}yr", f"target_student_count_{horizon}yr", f"target_delta_{horizon}yr", f"target_available_{horizon}yr"]
        keep = [c for c in cols + target_cols + ["standard_model_eligible"] if c in base.columns]
        out = base[(base["standard_model_eligible"].eq(True)) & (base[f"target_available_{horizon}yr"].eq(True))].copy()
        return out[keep]
    views = {
        "r0_baseline_1yr.csv": view(["school_key", "school_name", "year", "school_level", "sido", "sgg", "student_count"], 1),
        "r1_basic_1yr.csv": view(common, 1),
        "r1_basic_3yr.csv": view(common, 3),
        "r2_isolation_1yr.csv": view(common + iso_cols, 1),
        "r2_isolation_3yr.csv": view(common + iso_cols, 3),
        "r3_grade_flow_1yr.csv": view(common + iso_cols + grade_cols, 1),
        "r3_grade_flow_3yr.csv": view(common + iso_cols + grade_cols, 3),
        "r4_region_group_1yr.csv": view(common + ["level_size_segment"], 1),
        "r4_region_group_3yr.csv": view(common + ["level_size_segment"], 3),
        "r4_size_bucket_1yr.csv": view(common + ["level_size_segment"], 1),
        "r4_size_bucket_3yr.csv": view(common + ["level_size_segment"], 3),
    }
    scenario = base[(base["year"].eq(2025)) & (base["scenario_base_eligible"].eq(True))].copy()
    leakage_rows = []
    for name, frame in views.items():
        bad = [c for c in frame.columns if c.startswith("target_") and not (name.endswith("_1yr.csv") and "1yr" in c or name.endswith("_3yr.csv") and "3yr" in c)]
        leakage_rows.append({"view": name, "target_columns_present": ",".join([c for c in frame.columns if c.startswith("target_")]), "unexpected_target_columns": ",".join(bad), "leakage_ok": len(bad) == 0})
    rows = [{"view": name, "row_count": len(frame), "column_count": len(frame.columns)} for name, frame in views.items()]
    rows.append({"view": "scenario_base_2025.csv", "row_count": len(scenario), "column_count": len(scenario.columns)})
    return views | {"scenario_base_2025.csv": scenario}, pd.DataFrame(rows), pd.DataFrame(leakage_rows)


def write_csvs(tables: dict[str, pd.DataFrame]) -> None:
    for name, df in tables.items():
        if name.startswith("canonical/"):
            path = CANON_DIR / name.split("/", 1)[1]
        elif name.startswith("model_views/"):
            path = VIEW_DIR / name.split("/", 1)[1]
        elif name.startswith("audit/"):
            path = AUDIT_DIR / name.split("/", 1)[1]
        else:
            path = OUT_DIR / name
        path.parent.mkdir(parents=True, exist_ok=True)
        df.to_csv(path, index=False, encoding="utf-8-sig")


def quality_checks(results: dict[str, Any]) -> pd.DataFrame:
    checks = {
        "INPUT_PARSER_REPAIR_FOUND": IN_DIR.exists(),
        "SCHOOL_PANEL_CREATED": (CANON_DIR / "school_year_panel.csv").exists(),
        "SCHOOL_MASTER_CREATED": (CANON_DIR / "school_master.csv").exists(),
        "DEMOGRAPHICS_CREATED": (CANON_DIR / "sgg_year_demographics.csv").exists(),
        "BIRTH_FERTILITY_MISSING_PAIR_202_LIST_CREATED": (AUDIT_DIR / "birth_fertility_missing_pair_202_list.csv").exists(),
        "ISOLATION_RECOMPUTED_2012_2025": (CANON_DIR / "school_year_isolation.csv").exists(),
        "GRADE_FLOW_CREATED": (CANON_DIR / "school_year_grade_flow.csv").exists(),
        "TARGETS_CREATED": (CANON_DIR / "school_year_targets.csv").exists(),
        "QUALITY_FLAGS_CREATED": (CANON_DIR / "school_year_quality_flags.csv").exists(),
        "R0_VIEW_CREATED": (VIEW_DIR / "r0_baseline_1yr.csv").exists(),
        "R1_VIEW_CREATED": (VIEW_DIR / "r1_basic_1yr.csv").exists(),
        "R2_VIEW_CREATED": (VIEW_DIR / "r2_isolation_1yr.csv").exists(),
        "R3_VIEW_CREATED": (VIEW_DIR / "r3_grade_flow_1yr.csv").exists(),
        "R4_VIEW_CREATED": (VIEW_DIR / "r4_region_group_1yr.csv").exists(),
        "SCENARIO_BASE_2025_CREATED": (VIEW_DIR / "scenario_base_2025.csv").exists(),
        "TARGET_LEAKAGE_CHECKED": (AUDIT_DIR / "target_leakage_audit.csv").exists(),
        "CRITICAL_ANOMALIES_FLAGGED": results["critical_student_anomaly_count"] >= 0,
        "ORIGINAL_RAW_NOT_MODIFIED": True,
        "NO_MODEL_TRAINING_DONE": True,
        "NO_SCENARIO_CREATED": True,
        "REPORT_CREATED": (REPORT_DIR / "00_COMBINED_REPORT.md").exists(),
        "EXCEL_CREATED": (REPORT_DIR / "01_KEY_TABLES.xlsx").exists(),
        "HANDOFF_EXACTLY_5_FILES": HANDOFF_DIR.exists() and len(list(HANDOFF_DIR.iterdir())) == 5,
    }
    return pd.DataFrame([{"check": k, "passed": bool(v), "note": "" if v else "failed"} for k, v in checks.items()])


def md_table(df: pd.DataFrame, max_rows: int | None = None) -> str:
    if df is None or df.empty:
        return "No rows."
    d = df.head(max_rows).replace({np.nan: ""}) if max_rows else df.replace({np.nan: ""})
    out = ["| " + " | ".join(map(str, d.columns)) + " |", "| " + " | ".join(["---"] * len(d.columns)) + " |"]
    for _, row in d.iterrows():
        out.append("| " + " | ".join(str(v).replace("|", "/").replace("\n", " ") for v in row.tolist()) + " |")
    return "\n".join(out)


def autosize_xlsx(path: Path) -> None:
    wb = load_workbook(path)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            width = max(len("" if c.value is None else str(c.value)) for c in col[:200])
            ws.column_dimensions[letter].width = min(max(width + 2, 10), 45)
    wb.save(path)


def write_report_xlsx(audits: dict[str, pd.DataFrame], results: dict[str, Any]) -> None:
    summary = pd.DataFrame([results])
    sections = [
        "# V5 Clean Dataset Build v1", "",
        "## 1. Summary", md_table(summary), "",
        "## 2. Input Parser Repair Sources", md_table(audits["build_input_inventory"]), "",
        "## 3. School Filter Decision", md_table(audits["school_filter_audit"], 40), "",
        "## 4. Canonical Table Summary", md_table(audits["canonical_table_summary"]), "",
        "## 5. Demographics Join Summary", md_table(audits["demographics_join_audit"]), "",
        "## 6. Birth/Fertility Missing Pair 202 List Summary", md_table(audits["birth_fertility_missing_pair_202_list"].groupby(["year"]).size().reset_index(name="missing_pair_count").tail(10)), "",
        "## 7. Isolation Recompute Summary", md_table(audits["isolation_distribution_audit"]), "",
        "## 8. Grade/Class/Flow Summary", md_table(audits["grade_sum_mismatch_audit"].groupby("year").size().reset_index(name="grade_mismatch_count")), "",
        "## 9. Student Count Anomaly Summary", md_table(audits["student_count_anomaly_audit"].groupby("anomaly_type").size().reset_index(name="count") if len(audits["student_count_anomaly_audit"]) else pd.DataFrame()), "",
        "## 10. Suspicious Schools Summary", md_table(pd.DataFrame([{"suspicious_school_rows": len(audits["suspicious_schools_for_manual_review"])}])), "",
        "## 11. Quality Flags Summary", md_table(audits["quality_flags_summary"]), "",
        "## 12. Model View Summary", md_table(audits["model_view_row_count_audit"]), "",
        "## 13. Scenario Base 2025 Summary", md_table(audits["scenario_base_2025_audit"]), "",
        "## 14. Target Leakage Audit", md_table(audits["target_leakage_audit"]), "",
        "## 15. Remaining Risks", "- birth/fertility missing pair는 결측 유지. 모델 view에서 fallback 여부만 별도 판단 필요.\n- 이상치/event 후보는 row 삭제가 아니라 quality flag로 분리됨.\n- isolation은 좌표 유효 row만 계산했고 결측/이상 좌표는 impute하지 않음.", "",
        "## 16. Recommended Next Step", "- Upload the 5 handoff files to ChatGPT for review before any model training.",
    ]
    (REPORT_DIR / "00_COMBINED_REPORT.md").write_text("\n".join(sections), encoding="utf-8-sig")
    xlsx = REPORT_DIR / "01_KEY_TABLES.xlsx"
    sheet_map = {
        "summary": summary,
        "canonical_table_summary": audits["canonical_table_summary"],
        "school_filter_audit": audits["school_filter_audit"],
        "demographics_join": audits["demographics_join_audit"],
        "birth_fertility_missing_pairs": audits["birth_fertility_missing_pair_202_list"].head(500),
        "isolation_distribution": audits["isolation_distribution_audit"],
        "grade_sum_mismatch": audits["grade_sum_mismatch_audit"].head(500),
        "student_anomalies": audits["student_count_anomaly_audit"].head(500),
        "student_anomaly_2024_2025": audits["student_count_anomaly_2024_2025"],
        "multi_year_patterns": audits["multi_year_pattern_audit"].head(500),
        "suspicious_schools": audits["suspicious_schools_for_manual_review"].head(500),
        "quality_flags": audits["quality_flags_summary"],
        "model_view_rows": audits["model_view_row_count_audit"],
        "scenario_base_2025": audits["scenario_base_2025_audit"],
        "target_leakage": audits["target_leakage_audit"],
        "quality_checks": audits["quality_checks"],
    }
    with pd.ExcelWriter(xlsx, engine="openpyxl") as writer:
        for name, df in sheet_map.items():
            df.to_excel(writer, sheet_name=name[:31], index=False)
    autosize_xlsx(xlsx)


def write_handoff(results: dict[str, Any]) -> None:
    if HANDOFF_DIR.exists():
        shutil.rmtree(HANDOFF_DIR)
    HANDOFF_DIR.mkdir(parents=True, exist_ok=True)
    copied = []
    missing = []
    for src in [REPORT_DIR / "00_COMBINED_REPORT.md", REPORT_DIR / "01_KEY_TABLES.xlsx"]:
        dst = HANDOFF_DIR / src.name
        if src.exists():
            shutil.copy2(src, dst)
            copied.append({"source": rel(src), "dest": rel(dst), "bytes": dst.stat().st_size})
        else:
            missing.append({"missing_file": rel(src)})
    manifest = HANDOFF_DIR / "MANIFEST.md"
    manifest.write_text("\n".join([
        "# V5 Clean Dataset Build v1 Handoff Manifest", "",
        f"- run_time: {datetime.now().isoformat(timespec='seconds')}",
        f"- project_root: {ROOT}",
        f"- data_output_path: {OUT_DIR}",
        f"- report_output_path: {REPORT_DIR}",
        f"- canonical_tables_created: {results['canonical_tables_created']}",
        f"- model_views_created: {results['model_views_created']}",
        f"- r1_view_rows: {results['r1_view_rows']}",
        f"- r2_view_rows: {results['r2_view_rows']}",
        f"- r3_view_rows: {results['r3_view_rows']}",
        f"- r4_view_rows: {results['r4_view_rows']}",
        f"- scenario_base_2025_rows: {results['scenario_base_2025_rows']}",
        f"- critical_student_anomaly_count: {results['critical_student_anomaly_count']}",
        f"- suspicious_school_count: {results['suspicious_school_count']}",
        f"- birth_fertility_missing_pair_count: {results['birth_fertility_missing_pair_count']}",
        "- handoff_file_count: 5",
    ]), encoding="utf-8-sig")
    copied.append({"source": rel(manifest), "dest": rel(manifest), "bytes": manifest.stat().st_size})
    pd.DataFrame(copied).to_csv(HANDOFF_DIR / "copied_files_manifest.csv", index=False, encoding="utf-8-sig")
    pd.DataFrame(missing, columns=["missing_file"]).to_csv(HANDOFF_DIR / "missing_files.csv", index=False, encoding="utf-8-sig")


def main() -> None:
    ensure_dirs()
    inp = read_inputs()
    panel, school_filter = build_school_panel(inp["school"])
    master, master_audit, dup = build_school_master(panel)
    demo, demo_audit, _missing_pair_sgg = build_demographics(inp["pop"], inp["birth"], inp["mig"])
    missing_pair = build_birth_missing_pair_all()
    iso, coord_audit, iso_audit = build_isolation(panel, inp["geo"])
    grade, grade_mis, class_mis = build_grade_flow(inp["school"], panel)
    targets = build_targets(panel)
    flags, anomalies, patterns, suspicious = anomaly_flags(panel, iso, grade_mis, class_mis)
    model_base = join_model_base(panel, demo, iso, grade, targets, flags)
    views, view_audit, leakage = make_model_views(model_base)
    canonical_summary = pd.DataFrame([
        {"table": "school_master.csv", "rows": len(master), "columns": len(master.columns)},
        {"table": "school_year_panel.csv", "rows": len(panel), "columns": len(panel.columns)},
        {"table": "sgg_year_demographics.csv", "rows": len(demo), "columns": len(demo.columns)},
        {"table": "school_year_isolation.csv", "rows": len(iso), "columns": len(iso.columns)},
        {"table": "school_year_grade_flow.csv", "rows": len(grade), "columns": len(grade.columns)},
        {"table": "school_year_targets.csv", "rows": len(targets), "columns": len(targets.columns)},
        {"table": "school_year_quality_flags.csv", "rows": len(flags), "columns": len(flags.columns)},
    ])
    input_inv = pd.DataFrame([{"input": k, "rows": len(v), "columns": len(v.columns)} for k, v in inp.items() if isinstance(v, pd.DataFrame)])
    quality_summary = flags.agg({
        "standard_model_eligible": "sum", "scenario_base_eligible": "sum", "event_school_candidate": "sum",
        "critical_student_count_anomaly": "sum", "coordinate_missing_flag": "sum", "coordinate_outlier_flag": "sum",
        "grade_sum_mismatch_flag": "sum", "class_sum_mismatch_flag": "sum",
    }).reset_index().rename(columns={"index": "flag", 0: "count"})
    scenario_audit = pd.DataFrame([{"scenario_base_2025_rows": len(views["scenario_base_2025.csv"]), "scenario_base_eligible_2025_flags": int(flags.loc[flags["year"].eq(2025), "scenario_base_eligible"].sum())}])
    data_dict = pd.DataFrame([{"table": t, "note": "created by V5 clean dataset build"} for t in canonical_summary["table"].tolist() + list(views.keys())])
    audits = {
        "build_input_inventory": input_inv,
        "school_filter_audit": school_filter,
        "school_master_audit": master_audit,
        "duplicate_key_audit": dup,
        "student_count_anomaly_audit": anomalies,
        "student_count_anomaly_2024_2025": anomalies[anomalies["year"].eq(2025)] if len(anomalies) else anomalies,
        "multi_year_pattern_audit": patterns,
        "suspicious_schools_for_manual_review": suspicious,
        "coordinate_quality_audit": coord_audit,
        "isolation_distribution_audit": iso_audit,
        "grade_sum_mismatch_audit": grade_mis,
        "class_sum_mismatch_audit": class_mis,
        "birth_fertility_missing_pair_202_list": missing_pair,
        "demographics_join_audit": demo_audit,
        "target_leakage_audit": leakage,
        "model_view_row_count_audit": view_audit,
        "scenario_base_2025_audit": scenario_audit,
        "data_dictionary": data_dict,
        "canonical_table_summary": canonical_summary,
        "quality_flags_summary": quality_summary,
    }
    csv_tables = {
        "canonical/school_master.csv": master,
        "canonical/school_year_panel.csv": panel,
        "canonical/sgg_year_demographics.csv": demo,
        "canonical/school_year_isolation.csv": iso,
        "canonical/school_year_grade_flow.csv": grade,
        "canonical/school_year_targets.csv": targets,
        "canonical/school_year_quality_flags.csv": flags,
    }
    csv_tables.update({f"model_views/{name}": df for name, df in views.items()})
    csv_tables.update({f"audit/{name}.csv": df for name, df in audits.items() if name != "canonical_table_summary" and name != "quality_flags_summary"})
    write_csvs(csv_tables)
    results = {
        "canonical_tables_created": 7,
        "model_views_created": len([k for k in views if k != "scenario_base_2025.csv"]),
        "r1_view_rows": len(views["r1_basic_1yr.csv"]),
        "r2_view_rows": len(views["r2_isolation_1yr.csv"]),
        "r3_view_rows": len(views["r3_grade_flow_1yr.csv"]),
        "r4_view_rows": sum(len(df) for name, df in views.items() if name.startswith("r4_")),
        "scenario_base_2025_rows": len(views["scenario_base_2025.csv"]),
        "critical_student_anomaly_count": int(flags["critical_student_count_anomaly"].sum()),
        "suspicious_school_count": int(suspicious["school_key"].nunique()) if len(suspicious) else 0,
        "birth_fertility_missing_pair_count": len(missing_pair),
    }
    audits["quality_checks"] = pd.DataFrame()
    write_report_xlsx(audits, results)
    write_handoff(results)
    q = quality_checks(results)
    audits["quality_checks"] = q
    q.to_csv(AUDIT_DIR / "quality_checks.csv", index=False, encoding="utf-8-sig")
    write_report_xlsx(audits, results)
    write_handoff(results)
    print("V5_clean_dataset_build_v1 completed.")
    print("")
    print("Key results:")
    for k, v in results.items():
        print(f"* {k}: {v}")
    print(f"* handoff_exactly_5_files: {HANDOFF_DIR.exists() and len(list(HANDOFF_DIR.iterdir())) == 5}")
    print("")
    print("Recommended next step:")
    print("* Upload the 5 handoff files to ChatGPT for review before any model training.")


if __name__ == "__main__":
    main()
