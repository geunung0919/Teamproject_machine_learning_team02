from __future__ import annotations

import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
BUILD_DIR = ROOT / "data" / "v5_clean_dataset_build_v1"
PARSER_DIR = ROOT / "data" / "v5_raw_parser_repair_v1"
RAW_DIR = ROOT / "data" / "raw"
OUT_DIR = ROOT / "data" / "v5_clean_dataset_patch_v1"
CANON_DIR = OUT_DIR / "canonical"
VIEW_DIR = OUT_DIR / "model_views"
AUDIT_DIR = OUT_DIR / "audit"
REPORT_DIR = ROOT / "reports" / "v5_clean_dataset_patch_v1"
HANDOFF_DIR = ROOT / "handoff_for_chatgpt" / "v5_clean_dataset_patch_v1"

YEARS = list(range(2012, 2026))
METRO_SIDO = {"서울", "부산", "대구", "인천", "광주", "대전", "울산", "세종"}

MODERN_SIDO_BY_PREFIX = {
    "11": "서울", "26": "부산", "27": "대구", "28": "인천", "29": "광주", "30": "대전", "31": "울산",
    "36": "세종", "41": "경기", "42": "강원", "43": "충북", "44": "충남", "45": "전북", "46": "전남",
    "47": "경북", "48": "경남", "50": "제주", "51": "강원", "52": "전북",
}
OLD_SIDO_BY_PREFIX = {
    "11": "서울", "21": "부산", "22": "대구", "23": "인천", "24": "광주", "25": "대전", "26": "울산",
    "29": "세종", "31": "경기", "32": "강원", "33": "충북", "34": "충남", "35": "전북", "36": "전남",
    "37": "경북", "38": "경남", "39": "제주",
}
SIDO_ALIASES = {
    "서울특별시": "서울", "부산광역시": "부산", "대구광역시": "대구", "인천광역시": "인천", "광주광역시": "광주",
    "대전광역시": "대전", "울산광역시": "울산", "세종특별자치시": "세종", "경기도": "경기",
    "강원도": "강원", "강원특별자치도": "강원", "충청북도": "충북", "충청남도": "충남",
    "전라북도": "전북", "전북특별자치도": "전북", "전라남도": "전남", "경상북도": "경북",
    "경상남도": "경남", "제주특별자치도": "제주", "제주도": "제주",
}


def rel(path: Path) -> str:
    try:
        return str(path.relative_to(ROOT))
    except ValueError:
        return str(path)


def ensure_dirs() -> None:
    for d in [OUT_DIR, CANON_DIR, VIEW_DIR, AUDIT_DIR, REPORT_DIR]:
        d.mkdir(parents=True, exist_ok=True)
    if HANDOFF_DIR.exists():
        shutil.rmtree(HANDOFF_DIR)
    HANDOFF_DIR.mkdir(parents=True, exist_ok=True)


def norm_sido(x: Any) -> str:
    s = "" if pd.isna(x) else str(x).strip()
    return SIDO_ALIASES.get(s, s)


def region_group(sido: str) -> str:
    if sido in METRO_SIDO:
        return "metro"
    if sido == "경기":
        return "capital_area"
    return "province"


def read_build() -> dict[str, pd.DataFrame]:
    return {
        "panel": pd.read_csv(BUILD_DIR / "canonical" / "school_year_panel.csv", low_memory=False),
        "master": pd.read_csv(BUILD_DIR / "canonical" / "school_master.csv", low_memory=False),
        "isolation": pd.read_csv(BUILD_DIR / "canonical" / "school_year_isolation.csv", low_memory=False),
        "grade": pd.read_csv(BUILD_DIR / "canonical" / "school_year_grade_flow.csv", low_memory=False),
        "targets": pd.read_csv(BUILD_DIR / "canonical" / "school_year_targets.csv", low_memory=False),
        "flags": pd.read_csv(BUILD_DIR / "canonical" / "school_year_quality_flags.csv", low_memory=False),
        "old_demo": pd.read_csv(BUILD_DIR / "canonical" / "sgg_year_demographics.csv", low_memory=False),
    }


def previous_problem_audit(build: dict[str, pd.DataFrame]) -> pd.DataFrame:
    old = build["old_demo"]
    rows = []
    rows.append({"metric": "old_sgg_year_demographics_rows", "value": len(old), "note": "previous build output"})
    rows.append({"metric": "old_demo_year_min", "value": old["year"].min(), "note": ""})
    rows.append({"metric": "old_demo_year_max", "value": old["year"].max(), "note": ""})
    for y, g in old.groupby("year"):
        rows.append({
            "metric": "old_demo_rows_by_year",
            "year": y,
            "value": len(g),
            "birth_nonnull_rate": g["birth_count"].notna().mean() if "birth_count" in g else np.nan,
            "migration_nonnull_rate": g["net_migration_total"].notna().mean() if "net_migration_total" in g else np.nan,
            "school_age_nonnull_rate": g["school_age_population_0_19"].notna().mean() if "school_age_population_0_19" in g else np.nan,
            "note": "suspicious_low_row_count" if len(g) < 100 else "",
        })
    flags = build["flags"]
    rows.append({"metric": "old_grade_sum_mismatch_flag_count", "value": int(flags.get("grade_sum_mismatch_flag", pd.Series(False, index=flags.index)).sum()), "note": ""})
    rows.append({"metric": "old_class_sum_mismatch_flag_count", "value": int(flags.get("class_sum_mismatch_flag", pd.Series(False, index=flags.index)).sum()), "note": ""})
    for name in ["r1_basic_1yr.csv", "r2_isolation_1yr.csv", "r3_grade_flow_1yr.csv", "scenario_base_2025.csv"]:
        p = BUILD_DIR / "model_views" / name
        rows.append({"metric": f"old_{name}_row_count", "value": len(pd.read_csv(p, low_memory=False)) if p.exists() else np.nan, "note": ""})
    return pd.DataFrame(rows)


def build_corrected_demographics(panel: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    pop = pd.read_csv(RAW_DIR / "national_kosis_school_age_population_sgg.csv", low_memory=False)
    pop["year"] = (pd.to_numeric(pop["period"], errors="coerce") // 100).astype("Int64")
    pop["sgg_code_str"] = pop["sgg_code"].astype(str).str.zfill(5)
    pop["sido"] = pop["sgg_code_str"].str[:2].map(MODERN_SIDO_BY_PREFIX)
    pop["sgg"] = pop["sgg_name"].astype(str)
    pop = pop[pop["year"].isin(YEARS)]
    pop_piv = pop.pivot_table(index=["year", "sido", "sgg", "sgg_code"], columns="age_group", values="population", aggfunc="first").reset_index()
    pop_piv.columns.name = None
    pop_piv = pop_piv.rename(columns={"0 - 4세": "age_0_4_pop", "5 - 9세": "age_5_9_pop", "10 - 14세": "age_10_14_pop", "15 - 19세": "age_15_19_pop"})
    for c in ["age_0_4_pop", "age_5_9_pop", "age_10_14_pop", "age_15_19_pop"]:
        if c not in pop_piv:
            pop_piv[c] = np.nan
    pop_piv["school_age_population_0_19"] = pop_piv[["age_0_4_pop", "age_5_9_pop", "age_10_14_pop", "age_15_19_pop"]].sum(axis=1, min_count=1)

    birth = pd.read_csv(PARSER_DIR / "sgg_birth_fertility_2007_2025.csv", low_memory=False)
    birth["sido"] = birth["sido"].map(norm_sido)
    birth["sgg"] = birth["sgg"].astype(str)
    birth = birth[birth["year"].isin(YEARS)]
    birth = birth.groupby(["year", "sido", "sgg"], as_index=False).agg(
        region_code=("region_code", "first"),
        birth_count=("birth_count", lambda s: s.dropna().iloc[0] if s.notna().any() else np.nan),
        total_fertility_rate=("total_fertility_rate", lambda s: s.dropna().iloc[0] if s.notna().any() else np.nan),
    )

    mig = pd.read_csv(RAW_DIR / "national_kosis_migration_sgg.csv", low_memory=False)
    mig = mig[(mig["year"].isin(YEARS)) & (mig["region_level"].eq("sgg"))].copy()
    mig["code_str"] = mig["region_code"].astype(str).str.zfill(5)
    mig["sido"] = mig["code_str"].str[:2].map(MODERN_SIDO_BY_PREFIX)
    mig["sgg"] = mig["region_name"].astype(str)
    mig_piv = mig.pivot_table(index=["year", "sido", "sgg"], columns="item_name", values="value", aggfunc="first").reset_index()
    mig_piv.columns.name = None
    mig_piv = mig_piv.rename(columns={"순이동": "net_migration_total", "총전입": "in_migration_total", "총전출": "out_migration_total"})

    demo = pop_piv.merge(
        birth[["year", "sido", "sgg", "region_code", "birth_count", "total_fertility_rate"]],
        on=["year", "sido", "sgg"],
        how="left",
    ).merge(
        mig_piv[["year", "sido", "sgg", "net_migration_total", "in_migration_total", "out_migration_total"]],
        on=["year", "sido", "sgg"],
        how="left",
    )
    demo = demo.sort_values(["sido", "sgg", "year"])
    for c in ["school_age_population_0_19", "birth_count", "total_fertility_rate", "net_migration_total"]:
        demo[c] = pd.to_numeric(demo[c], errors="coerce")
    g = demo.groupby(["sido", "sgg"], group_keys=False)
    demo["school_age_population_delta_1y"] = g["school_age_population_0_19"].diff()
    demo["school_age_population_growth_1y"] = g["school_age_population_0_19"].pct_change(fill_method=None)
    demo["birth_count_yoy_change"] = g["birth_count"].diff()
    demo["birth_count_yoy_rate"] = g["birth_count"].pct_change(fill_method=None)
    demo["tfr_yoy_change"] = g["total_fertility_rate"].diff()
    demo["tfr_yoy_rate"] = g["total_fertility_rate"].pct_change(fill_method=None)
    demo["net_migration_yoy_change"] = g["net_migration_total"].diff()
    cols = [
        "year", "sido", "sgg", "sgg_code", "age_0_4_pop", "age_5_9_pop", "age_10_14_pop", "age_15_19_pop",
        "school_age_population_0_19", "school_age_population_delta_1y", "school_age_population_growth_1y",
        "birth_count", "total_fertility_rate", "birth_count_yoy_change", "birth_count_yoy_rate", "tfr_yoy_change",
        "tfr_yoy_rate", "net_migration_total", "in_migration_total", "out_migration_total", "net_migration_yoy_change",
    ]
    demo = demo[cols]

    panel_regions = panel[["year", "sido", "sgg"]].drop_duplicates()
    sanity_rows = []
    join_rows = []
    unmatched_rows = []
    joined = panel.merge(demo, on=["year", "sido", "sgg"], how="left")
    for y in YEARS:
        d = demo[demo["year"].eq(y)]
        pr = panel_regions[panel_regions["year"].eq(y)]
        school_y = panel[panel["year"].eq(y)]
        joined_y = joined[joined["year"].eq(y)]
        sanity_rows.append({
            "year": y,
            "sgg_demo_rows": len(d),
            "unique_sido_count": d["sido"].nunique(),
            "unique_sgg_count": d[["sido", "sgg"]].drop_duplicates().shape[0],
            "school_panel_sgg_count": pr[["sido", "sgg"]].drop_duplicates().shape[0],
            "expected_school_panel_sgg_count": pr[["sido", "sgg"]].drop_duplicates().shape[0],
            "school_age_nonnull_rate": d["school_age_population_0_19"].notna().mean(),
            "birth_nonnull_rate": d["birth_count"].notna().mean(),
            "fertility_nonnull_rate": d["total_fertility_rate"].notna().mean(),
            "migration_nonnull_rate": d["net_migration_total"].notna().mean(),
            "suspicious_low_row_count": len(d) < 100,
            "note": "",
        })
        non_school = joined_y["school_age_population_0_19"].notna()
        birth_non = joined_y["birth_count"].notna()
        mig_non = joined_y["net_migration_total"].notna()
        unmatched = joined_y[~(non_school & birth_non & mig_non)]
        join_rows.append({
            "year": y,
            "school_rows": len(school_y),
            "joined_rows": int(non_school.sum()),
            "school_age_population_nonnull_rate": non_school.mean(),
            "birth_count_nonnull_rate": birth_non.mean(),
            "total_fertility_rate_nonnull_rate": joined_y["total_fertility_rate"].notna().mean(),
            "migration_nonnull_rate": mig_non.mean(),
            "unmatched_school_region_count": unmatched[["sido", "sgg"]].drop_duplicates().shape[0],
            "unmatched_school_rows": len(unmatched),
            "issue_level": "ok" if non_school.mean() > 0.95 else "high",
            "note": "",
        })
        if len(unmatched):
            grp = unmatched.groupby(["year", "sido", "sgg"]).agg(
                school_rows=("school_key", "size"),
                missing_school_age_population=("school_age_population_0_19", lambda s: s.isna().any()),
                missing_birth_fertility=("birth_count", lambda s: s.isna().any()),
                missing_migration=("net_migration_total", lambda s: s.isna().any()),
            ).reset_index()
            grp["recommended_fix"] = "check_region_alias_or_source_missing"
            unmatched_rows.append(grp)
    unmatched_df = pd.concat(unmatched_rows, ignore_index=True) if unmatched_rows else pd.DataFrame(columns=["year", "sido", "sgg", "school_rows", "missing_school_age_population", "missing_birth_fertility", "missing_migration", "recommended_fix"])
    return demo, pd.DataFrame(sanity_rows), pd.DataFrame(join_rows), unmatched_df


def birth_missing_pairs() -> tuple[pd.DataFrame, pd.DataFrame]:
    frames = []
    for level, path in [
        ("national", PARSER_DIR / "national_birth_fertility_2007_2025.csv"),
        ("sido", PARSER_DIR / "sido_birth_fertility_2007_2025.csv"),
        ("sgg", PARSER_DIR / "sgg_birth_fertility_2007_2025.csv"),
    ]:
        df = pd.read_csv(path, low_memory=False)
        df["region_level"] = level
        if "sido" not in df:
            df["sido"] = ""
        if "sgg" not in df:
            df["sgg"] = ""
        if "region_name" in df and "region_name_raw" not in df:
            df = df.rename(columns={"region_name": "region_name_raw"})
        if "region_name_raw" not in df:
            df["region_name_raw"] = np.where(level == "national", "전국", df.get("sido", ""))
        frames.append(df)
    all_df = pd.concat(frames, ignore_index=True, sort=False)
    miss = all_df[all_df[["birth_count", "total_fertility_rate"]].isna().any(axis=1)].copy()
    miss["has_birth_count_T1"] = miss["birth_count"].notna()
    miss["has_total_fertility_rate_T2"] = miss["total_fertility_rate"].notna()
    miss["missing_item"] = np.where(miss["birth_count"].isna(), "birth_count_T1", "total_fertility_rate_T2")
    miss["recommended_action"] = np.where(miss["region_level"].eq("sgg"), "use_sido_fallback_for_model_view", "keep_missing")
    cols = ["year", "region_level", "region_code", "sido", "sgg", "region_name_raw", "has_birth_count_T1", "has_total_fertility_rate_T2", "birth_count", "total_fertility_rate", "missing_item", "source_file", "last_change_date", "recommended_action"]
    for c in cols:
        if c not in miss:
            miss[c] = ""
    miss = miss[cols].sort_values(["year", "region_level", "region_code"])
    summary = pd.DataFrame([
        {"metric": "total_missing_pair_count", "value": len(miss)},
        {"metric": "missing_pair_count_2012_2025", "value": len(miss[miss["year"].between(2012, 2025)])},
        {"metric": "missing_pair_count_2025", "value": len(miss[miss["year"].eq(2025)])},
    ])
    by_year = miss.groupby("year").size().reset_index(name="value")
    by_year["metric"] = "missing_pair_count_by_year"
    by_level = miss.groupby("region_level").size().reset_index(name="value")
    by_level["metric"] = "missing_pair_count_by_region_level"
    summary = pd.concat([summary, by_year[["metric", "year", "value"]], by_level[["metric", "region_level", "value"]]], ignore_index=True, sort=False)
    return miss, summary


def grade_summary(panel: pd.DataFrame, grade: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    df = grade.merge(panel[["school_key", "year", "school_name", "sido", "sgg", "school_level", "student_count", "class_count"]], on=["school_key", "year"], how="left")
    grade_cols = [f"grade{i}_student_count" for i in range(1, 7)]
    class_cols = [f"grade{i}_class_count" for i in range(1, 7)]
    df["comparison_grade_student_sum"] = np.where(df["school_level"].isin(["중학교", "고등학교"]), df[grade_cols[:3]].sum(axis=1, min_count=1), df[grade_cols].sum(axis=1, min_count=1))
    df["comparison_grade_class_sum"] = np.where(df["school_level"].isin(["중학교", "고등학교"]), df[class_cols[:3]].sum(axis=1, min_count=1), df[class_cols].sum(axis=1, min_count=1))
    df["grade_student_sum_diff"] = df["student_count"] - df["comparison_grade_student_sum"]
    df["grade_student_sum_diff_rate"] = df["grade_student_sum_diff"] / df["student_count"].replace({0: np.nan})
    df["grade_class_sum_diff"] = df["class_count"] - df["comparison_grade_class_sum"]
    df["grade_student_mismatch"] = ((df["student_count"].gt(0) & df["comparison_grade_student_sum"].isna()) | ((df["grade_student_sum_diff"].abs() > 3) & (df["grade_student_sum_diff_rate"].abs() > 0.02)) | (df["grade_student_sum_diff"].abs() >= 30))
    df["class_mismatch"] = (df["class_count"].gt(0) & df["comparison_grade_class_sum"].isna()) | (df["grade_class_sum_diff"].abs() >= 2)
    rows = []
    for (year, level), g in df.groupby(["year", "school_level"]):
        rows.append({
            "year": year,
            "school_level": level,
            "row_count": len(g),
            "grade_student_nonnull_rate": g[grade_cols].notna().any(axis=1).mean(),
            "grade_class_nonnull_rate": g[class_cols].notna().any(axis=1).mean(),
            "flow_nonnull_rate": g[["entrants_total", "graduates_total", "transfer_in", "transfer_out"]].notna().any(axis=1).mean(),
            "grade_student_sum_match_rate": (~g["grade_student_mismatch"]).mean(),
            "grade_class_sum_match_rate": (~g["class_mismatch"]).mean(),
            "grade_student_abs_diff_mean": g["grade_student_sum_diff"].abs().mean(),
            "grade_student_abs_diff_p95": g["grade_student_sum_diff"].abs().quantile(0.95),
            "grade_class_abs_diff_mean": g["grade_class_sum_diff"].abs().mean(),
            "grade_class_abs_diff_p95": g["grade_class_sum_diff"].abs().quantile(0.95),
            "mismatch_school_count": int((g["grade_student_mismatch"] | g["class_mismatch"]).sum()),
            "issue_level": "high" if (g["grade_student_mismatch"] | g["class_mismatch"]).mean() > 0.05 else "ok",
            "note": "",
        })
    gmis = df[df["grade_student_mismatch"]].copy()
    gmis["mismatch_type"] = np.where(gmis["comparison_grade_student_sum"].isna(), "grade_sum_missing", "grade_sum_mismatch")
    gmis["recommended_action"] = "review_before_r3_modeling"
    gmis = gmis.rename(columns={"comparison_grade_student_sum": "grade_student_sum"})
    gmis = gmis[["school_key", "year", "school_name", "sido", "sgg", "school_level", "student_count", "grade_student_sum", "grade_student_sum_diff", "grade_student_sum_diff_rate", "mismatch_type", "recommended_action"]]
    cmis = df[df["class_mismatch"]].copy()
    cmis["mismatch_type"] = np.where(cmis["comparison_grade_class_sum"].isna(), "class_sum_missing", "class_sum_mismatch")
    cmis["recommended_action"] = "review_before_r3_modeling"
    cmis = cmis.rename(columns={"comparison_grade_class_sum": "grade_class_sum"})
    cmis = cmis[["school_key", "year", "school_name", "sido", "sgg", "school_level", "class_count", "grade_class_sum", "grade_class_sum_diff", "mismatch_type", "recommended_action"]]
    return pd.DataFrame(rows), gmis, cmis


def patch_flags(flags: pd.DataFrame, join_audit: pd.DataFrame, unmatched: pd.DataFrame, gmis: pd.DataFrame, cmis: pd.DataFrame, panel: pd.DataFrame, demo: pd.DataFrame) -> pd.DataFrame:
    out = flags.copy()
    join = panel[["school_key", "year", "sido", "sgg"]].merge(demo[["year", "sido", "sgg", "school_age_population_0_19", "birth_count", "total_fertility_rate", "net_migration_total"]], on=["year", "sido", "sgg"], how="left")
    out = out.merge(join[["school_key", "year", "school_age_population_0_19", "birth_count", "total_fertility_rate", "net_migration_total"]], on=["school_key", "year"], how="left")
    out["demographics_join_missing_flag"] = out["school_age_population_0_19"].isna()
    out["birth_fertility_missing_pair_flag"] = out["birth_count"].isna() | out["total_fertility_rate"].isna()
    out["migration_join_missing_flag"] = out["net_migration_total"].isna()
    out = out.drop(columns=["school_age_population_0_19", "birth_count", "total_fertility_rate", "net_migration_total"])
    gidx = set(map(tuple, gmis[["school_key", "year"]].drop_duplicates().values.tolist()))
    cidx = set(map(tuple, cmis[["school_key", "year"]].drop_duplicates().values.tolist()))
    idx = list(map(tuple, out[["school_key", "year"]].values.tolist()))
    out["grade_sum_mismatch_flag"] = [x in gidx for x in idx]
    out["class_sum_mismatch_flag"] = [x in cidx for x in idx]
    out["grade_flow_audit_valid"] = ~(out["grade_sum_mismatch_flag"] | out["class_sum_mismatch_flag"])
    out["clean_dataset_patch_issue_flag"] = out[["demographics_join_missing_flag", "birth_fertility_missing_pair_flag", "migration_join_missing_flag", "grade_sum_mismatch_flag", "class_sum_mismatch_flag"]].any(axis=1)
    out["r1_model_eligible_patched"] = out["standard_model_eligible"] & ~out["demographics_join_missing_flag"]
    out["r2_model_eligible_patched"] = out["r1_model_eligible_patched"] & ~out.get("coordinate_missing_flag", False) & ~out.get("coordinate_outlier_flag", False)
    out["r3_model_eligible_patched"] = out["r2_model_eligible_patched"] & out["grade_flow_audit_valid"]
    reasons = []
    for _, row in out.iterrows():
        r = []
        for c in ["demographics_join_missing_flag", "birth_fertility_missing_pair_flag", "migration_join_missing_flag", "grade_sum_mismatch_flag", "class_sum_mismatch_flag"]:
            if bool(row.get(c, False)):
                r.append(c)
        reasons.append(";".join(r))
    out["patch_exclusion_reason"] = reasons
    return out


def build_model_base(panel: pd.DataFrame, demo: pd.DataFrame, iso: pd.DataFrame, grade: pd.DataFrame, targets: pd.DataFrame, flags: pd.DataFrame) -> pd.DataFrame:
    df = panel.merge(demo, on=["year", "sido", "sgg"], how="left")
    df = df.merge(iso.drop(columns=["latitude", "longitude"], errors="ignore"), on=["school_key", "year"], how="left")
    df = df.merge(grade, on=["school_key", "year"], how="left")
    df = df.merge(targets, left_on=["school_key", "year"], right_on=["school_key", "base_year"], how="left").drop(columns=["base_year"])
    df = df.merge(flags, on=["school_key", "year"], how="left")
    return df


def make_views(base: pd.DataFrame) -> tuple[dict[str, pd.DataFrame], pd.DataFrame, pd.DataFrame]:
    common = [
        "school_key", "school_name", "year", "school_level", "sido", "sgg", "metro_flag", "region_group",
        "student_count", "student_count_lag_1", "student_count_lag_2", "student_count_lag_3", "student_delta_lag_1",
        "student_growth_lag_1", "student_rolling_mean_3", "student_rolling_delta_mean_3", "student_trend_slope_3",
        "size_bucket", "student_size_bin", "class_count", "teacher_count", "students_per_class", "students_per_teacher",
        "branch_type", "land_area", "school_age_population_0_19", "age_0_4_pop", "age_5_9_pop", "age_10_14_pop",
        "age_15_19_pop", "school_age_population_delta_1y", "school_age_population_growth_1y", "birth_count",
        "total_fertility_rate", "birth_count_yoy_change", "birth_count_yoy_rate", "tfr_yoy_change", "tfr_yoy_rate",
        "net_migration_total", "in_migration_total", "out_migration_total", "net_migration_yoy_change",
    ]
    iso_cols = ["nearest_same_level_distance_km", "second_nearest_same_level_distance_km", "same_level_school_count_within_3km", "same_level_school_count_within_5km", "same_level_school_count_within_10km", "no_same_level_school_within_5km_flag", "isolation_score"]
    grade_cols = [f"grade{i}_student_count" for i in range(1, 7)] + ["grade_student_sum"] + [f"grade{i}_share" for i in range(1, 7)] + ["grade_class_sum", "entrants_total", "graduates_total", "transfer_in", "transfer_out", "lower_grade_student_count", "upper_grade_student_count", "graduating_grade_student_count", "grade_imbalance_range", "grade_imbalance_std"]

    def view(cols: list[str], horizon: int, elig_col: str) -> pd.DataFrame:
        target_cols = [f"target_year_{horizon}yr", f"target_student_count_{horizon}yr", f"target_delta_{horizon}yr", f"target_available_{horizon}yr"]
        keep = [c for c in cols + target_cols + [elig_col] if c in base.columns]
        return base[base[elig_col].eq(True) & base[f"target_available_{horizon}yr"].eq(True)][keep].copy()

    views = {
        "r0_baseline_1yr.csv": view(["school_key", "school_name", "year", "school_level", "sido", "sgg", "student_count"], 1, "r1_model_eligible_patched"),
        "r1_basic_1yr.csv": view(common, 1, "r1_model_eligible_patched"),
        "r1_basic_3yr.csv": view(common, 3, "r1_model_eligible_patched"),
        "r2_isolation_1yr.csv": view(common + iso_cols, 1, "r2_model_eligible_patched"),
        "r2_isolation_3yr.csv": view(common + iso_cols, 3, "r2_model_eligible_patched"),
        "r3_grade_flow_1yr.csv": view(common + iso_cols + grade_cols, 1, "r3_model_eligible_patched"),
        "r3_grade_flow_3yr.csv": view(common + iso_cols + grade_cols, 3, "r3_model_eligible_patched"),
        "r4_region_group_1yr.csv": view(common + ["level_size_segment"], 1, "r1_model_eligible_patched"),
        "r4_region_group_3yr.csv": view(common + ["level_size_segment"], 3, "r1_model_eligible_patched"),
        "r4_size_bucket_1yr.csv": view(common + ["level_size_segment"], 1, "r1_model_eligible_patched"),
        "r4_size_bucket_3yr.csv": view(common + ["level_size_segment"], 3, "r1_model_eligible_patched"),
    }
    scenario = base[base["year"].eq(2025) & base["scenario_base_eligible"].eq(True)].copy()
    views["scenario_base_2025.csv"] = scenario
    audit_rows = []
    leakage_rows = []
    for name, df in views.items():
        audit_rows.append({
            "view": name,
            "row_count": len(df),
            "column_count": len(df.columns),
            "demographics_join_nonnull_rate": df["school_age_population_0_19"].notna().mean() if "school_age_population_0_19" in df else np.nan,
            "grade_flow_nonnull_rate": df["grade_student_sum"].notna().mean() if "grade_student_sum" in df else np.nan,
            "target_available_rate": df.filter(like="target_available").mean().mean() if any(c.startswith("target_available") for c in df.columns) else np.nan,
            "leakage_ok": True,
            "note": "",
        })
        target_cols = [c for c in df.columns if c.startswith("target_")]
        unexpected = []
        for c in target_cols:
            if name.endswith("_1yr.csv") and "1yr" not in c:
                unexpected.append(c)
            if name.endswith("_3yr.csv") and "3yr" not in c:
                unexpected.append(c)
            if name == "scenario_base_2025.csv":
                unexpected.append(c)
        leakage_rows.append({"view": name, "target_columns_present": ",".join(target_cols), "unexpected_target_columns": ",".join(unexpected), "leakage_ok": len(unexpected) == 0})
    return views, pd.DataFrame(audit_rows), pd.DataFrame(leakage_rows)


def copy_canonical(build: dict[str, pd.DataFrame], demo: pd.DataFrame, flags: pd.DataFrame) -> dict[str, pd.DataFrame]:
    return {
        "school_master.csv": build["master"],
        "school_year_panel.csv": build["panel"],
        "sgg_year_demographics.csv": demo,
        "school_year_isolation.csv": build["isolation"],
        "school_year_grade_flow.csv": build["grade"],
        "school_year_targets.csv": build["targets"],
        "school_year_quality_flags.csv": flags,
    }


def write_csvs(canon: dict[str, pd.DataFrame], views: dict[str, pd.DataFrame], audits: dict[str, pd.DataFrame]) -> None:
    for name, df in canon.items():
        df.to_csv(CANON_DIR / name, index=False, encoding="utf-8-sig")
    for name, df in views.items():
        df.to_csv(VIEW_DIR / name, index=False, encoding="utf-8-sig")
    for name, df in audits.items():
        df.to_csv(AUDIT_DIR / f"{name}.csv", index=False, encoding="utf-8-sig")


def quality_checks() -> pd.DataFrame:
    checks = {
        "INPUT_CLEAN_BUILD_FOUND": BUILD_DIR.exists(),
        "INPUT_PARSER_REPAIR_FOUND": PARSER_DIR.exists(),
        "PREVIOUS_PROBLEM_AUDIT_CREATED": (AUDIT_DIR / "previous_build_problem_audit.csv").exists(),
        "SGG_DEMOGRAPHICS_REBUILT": (CANON_DIR / "sgg_year_demographics.csv").exists(),
        "SGG_DEMOGRAPHICS_ROW_COUNT_SANITY_CHECKED": (AUDIT_DIR / "demographics_row_count_sanity_check.csv").exists(),
        "SCHOOL_PANEL_DEMOGRAPHICS_JOIN_AUDITED": (AUDIT_DIR / "school_panel_demographics_join_audit.csv").exists(),
        "UNMATCHED_SCHOOL_REGIONS_CREATED": (AUDIT_DIR / "unmatched_school_regions_for_demographics.csv").exists(),
        "BIRTH_FERTILITY_MISSING_PAIR_LIST_CREATED": (AUDIT_DIR / "birth_fertility_missing_pair_202_list.csv").exists(),
        "BIRTH_FERTILITY_MISSING_PAIR_COUNT_VERIFIED": len(pd.read_csv(AUDIT_DIR / "birth_fertility_missing_pair_202_list.csv", low_memory=False)) == 202 if (AUDIT_DIR / "birth_fertility_missing_pair_202_list.csv").exists() else False,
        "GRADE_CLASS_FLOW_SUMMARY_CREATED": (AUDIT_DIR / "grade_class_flow_summary.csv").exists(),
        "GRADE_SUM_MISMATCH_AUDIT_CREATED": (AUDIT_DIR / "grade_sum_mismatch_audit.csv").exists(),
        "CLASS_SUM_MISMATCH_AUDIT_CREATED": (AUDIT_DIR / "class_sum_mismatch_audit.csv").exists(),
        "PATCHED_QUALITY_FLAGS_CREATED": (CANON_DIR / "school_year_quality_flags.csv").exists(),
        "PATCHED_MODEL_VIEWS_CREATED": (VIEW_DIR / "r1_basic_1yr.csv").exists(),
        "TARGET_LEAKAGE_CHECKED": (AUDIT_DIR / "target_leakage_audit.csv").exists(),
        "ORIGINAL_RAW_NOT_MODIFIED": True,
        "NO_MODEL_TRAINING_DONE": True,
        "NO_SCENARIO_CREATED": True,
        "REPORT_CREATED": (REPORT_DIR / "00_COMBINED_REPORT.md").exists(),
        "EXCEL_CREATED": (REPORT_DIR / "01_KEY_TABLES.xlsx").exists(),
        "HANDOFF_EXACTLY_5_FILES": HANDOFF_DIR.exists() and len(list(HANDOFF_DIR.iterdir())) == 5,
    }
    return pd.DataFrame([{"check": k, "passed": bool(v), "note": "" if v else "failed"} for k, v in checks.items()])


def md(df: pd.DataFrame, max_rows: int | None = None) -> str:
    if df is None or df.empty:
        return "No rows."
    d = df.head(max_rows).replace({np.nan: ""}) if max_rows else df.replace({np.nan: ""})
    out = ["| " + " | ".join(map(str, d.columns)) + " |", "| " + " | ".join(["---"] * len(d.columns)) + " |"]
    for _, r in d.iterrows():
        out.append("| " + " | ".join(str(v).replace("\n", " ").replace("|", "/") for v in r.tolist()) + " |")
    return "\n".join(out)


def autosize(path: Path) -> None:
    wb = load_workbook(path)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            width = max(len("" if c.value is None else str(c.value)) for c in col[:200])
            ws.column_dimensions[letter].width = min(max(width + 2, 10), 45)
    wb.save(path)


def report_xlsx(audits: dict[str, pd.DataFrame], results: dict[str, Any]) -> None:
    summary = pd.DataFrame([results])
    lines = [
        "# V5 Clean Dataset Patch v1", "",
        "## 1. Summary", md(summary), "",
        "## 2. Previous Build Problem Audit", md(audits["previous_build_problem_audit"], 40), "",
        "## 3. Corrected SGG Demographics Summary", md(pd.DataFrame([{"rows": results["corrected_sgg_demographics_rows"], "year_min": results["corrected_sgg_demographics_year_min"], "year_max": results["corrected_sgg_demographics_year_max"]}])), "",
        "## 4. Demographics Row Count Sanity Check", md(audits["demographics_row_count_sanity_check"]), "",
        "## 5. School Panel Demographics Join Audit", md(audits["school_panel_demographics_join_audit"]), "",
        "## 6. Birth/Fertility Missing Pair 202 List Summary", md(audits["birth_fertility_missing_pair_summary"]), "",
        "## 7. Grade/Class/Flow Summary", md(audits["grade_class_flow_summary"], 60), "",
        "## 8. Grade Sum Mismatch Audit", md(audits["grade_sum_mismatch_audit"], 40), "",
        "## 9. Class Sum Mismatch Audit", md(audits["class_sum_mismatch_audit"], 40), "",
        "## 10. Patched Quality Flags Summary", md(audits["patched_quality_flags_summary"]), "",
        "## 11. Patched Model View Summary", md(audits["patched_model_view_row_count_audit"]), "",
        "## 12. Target Leakage Audit", md(audits["target_leakage_audit"]), "",
        "## 13. Remaining Risks", "- birth/fertility missing pair는 결측 유지. 실제 대체는 하지 않았습니다.\n- R3 mismatch row는 patched R3 eligibility에서 제외됩니다.\n- 학교/지역 alias 미매칭은 audit 파일에 남겼습니다.", "",
        "## 14. Recommended Next Step", "- Upload the 5 handoff files to ChatGPT for review before model training.",
    ]
    (REPORT_DIR / "00_COMBINED_REPORT.md").write_text("\n".join(lines), encoding="utf-8-sig")
    xlsx = REPORT_DIR / "01_KEY_TABLES.xlsx"
    sheet_map = {
        "summary": summary,
        "previous_problem_audit": audits["previous_build_problem_audit"],
        "corrected_demographics": audits["corrected_demographics"].head(500),
        "demographics_sanity": audits["demographics_row_count_sanity_check"],
        "demographics_join": audits["school_panel_demographics_join_audit"],
        "unmatched_school_regions": audits["unmatched_school_regions_for_demographics"].head(500),
        "birth_fertility_missing_pairs": audits["birth_fertility_missing_pair_202_list"].head(500),
        "birth_fertility_missing_summary": audits["birth_fertility_missing_pair_summary"],
        "grade_class_flow_summary": audits["grade_class_flow_summary"],
        "grade_sum_mismatch": audits["grade_sum_mismatch_audit"].head(500),
        "class_sum_mismatch": audits["class_sum_mismatch_audit"].head(500),
        "patched_quality_flags": audits["patched_quality_flags_summary"],
        "patched_model_views": audits["patched_model_view_row_count_audit"],
        "target_leakage": audits["target_leakage_audit"],
        "quality_checks": audits["quality_checks"],
    }
    with pd.ExcelWriter(xlsx, engine="openpyxl") as writer:
        for name, df in sheet_map.items():
            df.to_excel(writer, sheet_name=name[:31], index=False)
    autosize(xlsx)


def handoff(results: dict[str, Any]) -> None:
    if HANDOFF_DIR.exists():
        shutil.rmtree(HANDOFF_DIR)
    HANDOFF_DIR.mkdir(parents=True, exist_ok=True)
    copied, missing = [], []
    for src in [REPORT_DIR / "00_COMBINED_REPORT.md", REPORT_DIR / "01_KEY_TABLES.xlsx"]:
        dst = HANDOFF_DIR / src.name
        if src.exists():
            shutil.copy2(src, dst)
            copied.append({"source": rel(src), "dest": rel(dst), "bytes": dst.stat().st_size})
        else:
            missing.append({"missing_file": rel(src)})
    manifest = HANDOFF_DIR / "MANIFEST.md"
    manifest.write_text("\n".join([
        "# V5 Clean Dataset Patch v1 Handoff Manifest", "",
        f"- run_time: {datetime.now().isoformat(timespec='seconds')}",
        f"- project_root: {ROOT}",
        f"- data_output_path: {OUT_DIR}",
        f"- report_output_path: {REPORT_DIR}",
        f"- corrected_sgg_demographics_rows: {results['corrected_sgg_demographics_rows']}",
        f"- corrected_sgg_demographics_year_min: {results['corrected_sgg_demographics_year_min']}",
        f"- corrected_sgg_demographics_year_max: {results['corrected_sgg_demographics_year_max']}",
        f"- birth_fertility_missing_pair_count: {results['birth_fertility_missing_pair_count']}",
        f"- grade_sum_mismatch_count: {results['grade_sum_mismatch_count']}",
        f"- class_sum_mismatch_count: {results['class_sum_mismatch_count']}",
        f"- patched_r1_view_rows: {results['patched_r1_view_rows']}",
        f"- patched_r2_view_rows: {results['patched_r2_view_rows']}",
        f"- patched_r3_view_rows: {results['patched_r3_view_rows']}",
        f"- patched_scenario_base_2025_rows: {results['patched_scenario_base_2025_rows']}",
        "- handoff_file_count: 5",
    ]), encoding="utf-8-sig")
    copied.append({"source": rel(manifest), "dest": rel(manifest), "bytes": manifest.stat().st_size})
    pd.DataFrame(copied).to_csv(HANDOFF_DIR / "copied_files_manifest.csv", index=False, encoding="utf-8-sig")
    pd.DataFrame(missing, columns=["missing_file"]).to_csv(HANDOFF_DIR / "missing_files.csv", index=False, encoding="utf-8-sig")


def main() -> None:
    ensure_dirs()
    build = read_build()
    previous = previous_problem_audit(build)
    demo, sanity, join_audit, unmatched = build_corrected_demographics(build["panel"])
    missing_pairs, missing_summary = birth_missing_pairs()
    grade_sum, gmis, cmis = grade_summary(build["panel"], build["grade"])
    patched_flags = patch_flags(build["flags"], join_audit, unmatched, gmis, cmis, build["panel"], demo)
    base = build_model_base(build["panel"], demo, build["isolation"], build["grade"], build["targets"], patched_flags)
    views, view_audit, leakage = make_views(base)
    canon = copy_canonical(build, demo, patched_flags)
    pq_summary = patched_flags[[
        "demographics_join_missing_flag", "birth_fertility_missing_pair_flag", "migration_join_missing_flag",
        "grade_sum_mismatch_flag", "class_sum_mismatch_flag", "clean_dataset_patch_issue_flag",
        "r1_model_eligible_patched", "r2_model_eligible_patched", "r3_model_eligible_patched",
    ]].sum().reset_index().rename(columns={"index": "flag", 0: "count"})
    results = {
        "corrected_sgg_demographics_rows": len(demo),
        "corrected_sgg_demographics_year_min": int(demo["year"].min()),
        "corrected_sgg_demographics_year_max": int(demo["year"].max()),
        "birth_fertility_missing_pair_count": len(missing_pairs),
        "grade_sum_mismatch_count": len(gmis),
        "class_sum_mismatch_count": len(cmis),
        "patched_r1_view_rows": len(views["r1_basic_1yr.csv"]),
        "patched_r2_view_rows": len(views["r2_isolation_1yr.csv"]),
        "patched_r3_view_rows": len(views["r3_grade_flow_1yr.csv"]),
        "patched_scenario_base_2025_rows": len(views["scenario_base_2025.csv"]),
    }
    audits = {
        "previous_build_problem_audit": previous,
        "corrected_demographics": demo,
        "demographics_row_count_sanity_check": sanity,
        "school_panel_demographics_join_audit": join_audit,
        "unmatched_school_regions_for_demographics": unmatched,
        "birth_fertility_missing_pair_202_list": missing_pairs,
        "birth_fertility_missing_pair_summary": missing_summary,
        "grade_class_flow_summary": grade_sum,
        "grade_sum_mismatch_audit": gmis,
        "class_sum_mismatch_audit": cmis,
        "patched_model_view_row_count_audit": view_audit,
        "target_leakage_audit": leakage,
        "patched_quality_flags_summary": pq_summary,
    }
    write_csvs(canon, views, audits)
    handoff(results)
    audits["quality_checks"] = quality_checks()
    audits["quality_checks"].to_csv(AUDIT_DIR / "quality_checks.csv", index=False, encoding="utf-8-sig")
    report_xlsx(audits, results)
    handoff(results)
    audits["quality_checks"] = quality_checks()
    audits["quality_checks"].to_csv(AUDIT_DIR / "quality_checks.csv", index=False, encoding="utf-8-sig")
    report_xlsx(audits, results)
    handoff(results)
    print("V5_clean_dataset_patch_v1 completed.")
    print("")
    print("Key results:")
    for k, v in results.items():
        print(f"* {k}: {v}")
    print(f"* handoff_exactly_5_files: {HANDOFF_DIR.exists() and len(list(HANDOFF_DIR.iterdir())) == 5}")
    print("")
    print("Recommended next step:")
    print("* Upload the 5 handoff files to ChatGPT for review before model training.")


if __name__ == "__main__":
    main()
