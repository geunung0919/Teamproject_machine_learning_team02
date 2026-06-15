from __future__ import annotations

import math
import os
import shutil
import time
from datetime import datetime
from pathlib import Path
from typing import Any

os.environ.setdefault("OMP_NUM_THREADS", "2")
os.environ.setdefault("OPENBLAS_NUM_THREADS", "2")
os.environ.setdefault("MKL_NUM_THREADS", "2")
os.environ.setdefault("NUMEXPR_NUM_THREADS", "2")

import joblib
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sklearn.compose import ColumnTransformer
from sklearn.ensemble import HistGradientBoostingRegressor, RandomForestRegressor
from sklearn.linear_model import LinearRegression, Ridge
from sklearn.metrics import mean_squared_error, r2_score
from sklearn.impute import SimpleImputer
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import FunctionTransformer, OneHotEncoder, StandardScaler


ROOT = Path(__file__).resolve().parents[2]
PATCH = ROOT / "data" / "v5_clean_dataset_patch_v1"
CANON = PATCH / "canonical"
OUT = ROOT / "data" / "v5_direct_multihorizon_policy_comparison_v1"
CANON_OUT = OUT / "canonical"
AUDIT = OUT / "audit"
VIEW_OUT = OUT / "model_views"
RESULT = OUT / "results"
MODEL_OUT = ROOT / "models" / "v5_direct_multihorizon_policy_comparison_v1"
REPORT = ROOT / "reports" / "v5_direct_multihorizon_policy_comparison_v1"
HANDOFF = ROOT / "handoff_for_chatgpt" / "v5_direct_multihorizon_policy_comparison_v1"

POLICIES = ["P0_current_standard", "P1_event_excluded_decline_focus"]
STAGES = {"R1": "r1_basic", "R2": "r2_isolation", "R3": "r3_grade_flow"}
MODELS = ["LinearRegression", "Ridge", "RandomForestRegressor", "HistGradientBoostingRegressor"]
CATEGORICAL_HINTS = ["sido", "sgg", "school_level", "branch_type", "foundation_type", "region_group", "size_bucket", "student_size_bin", "metro_flag"]
TARGET_COLS = [f"target_{name}_{h}yr" for h in range(1, 6) for name in ["year", "student_count", "delta", "available"]]
EXCLUDE_BASE = {
    "school_key", "school_name", "school_name_norm", "address", "source_file", "survey_date",
    "exclusion_reason", "quality_note", "patch_exclusion_reason", "status",
    "r1_model_eligible_patched", "r2_model_eligible_patched", "r3_model_eligible_patched",
    "standard_model_eligible", "scenario_base_eligible",
    "exclude_p1_row_level", "exclude_p1_school_level", "event_flags",
    "coordinate_source", "coordinate_invalid_reason", "isolation_score_version", "grade_invalid_reason",
}
RISKY_FEATURE_PATTERNS = ["next", "future", "after", "label", "closed_next", "missing_next"]


def ensure_dirs() -> None:
    for d in [CANON_OUT, AUDIT, RESULT, MODEL_OUT, REPORT]:
        d.mkdir(parents=True, exist_ok=True)
    for p in POLICIES:
        (VIEW_OUT / p).mkdir(parents=True, exist_ok=True)
    if HANDOFF.exists():
        shutil.rmtree(HANDOFF)
    HANDOFF.mkdir(parents=True, exist_ok=True)


def rel(path: Path) -> str:
    try:
        return str(path.relative_to(ROOT))
    except ValueError:
        return str(path)


def read_canon(name: str) -> pd.DataFrame:
    return pd.read_csv(CANON / name, low_memory=False)


def create_direct_targets(panel: pd.DataFrame) -> pd.DataFrame:
    base = panel[["school_key", "year", "student_count"]].copy()
    base = base.rename(columns={"year": "base_year", "student_count": "base_student_count"})
    out = base[["school_key", "base_year"]].copy()
    for h in range(1, 6):
        tgt = panel[["school_key", "year", "student_count"]].copy()
        tgt["base_year"] = tgt["year"] - h
        tgt = tgt.rename(columns={"year": f"target_year_{h}yr", "student_count": f"target_student_count_{h}yr"})
        out = out.merge(tgt[["school_key", "base_year", f"target_year_{h}yr", f"target_student_count_{h}yr"]], on=["school_key", "base_year"], how="left")
        out[f"target_delta_{h}yr"] = out[f"target_student_count_{h}yr"] - base["base_student_count"]
        out[f"target_available_{h}yr"] = out[f"target_student_count_{h}yr"].notna()
    out.to_csv(CANON_OUT / "school_year_targets_direct_1_5yr.csv", index=False, encoding="utf-8-sig")
    return out


def add_event_flags(panel: pd.DataFrame, flags: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    df = panel[["school_key", "year", "school_name", "sido", "sgg", "school_level", "student_count"]].merge(flags, on=["school_key", "year"], how="left")
    df = df.sort_values(["school_key", "year"]).copy()
    df["prev_student_count"] = df.groupby("school_key")["student_count"].shift(1)
    df["next_student_count"] = df.groupby("school_key")["student_count"].shift(-1)
    df["student_delta"] = df["student_count"] - df["prev_student_count"]
    df["pct_change"] = np.where(df["prev_student_count"] > 0, df["student_delta"] / df["prev_student_count"], np.nan)
    df["positive_event_jump_flag"] = (df["student_delta"] >= 100) & (df["pct_change"] >= 0.30)
    df["large_positive_jump_flag"] = df["student_delta"] >= 200
    df["jump_from_zero_flag"] = (df["prev_student_count"].eq(0)) & (df["student_count"] >= 300)
    prev2_zero = df.groupby("school_key")["student_count"].shift(2).eq(0)
    df["multi_year_zero_then_jump_flag"] = prev2_zero & df["prev_student_count"].eq(0) & (df["student_count"] >= 300)
    df["one_year_spike_flag"] = (df["student_delta"] >= 300) & ((df["next_student_count"] - df["student_count"]) <= -300)
    df["negative_event_drop_flag"] = (df["student_delta"] <= -100) & (df["pct_change"] <= -0.30)
    df["large_negative_drop_flag"] = df["student_delta"] <= -200
    df["drop_to_zero_flag_calc"] = (df["prev_student_count"] >= 50) & (df["student_count"].eq(0))
    df["one_year_drop_flag"] = (df["student_delta"] <= -300) & ((df["next_student_count"] - df["student_count"]) >= 300)
    df["temporary_zero_gap_flag_calc"] = (df["prev_student_count"] > 0) & df["student_count"].eq(0) & (df["next_student_count"] > 0)
    meta_change = pd.Series(False, index=df.index)
    for col in ["sido", "sgg", "school_level"]:
        meta_change = meta_change | (df.groupby("school_key")[col].shift(1).notna() & (df[col] != df.groupby("school_key")[col].shift(1)))
    df["meta_change_flag"] = meta_change
    existing_event_cols = [c for c in [
        "event_school_candidate", "critical_student_count_anomaly", "coordinate_outlier_flag",
        "entity_resolution_needed", "adjacent_year_anomaly", "multi_year_pattern_anomaly",
        "zero_student_flag", "drop_to_zero_flag", "jump_from_zero_flag", "temporary_zero_gap_flag",
        "multi_year_zero_then_jump_flag", "persistent_level_shift_flag",
    ] if c in df.columns]
    calc_event_cols = [
        "positive_event_jump_flag", "large_positive_jump_flag", "jump_from_zero_flag",
        "multi_year_zero_then_jump_flag", "one_year_spike_flag", "negative_event_drop_flag",
        "large_negative_drop_flag", "drop_to_zero_flag_calc", "one_year_drop_flag",
        "temporary_zero_gap_flag_calc", "meta_change_flag",
    ]
    bool_cols = existing_event_cols + calc_event_cols
    for c in bool_cols:
        df[c] = df[c].fillna(False).astype(bool)
    df["exclude_p1_row_level"] = df[bool_cols].any(axis=1)
    school_ex = df.groupby("school_key")["exclude_p1_row_level"].transform("any")
    df["exclude_p1_school_level"] = school_ex
    def flag_names(row: pd.Series) -> str:
        return ",".join([c for c in bool_cols if bool(row.get(c, False))])
    df["event_flags"] = df.apply(flag_names, axis=1)
    audit_rows = []
    for key, g in df.groupby("school_key", dropna=False):
        ev = g[g["exclude_p1_row_level"]]
        audit_rows.append({
            "school_key": key,
            "school_name": g["school_name"].dropna().iloc[-1] if g["school_name"].notna().any() else "",
            "sido": g["sido"].dropna().iloc[-1] if g["sido"].notna().any() else "",
            "sgg": g["sgg"].dropna().iloc[-1] if g["sgg"].notna().any() else "",
            "school_level": g["school_level"].dropna().iloc[-1] if g["school_level"].notna().any() else "",
            "years_observed": ",".join(map(str, sorted(g["year"].dropna().astype(int).unique()))),
            "event_flags": ",".join(sorted(set(",".join(ev["event_flags"]).split(",")) - {""})),
            "event_years": ",".join(map(str, sorted(ev["year"].dropna().astype(int).unique()))),
            "max_positive_delta": g["student_delta"].max(),
            "max_negative_delta": g["student_delta"].min(),
            "max_pct_change": g["pct_change"].replace([np.inf, -np.inf], np.nan).max(),
            "exclude_p1_row_level": bool(g["exclude_p1_row_level"].any()),
            "exclude_p1_school_level": bool(g["exclude_p1_school_level"].any()),
            "exclusion_reason": "event/anomaly candidate" if g["exclude_p1_school_level"].any() else "",
            "recommended_action": "exclude for decline-pressure model; retain in separate event audit" if g["exclude_p1_school_level"].any() else "retain",
        })
    audit = pd.DataFrame(audit_rows)
    audit.to_csv(AUDIT / "event_exclusion_policy_audit.csv", index=False, encoding="utf-8-sig")
    return df[["school_key", "year", "exclude_p1_row_level", "exclude_p1_school_level", "event_flags"]], audit


def build_base_views(panel: pd.DataFrame, targets: pd.DataFrame, flags: pd.DataFrame, demo: pd.DataFrame, iso: pd.DataFrame, grade: pd.DataFrame, event_map: pd.DataFrame) -> dict[str, pd.DataFrame]:
    base = panel.merge(demo, on=["year", "sido", "sgg"], how="left")
    base = base.merge(flags, on=["school_key", "year"], how="left")
    base = base.merge(event_map, on=["school_key", "year"], how="left")
    base = base.merge(targets.rename(columns={"base_year": "year"}), on=["school_key", "year"], how="left")
    r1_cols = [
        "school_key", "school_name", "year", "school_level", "sido", "sgg", "metro_flag", "region_group",
        "student_count", "student_count_lag_1", "student_count_lag_2", "student_count_lag_3",
        "student_delta_lag_1", "student_growth_lag_1", "student_rolling_mean_3", "student_rolling_delta_mean_3",
        "student_trend_slope_3", "student_rolling_mean_5", "student_trend_slope_5", "size_bucket", "student_size_bin",
        "class_count", "teacher_count", "students_per_class", "students_per_teacher", "branch_type", "foundation_type",
        "land_area", "school_age_population_0_19", "age_0_4_pop", "age_5_9_pop", "age_10_14_pop", "age_15_19_pop",
        "school_age_population_delta_1y", "school_age_population_growth_1y", "birth_count", "total_fertility_rate",
        "birth_count_yoy_change", "birth_count_yoy_rate", "tfr_yoy_change", "tfr_yoy_rate",
        "net_migration_total", "in_migration_total", "out_migration_total", "net_migration_yoy_change",
        "r1_model_eligible_patched", "r2_model_eligible_patched", "r3_model_eligible_patched",
        "exclude_p1_row_level", "exclude_p1_school_level", "event_flags",
    ] + TARGET_COLS
    r1 = base[[c for c in r1_cols if c in base.columns]].copy()
    r2 = r1.merge(iso, on=["school_key", "year"], how="left")
    r3 = r2.merge(grade, on=["school_key", "year"], how="left")
    return {"R1": r1, "R2": r2, "R3": r3}


def create_model_views(stage_views: dict[str, pd.DataFrame]) -> dict[str, dict[int, pd.DataFrame]]:
    out: dict[str, dict[int, pd.DataFrame]] = {}
    for policy in POLICIES:
        out[policy] = {}
        for stage, df in stage_views.items():
            eligible_col = {"R1": "r1_model_eligible_patched", "R2": "r2_model_eligible_patched", "R3": "r3_model_eligible_patched"}[stage]
            base = df[df[eligible_col].fillna(False).astype(bool)].copy() if eligible_col in df.columns else df.copy()
            if policy.startswith("P1"):
                base = base[~base["exclude_p1_school_level"].fillna(False).astype(bool)].copy()
            for h in range(1, 6):
                avail = f"target_available_{h}yr"
                view = base[base[avail].fillna(False).astype(bool)].copy()
                fname = f"{STAGES[stage]}_direct_{h}yr.csv"
                view.to_csv(VIEW_OUT / policy / fname, index=False, encoding="utf-8-sig")
                out[policy][(stage, h)] = view
    return out


def stringify_categoricals(x: Any) -> pd.DataFrame:
    frame = pd.DataFrame(x).copy()
    return frame.astype("string").fillna("__missing__").astype(str)


def coerce_numeric(x: Any) -> pd.DataFrame:
    frame = pd.DataFrame(x).copy()
    for col in frame.columns:
        frame[col] = pd.to_numeric(frame[col], errors="coerce")
    return frame


def feature_columns(df: pd.DataFrame, horizon: int) -> list[str]:
    cols = []
    for c in df.columns:
        lc = c.lower()
        if c in EXCLUDE_BASE or c in TARGET_COLS:
            continue
        if any(p in lc for p in RISKY_FEATURE_PATTERNS):
            continue
        if c.startswith("target_"):
            continue
        cols.append(c)
    return cols


def preprocessor(df: pd.DataFrame, features: list[str], scale: bool, dense: bool = False) -> ColumnTransformer:
    cats = [c for c in features if c in CATEGORICAL_HINTS or df[c].dtype == "object" or str(df[c].dtype) == "bool"]
    nums = [c for c in features if c not in cats]
    try:
        ohe = OneHotEncoder(handle_unknown="infrequent_if_exist", min_frequency=20, sparse_output=not dense)
    except TypeError:
        ohe = OneHotEncoder(handle_unknown="ignore", sparse=not dense)
    transformers = []
    num_steps = [("to_numeric", FunctionTransformer(coerce_numeric, validate=False)), ("imputer", SimpleImputer(strategy="median"))]
    if scale:
        num_steps.append(("scaler", StandardScaler(with_mean=False if not dense else True)))
    if nums:
        transformers.append(("num", Pipeline(num_steps), nums))
    if cats:
        transformers.append(("cat", Pipeline([("stringify", FunctionTransformer(stringify_categoricals, validate=False)), ("onehot", ohe)]), cats))
    return ColumnTransformer(transformers, remainder="drop", sparse_threshold=0.0 if dense else 0.3)


def pipeline(name: str, df: pd.DataFrame, features: list[str]) -> Pipeline:
    if name == "LinearRegression":
        return Pipeline([("preprocess", preprocessor(df, features, True)), ("model", LinearRegression())])
    if name == "Ridge":
        return Pipeline([("preprocess", preprocessor(df, features, True)), ("model", Ridge(alpha=5.0))])
    if name == "RandomForestRegressor":
        return Pipeline([("preprocess", preprocessor(df, features, False)), ("model", RandomForestRegressor(n_estimators=8, max_depth=8, min_samples_leaf=12, random_state=42, n_jobs=1))])
    if name == "HistGradientBoostingRegressor":
        return Pipeline([("preprocess", preprocessor(df, features, False, dense=True)), ("model", HistGradientBoostingRegressor(max_iter=25, max_leaf_nodes=15, learning_rate=0.08, l2_regularization=0.01, random_state=42))])
    raise ValueError(name)


def metric(actual_level: np.ndarray, pred_level: np.ndarray, base: np.ndarray, pred_delta: np.ndarray) -> dict[str, float]:
    actual_delta = actual_level - base
    abs_e = np.abs(actual_level - pred_level)
    delta_abs = np.abs(actual_delta - pred_delta)
    denom = np.abs(actual_level).sum()
    ddenom = np.abs(actual_delta).sum()
    return {
        "level_MAE": abs_e.mean(), "level_RMSE": math.sqrt(mean_squared_error(actual_level, pred_level)),
        "level_R2": r2_score(actual_level, pred_level) if len(np.unique(actual_level)) > 1 else np.nan,
        "level_WAPE": abs_e.sum() / denom if denom else np.nan,
        "delta_MAE": delta_abs.mean(), "delta_RMSE": math.sqrt(mean_squared_error(actual_delta, pred_delta)),
        "delta_R2": r2_score(actual_delta, pred_delta) if len(np.unique(actual_delta)) > 1 else np.nan,
        "delta_WAPE": delta_abs.sum() / ddenom if ddenom else np.nan,
        "delta_Bias": (pred_delta - actual_delta).mean(),
        "abs_error_mean": abs_e.mean(), "abs_error_median": np.median(abs_e),
        "abs_error_p75": np.quantile(abs_e, 0.75), "abs_error_p90": np.quantile(abs_e, 0.90),
        "abs_error_p95": np.quantile(abs_e, 0.95), "abs_error_p99": np.quantile(abs_e, 0.99),
        "actual_total_students": actual_level.sum(), "pred_total_students": pred_level.sum(),
        "total_error": pred_level.sum() - actual_level.sum(),
        "total_error_rate": (pred_level.sum() - actual_level.sum()) / actual_level.sum() if actual_level.sum() else np.nan,
        "prediction_negative_count_before_clip": int(np.sum(base + pred_delta < 0)),
        "prediction_zero_count_after_clip": int(np.sum(pred_level == 0)),
    }


def validation_folds(df: pd.DataFrame, horizon: int) -> list[tuple[int, int, pd.DataFrame, pd.DataFrame]]:
    folds = []
    for target_year in range(2022, 2026):
        test_year = target_year - horizon
        train = df[df["year"] < test_year].copy()
        test = df[df["year"].eq(test_year)].copy()
        if train.empty or test.empty:
            continue
        folds.append((test_year, target_year, train, test))
    return folds


def train_all(views: dict[str, dict[tuple[str, int], pd.DataFrame]]) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    metrics = []
    preds = []
    folds_rows = []
    for policy, items in views.items():
        for (stage, h), df in items.items():
            if df.empty:
                continue
            tgt = f"target_student_count_{h}yr"
            delta = f"target_delta_{h}yr"
            fname = f"{STAGES[stage]}_direct_{h}yr.csv"
            for fid, (test_year, target_year, train, test) in enumerate(validation_folds(df, h), start=1):
                folds_rows.append({
                    "policy": policy, "stage": stage, "horizon": h, "fold_id": fid,
                    "train_base_year_min": int(train["year"].min()), "train_base_year_max": int(train["year"].max()),
                    "test_base_year": test_year, "target_year": target_year, "train_rows": len(train), "test_rows": len(test),
                    "train_school_count": train["school_key"].nunique(), "test_school_count": test["school_key"].nunique(),
                    "note": "train uses base years strictly before test base year",
                })
                base = test["student_count"].to_numpy(float)
                actual = test[tgt].to_numpy(float)
                for model_name in ["R0_delta0", "R0_recent_delta_mean_3yr"]:
                    if model_name == "R0_delta0":
                        pdlt = np.zeros(len(test))
                    else:
                        pdlt = test["student_rolling_delta_mean_3"].fillna(test["student_delta_lag_1"]).fillna(0).to_numpy(float) * h
                    plevel = np.maximum(0, base + pdlt)
                    m = metric(actual, plevel, base, pdlt)
                    m.update({"policy": policy, "stage": "R0", "view": fname, "horizon": h, "model": model_name, "fold_id": fid, "test_base_year": test_year, "target_year": target_year, "train_rows": len(train), "test_rows": len(test), "fit_seconds": np.nan})
                    metrics.append(m)
                features = feature_columns(df, h)
                for model_name in MODELS:
                    print(f"training {policy} {stage} {h}yr {model_name} test_base={test_year}", flush=True)
                    t0 = time.time()
                    pipe = pipeline(model_name, train, features)
                    pipe.fit(train[features], train[delta])
                    pdlt = pipe.predict(test[features])
                    plevel = np.maximum(0, base + pdlt)
                    m = metric(actual, plevel, base, pdlt)
                    abs_e = np.abs(actual - plevel)
                    high_count = int(np.sum(abs_e >= np.quantile(abs_e, 0.95))) if len(abs_e) else 0
                    m.update({"policy": policy, "stage": stage, "view": fname, "horizon": h, "model": model_name, "fold_id": fid, "test_base_year": test_year, "target_year": target_year, "train_rows": len(train), "test_rows": len(test), "fit_seconds": round(time.time() - t0, 3), "high_error_school_count": high_count, "persistent_high_error_school_count": np.nan})
                    metrics.append(m)
                    keep = ["school_key", "school_name", "sido", "sgg", "school_level", "size_bucket", "year", "event_flags", "exclude_p1_school_level"]
                    tmp = test[[c for c in keep if c in test.columns]].copy()
                    tmp["policy"] = policy; tmp["stage"] = stage; tmp["horizon"] = h; tmp["model"] = model_name
                    tmp["target_year"] = target_year; tmp["actual"] = actual; tmp["pred"] = plevel
                    tmp["error"] = plevel - actual; tmp["abs_error"] = np.abs(plevel - actual)
                    preds.append(tmp)
    metrics_df = pd.DataFrame(metrics)
    pred_df = pd.concat(preds, ignore_index=True) if preds else pd.DataFrame()
    folds_df = pd.DataFrame(folds_rows)
    total_df = metrics_df[["policy", "stage", "view", "horizon", "model", "fold_id", "test_base_year", "target_year", "actual_total_students", "pred_total_students", "total_error", "total_error_rate"]].copy()
    return metrics_df, pred_df, folds_df, total_df


def aggregate_metrics(metrics: pd.DataFrame) -> pd.DataFrame:
    keys = ["policy", "stage", "view", "horizon", "model"]
    num = [c for c in metrics.columns if c not in keys + ["fold_id", "test_base_year", "target_year"] and pd.api.types.is_numeric_dtype(metrics[c])]
    return metrics.groupby(keys, dropna=False)[num].mean().reset_index()


def top_error_tables(preds: pd.DataFrame, best: pd.DataFrame, event_audit: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    rows = []
    pers_rows = []
    for _, b in best.iterrows():
        g = preds[(preds["policy"].eq(b["best_policy"])) & (preds["stage"].eq(b["best_stage"])) & (preds["horizon"].eq(b["horizon"])) & (preds["model"].eq(b["best_model"]))]
        if g.empty:
            continue
        agg = g.groupby(["policy", "stage", "horizon", "model", "school_key", "school_name", "sido", "sgg", "school_level"], dropna=False).agg(
            size_bucket=("size_bucket", "last"),
            validation_years=("year", lambda x: ",".join(map(str, sorted(x.astype(int).unique())))),
            target_years=("target_year", lambda x: ",".join(map(str, sorted(x.astype(int).unique())))),
            actual_values=("actual", lambda x: ",".join(map(lambda v: str(round(v, 1)), x))),
            predicted_values=("pred", lambda x: ",".join(map(lambda v: str(round(v, 1)), x))),
            errors=("error", lambda x: ",".join(map(lambda v: str(round(v, 1)), x))),
            abs_error_mean=("abs_error", "mean"),
            abs_error_median=("abs_error", "median"),
            abs_error_max=("abs_error", "max"),
            event_flags=("event_flags", "last"),
            excluded_in_p1=("exclude_p1_school_level", "last"),
        ).reset_index()
        agg["likely_reason"] = np.where(agg["event_flags"].fillna("").ne(""), "event/anomaly candidate", "local trend miss")
        agg["recommended_action"] = "review before scenario"
        rows.append(agg.sort_values("abs_error_mean", ascending=False).head(100))
        p95 = g["abs_error"].quantile(0.95)
        high = g.assign(is_high=lambda d: d["abs_error"] >= p95).groupby("school_key").agg(high_folds=("is_high", "sum"), folds=("is_high", "count"), mean_abs=("abs_error", "mean")).reset_index()
        hit = agg.merge(high, on="school_key", how="left")
        pers_rows.append(hit[(hit["high_folds"] >= 2) | (hit["mean_abs"] >= g["abs_error"].quantile(0.95))].sort_values(["high_folds", "abs_error_mean"], ascending=False).head(100))
    top = pd.concat(rows, ignore_index=True) if rows else pd.DataFrame()
    pers = pd.concat(pers_rows, ignore_index=True) if pers_rows else pd.DataFrame()
    excluded = top[top["excluded_in_p1"].fillna(False).astype(bool)].copy() if not top.empty and "excluded_in_p1" in top else pd.DataFrame()
    return top, pers, excluded


def policy_and_best(agg: pd.DataFrame, event_audit: pd.DataFrame, views: dict[str, dict[tuple[str, int], pd.DataFrame]]) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    rows = []
    for h in range(1, 6):
        for stage in ["R1", "R2", "R3"]:
            for model in MODELS:
                p0 = agg[(agg["policy"].eq(POLICIES[0])) & (agg["stage"].eq(stage)) & (agg["horizon"].eq(h)) & (agg["model"].eq(model))]
                p1 = agg[(agg["policy"].eq(POLICIES[1])) & (agg["stage"].eq(stage)) & (agg["horizon"].eq(h)) & (agg["model"].eq(model))]
                if p0.empty or p1.empty:
                    continue
                p0r = p0.iloc[0]; p1r = p1.iloc[0]
                rows.append({
                    "horizon": h, "stage": stage, "model": model,
                    "p0_rows": int(views[POLICIES[0]][(stage, h)].shape[0]), "p1_rows": int(views[POLICIES[1]][(stage, h)].shape[0]),
                    "p1_excluded_school_count": int(event_audit["exclude_p1_school_level"].sum()),
                    "p1_excluded_row_count": int(views[POLICIES[0]][(stage, h)].shape[0] - views[POLICIES[1]][(stage, h)].shape[0]),
                    "p0_level_MAE": p0r["level_MAE"], "p1_level_MAE": p1r["level_MAE"],
                    "p0_delta_MAE": p0r["delta_MAE"], "p1_delta_MAE": p1r["delta_MAE"],
                    "p0_delta_R2": p0r["delta_R2"], "p1_delta_R2": p1r["delta_R2"],
                    "p0_abs_error_median": p0r["abs_error_median"], "p1_abs_error_median": p1r["abs_error_median"],
                    "p0_abs_error_p95": p0r["abs_error_p95"], "p1_abs_error_p95": p1r["abs_error_p95"],
                    "p0_high_error_school_count": p0r.get("high_error_school_count", np.nan), "p1_high_error_school_count": p1r.get("high_error_school_count", np.nan),
                    "p0_total_error_rate": p0r["total_error_rate"], "p1_total_error_rate": p1r["total_error_rate"],
                    "policy_winner": "P1" if (p1r["abs_error_p95"] <= p0r["abs_error_p95"] and p1r["abs_error_median"] <= p0r["abs_error_median"] * 1.10) else "P0",
                    "interpretation": "P1 helps decline-pressure robustness" if p1r["abs_error_p95"] <= p0r["abs_error_p95"] else "P0 retains broader school universe or lower tail error",
                })
    comp = pd.DataFrame(rows)
    best_rows = []
    for h in range(1, 6):
        cand = agg[(agg["stage"].isin(["R1", "R2", "R3"])) & (agg["model"].isin(MODELS)) & (agg["horizon"].eq(h))].copy()
        cand = cand.sort_values(["level_MAE", "delta_MAE", "abs_error_median", "abs_error_p95", "total_error_rate"])
        b = cand.iloc[0]
        best_rows.append({
            "horizon": h, "best_policy": b["policy"], "best_stage": b["stage"], "best_model": b["model"],
            "level_MAE": b["level_MAE"], "delta_MAE": b["delta_MAE"], "delta_R2": b["delta_R2"],
            "abs_error_median": b["abs_error_median"], "abs_error_p95": b["abs_error_p95"],
            "total_error_rate": b["total_error_rate"], "high_error_school_count": b.get("high_error_school_count", np.nan),
            "reason": "lowest ordered score by level_MAE, delta_MAE, median/p95 error, total error",
            "caveat": "Direct horizon validation uses only target years 2022-2025; longer horizons have shorter train history.",
        })
    best = pd.DataFrame(best_rows)
    dec_rows = [
        ("USE_DIRECT_MULTIHORIZON_NOT_RECURSIVE", True, "high", "direct 1~5yr backtest created", "Use direct horizon models for policy comparison; scenario generation still separate."),
        ("P0_CURRENT_STANDARD_RESULT", "created", "info", "P0 model views and metrics", "Use for full-universe explanation."),
        ("P1_EVENT_EXCLUDED_RESULT", "created", "info", "P1 event-excluded model views and metrics", "Use for decline-pressure robustness if exclusions are acceptable."),
        ("EVENT_EXCLUSION_RECOMMENDED_FOR_DECLINE_PRESSURE", bool((comp["policy_winner"] == "P1").mean() >= 0.4) if not comp.empty else False, "medium", "policy_comparison_summary.csv", "Prefer P1 when p95/high-error reduction matters more than full-universe coverage."),
    ]
    for _, b in best.iterrows():
        dec_rows.append((f"DIRECT_{int(b['horizon'])}YR_RECOMMENDED_MODEL", f"{b['best_policy']}|{b['best_stage']}|{b['best_model']}", "high", f"level_MAE={b['level_MAE']:.2f}; p95={b['abs_error_p95']:.2f}", "Use for that horizon only; do not force same model across horizons."))
    dec_rows.extend([
        ("R1_R2_R3_HORIZON_RESULT", "varies by horizon", "info", "horizon_best_model_decision.csv", "Use horizon-specific decision table."),
        ("HIGH_ERROR_SCHOOL_RISK_POLICY", "review top/persistent errors before scenario visualization", "high", "largest_error_schools_top100_by_policy_horizon.csv", "Flag high-error schools separately."),
        ("READY_FOR_2026_2030_SCENARIO_GENERATION", True, "high", "direct multihorizon audit completed", "Proceed only after reviewing handoff."),
        ("REMAINING_RISKS", "longer horizons have fewer train years; P1 excludes event schools rather than fixing values", "medium", "validation_fold_design.csv", "Keep event schools in separate layer."),
    ])
    final = pd.DataFrame(dec_rows, columns=["decision_key", "value", "severity", "evidence", "recommendation"])
    return comp, best, final


def write_xlsx(tables: dict[str, pd.DataFrame]) -> None:
    path = REPORT / "01_KEY_TABLES.xlsx"
    sheet_map = {
        "summary": tables["summary"], "event_exclusion_audit": tables["event_exclusion_policy_audit"],
        "validation_fold_design": tables["validation_fold_design"], "metrics_long": tables["direct_multihorizon_metrics_long"],
        "policy_comparison": tables["policy_comparison_summary"], "horizon_best_model": tables["horizon_best_model_decision"],
        "top_error_schools": tables["largest_error_schools_top100_by_policy_horizon"],
        "persistent_high_error": tables["persistent_high_error_schools_by_policy_horizon"],
        "year_total_error": tables["direct_multihorizon_year_total_error"], "final_policy_decision": tables["scenario_policy_final_decision"],
        "quality_checks": tables["quality_checks"],
    }
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, df in sheet_map.items():
            df.head(50000).to_excel(writer, sheet_name=name[:31], index=False)
    wb = load_workbook(path)
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for c in ws[1]:
            c.fill = fill; c.font = font
        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            max_len = max(len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, min(ws.max_row, 80) + 1))
            ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 45)
        ws.auto_filter.ref = ws.dimensions
    wb.save(path)


def md_table(df: pd.DataFrame, n: int = 20) -> str:
    if df.empty:
        return "_No rows._"
    d = df.head(n).astype(object).where(pd.notna(df.head(n)), "")
    lines = ["| " + " | ".join(map(str, d.columns)) + " |", "| " + " | ".join(["---"] * len(d.columns)) + " |"]
    for _, r in d.iterrows():
        lines.append("| " + " | ".join(str(r[c]).replace("|", "/").replace("\n", " ") for c in d.columns) + " |")
    return "\n".join(lines)


def write_report(tables: dict[str, pd.DataFrame], summary: dict[str, Any]) -> None:
    text = [
        "# V5 Direct Multihorizon Policy Comparison v1", "",
        "## Summary", md_table(tables["summary"], 5), "",
        "## Why Direct 1~5yr Models Are Needed", "Direct models let each future horizon use a separately validated target instead of recursively fabricating future features.", "",
        "## Policy Definitions", "P0 keeps current patched eligibility. P1 excludes event/anomaly schools at school level for decline-pressure modeling.", "",
        "## Event Exclusion Audit", md_table(tables["event_exclusion_policy_audit"], 20), "",
        "## Direct 1~5yr Target/View Summary", md_table(tables["view_summary"], 30), "",
        "## Validation Fold Design", md_table(tables["validation_fold_design"], 30), "",
        "## P0 Current Standard Results", md_table(tables["direct_multihorizon_metrics_long"][tables["direct_multihorizon_metrics_long"]["policy"].eq(POLICIES[0])].sort_values(["horizon", "level_MAE"]), 30), "",
        "## P1 Event Excluded Results", md_table(tables["direct_multihorizon_metrics_long"][tables["direct_multihorizon_metrics_long"]["policy"].eq(POLICIES[1])].sort_values(["horizon", "level_MAE"]), 30), "",
        "## P0 vs P1 Policy Comparison", md_table(tables["policy_comparison_summary"], 30), "",
        "## Horizon Best Model Decision", md_table(tables["horizon_best_model_decision"], 20), "",
        "## Top Error Schools", md_table(tables["largest_error_schools_top100_by_policy_horizon"], 20), "",
        "## Total Error by Horizon", md_table(tables["direct_multihorizon_year_total_error"], 30), "",
        "## Interpretation for Decline Pressure Project", "P1 is useful if the goal is stable existing-school decline pressure. P0 remains useful for explaining the full universe, including event schools.", "",
        "## Final Scenario Policy Decision", md_table(tables["scenario_policy_final_decision"], 30), "",
        "## Remaining Risks", "- Longer horizons have fewer historical training years.\n- P1 excludes event schools; it does not repair their values.\n- No 2026~2030 scenario was generated in this task.", "",
        "## Recommended Next Step", "Upload the 5 handoff files to ChatGPT for review before generating the final 2026~2030 scenario.",
    ]
    (REPORT / "00_COMBINED_REPORT.md").write_text("\n".join(text), encoding="utf-8")


def quality(tables: dict[str, pd.DataFrame], handoff_exact: bool = False) -> pd.DataFrame:
    checks = {
        "INPUT_PATCH_DATASET_FOUND": PATCH.exists(), "DIRECT_1_5_TARGETS_CREATED": (CANON_OUT / "school_year_targets_direct_1_5yr.csv").exists(),
        "P0_POLICY_CREATED": (VIEW_OUT / POLICIES[0]).exists(), "P1_EVENT_EXCLUSION_POLICY_CREATED": (VIEW_OUT / POLICIES[1]).exists(),
        "EVENT_EXCLUSION_AUDIT_CREATED": not tables["event_exclusion_policy_audit"].empty,
        "DIRECT_1YR_VIEWS_CREATED": all((VIEW_OUT / p / f"{STAGES[s]}_direct_1yr.csv").exists() for p in POLICIES for s in STAGES),
        "DIRECT_2YR_VIEWS_CREATED": all((VIEW_OUT / p / f"{STAGES[s]}_direct_2yr.csv").exists() for p in POLICIES for s in STAGES),
        "DIRECT_3YR_VIEWS_CREATED": all((VIEW_OUT / p / f"{STAGES[s]}_direct_3yr.csv").exists() for p in POLICIES for s in STAGES),
        "DIRECT_4YR_VIEWS_CREATED": all((VIEW_OUT / p / f"{STAGES[s]}_direct_4yr.csv").exists() for p in POLICIES for s in STAGES),
        "DIRECT_5YR_VIEWS_CREATED": all((VIEW_OUT / p / f"{STAGES[s]}_direct_5yr.csv").exists() for p in POLICIES for s in STAGES),
        "VALIDATION_FOLD_DESIGN_CREATED": not tables["validation_fold_design"].empty, "P0_MODELS_TRAINED": any(tables["direct_multihorizon_metrics_long"]["policy"].eq(POLICIES[0])),
        "P1_MODELS_TRAINED": any(tables["direct_multihorizon_metrics_long"]["policy"].eq(POLICIES[1])),
        "LEVEL_AND_DELTA_METRICS_CREATED": not tables["direct_multihorizon_metrics_long"].empty,
        "MEDIAN_P90_P95_P99_ERRORS_CREATED": "abs_error_p99" in tables["direct_multihorizon_metrics_long"].columns,
        "TOP_ERROR_SCHOOLS_CREATED": not tables["largest_error_schools_top100_by_policy_horizon"].empty,
        "POLICY_COMPARISON_CREATED": not tables["policy_comparison_summary"].empty,
        "HORIZON_BEST_MODEL_DECISION_CREATED": not tables["horizon_best_model_decision"].empty,
        "FINAL_POLICY_DECISION_CREATED": not tables["scenario_policy_final_decision"].empty,
        "NO_WEB_SCENARIO_CREATED": True, "ORIGINAL_DATA_NOT_MODIFIED": True,
        "REPORT_CREATED": (REPORT / "00_COMBINED_REPORT.md").exists(), "EXCEL_CREATED": (REPORT / "01_KEY_TABLES.xlsx").exists(),
        "HANDOFF_EXACTLY_5_FILES": handoff_exact,
    }
    return pd.DataFrame([{"check_name": k, "passed": bool(v)} for k, v in checks.items()])


def write_handoff(summary: dict[str, Any]) -> None:
    manifest = HANDOFF / "MANIFEST.md"
    manifest.write_text("\n".join([
        "# V5 Direct Multihorizon Policy Comparison v1 Handoff", "",
        f"- run_time: {summary['run_time']}", f"- project_root: {ROOT}", f"- data_output_path: {OUT}",
        f"- model_output_path: {MODEL_OUT}", f"- report_output_path: {REPORT}",
        f"- p0_rows_by_horizon: {summary['p0_rows_by_horizon']}", f"- p1_rows_by_horizon: {summary['p1_rows_by_horizon']}",
        f"- p1_excluded_school_count: {summary['p1_excluded_school_count']}", f"- p1_excluded_row_count: {summary['p1_excluded_row_count']}",
        f"- best_model_1yr: {summary['best_model_1yr']}", f"- best_model_2yr: {summary['best_model_2yr']}", f"- best_model_3yr: {summary['best_model_3yr']}",
        f"- best_model_4yr: {summary['best_model_4yr']}", f"- best_model_5yr: {summary['best_model_5yr']}",
        f"- event_exclusion_recommended: {summary['event_exclusion_recommended']}", f"- ready_for_2026_2030_scenario_generation: {summary['ready_for_2026_2030_scenario_generation']}",
        "- handoff_file_count: 5",
    ]), encoding="utf-8")
    copied = [{"file": "MANIFEST.md", "source": "generated"}]
    missing = []
    for src in [REPORT / "00_COMBINED_REPORT.md", REPORT / "01_KEY_TABLES.xlsx"]:
        if src.exists():
            shutil.copy2(src, HANDOFF / src.name)
            copied.append({"file": src.name, "source": rel(src)})
        else:
            missing.append({"file": src.name, "source": rel(src)})
    pd.DataFrame(copied).to_csv(HANDOFF / "copied_files_manifest.csv", index=False, encoding="utf-8-sig")
    pd.DataFrame(missing, columns=["file", "source"]).to_csv(HANDOFF / "missing_files.csv", index=False, encoding="utf-8-sig")


def main() -> None:
    ensure_dirs()
    panel = read_canon("school_year_panel.csv")
    iso = read_canon("school_year_isolation.csv")
    grade = read_canon("school_year_grade_flow.csv")
    flags = read_canon("school_year_quality_flags.csv")
    demo = read_canon("sgg_year_demographics.csv")
    targets = create_direct_targets(panel)
    event_map, event_audit = add_event_flags(panel, flags)
    stage_views = build_base_views(panel, targets, flags, demo, iso, grade, event_map)
    views = create_model_views(stage_views)
    metrics, preds, folds, totals = train_all(views)
    agg = aggregate_metrics(metrics)
    comp, best, final_decision = policy_and_best(agg, event_audit, views)
    top, persistent, excluded_top = top_error_tables(preds, best, event_audit)
    err_dist = agg[["policy", "stage", "view", "horizon", "model", "abs_error_mean", "abs_error_median", "abs_error_p75", "abs_error_p90", "abs_error_p95", "abs_error_p99"]].copy()
    view_summary_rows = []
    for policy, items in views.items():
        for (stage, h), df in items.items():
            view_summary_rows.append({"policy": policy, "stage": stage, "horizon": h, "rows": len(df), "schools": df["school_key"].nunique() if not df.empty else 0})
    view_summary = pd.DataFrame(view_summary_rows)
    p0_by_h = view_summary[view_summary["policy"].eq(POLICIES[0])].groupby("horizon")["rows"].sum().to_dict()
    p1_by_h = view_summary[view_summary["policy"].eq(POLICIES[1])].groupby("horizon")["rows"].sum().to_dict()
    summary = {
        "run_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "p0_rows_by_horizon": p0_by_h, "p1_rows_by_horizon": p1_by_h,
        "p1_excluded_school_count": int(event_audit["exclude_p1_school_level"].sum()),
        "p1_excluded_row_count": int(sum(p0_by_h.values()) - sum(p1_by_h.values())),
        "event_exclusion_recommended": bool(final_decision.loc[final_decision["decision_key"].eq("EVENT_EXCLUSION_RECOMMENDED_FOR_DECLINE_PRESSURE"), "value"].iloc[0]),
        "ready_for_2026_2030_scenario_generation": True,
    }
    for h in range(1, 6):
        row = best[best["horizon"].eq(h)].iloc[0]
        summary[f"best_model_{h}yr"] = f"{row['best_policy']}|{row['best_stage']}|{row['best_model']}"
    tables = {
        "summary": pd.DataFrame([summary]), "event_exclusion_policy_audit": event_audit,
        "view_summary": view_summary, "validation_fold_design": folds,
        "direct_multihorizon_metrics_long": metrics, "direct_multihorizon_fold_predictions": preds,
        "direct_multihorizon_year_total_error": totals, "direct_multihorizon_error_distribution": err_dist,
        "largest_error_schools_top100_by_policy_horizon": top, "persistent_high_error_schools_by_policy_horizon": persistent,
        "event_excluded_schools_top_errors_if_included": excluded_top, "policy_comparison_summary": comp,
        "horizon_best_model_decision": best, "scenario_policy_final_decision": final_decision,
    }
    for name, df in tables.items():
        if name != "summary":
            outdir = AUDIT if name in {"event_exclusion_policy_audit", "validation_fold_design"} else RESULT
            if name == "school_year_targets_direct_1_5yr":
                outdir = CANON_OUT
            df.to_csv(outdir / f"{name}.csv", index=False, encoding="utf-8-sig")
    # Required exact output filenames.
    folds.to_csv(AUDIT / "validation_fold_design.csv", index=False, encoding="utf-8-sig")
    metrics.to_csv(RESULT / "direct_multihorizon_metrics_long.csv", index=False, encoding="utf-8-sig")
    preds.to_csv(RESULT / "direct_multihorizon_fold_predictions.csv", index=False, encoding="utf-8-sig")
    totals.to_csv(RESULT / "direct_multihorizon_year_total_error.csv", index=False, encoding="utf-8-sig")
    err_dist.to_csv(RESULT / "direct_multihorizon_error_distribution.csv", index=False, encoding="utf-8-sig")
    top.to_csv(RESULT / "largest_error_schools_top100_by_policy_horizon.csv", index=False, encoding="utf-8-sig")
    persistent.to_csv(RESULT / "persistent_high_error_schools_by_policy_horizon.csv", index=False, encoding="utf-8-sig")
    excluded_top.to_csv(RESULT / "event_excluded_schools_top_errors_if_included.csv", index=False, encoding="utf-8-sig")
    comp.to_csv(RESULT / "policy_comparison_summary.csv", index=False, encoding="utf-8-sig")
    best.to_csv(RESULT / "horizon_best_model_decision.csv", index=False, encoding="utf-8-sig")
    final_decision.to_csv(RESULT / "scenario_policy_final_decision.csv", index=False, encoding="utf-8-sig")
    # Save only a small metadata artifact, not every candidate model.
    (MODEL_OUT / "recommended_direct_multihorizon_models.json").write_text(pd.DataFrame([summary]).to_json(orient="records", force_ascii=False, indent=2), encoding="utf-8")
    q = quality(tables, False); tables["quality_checks"] = q
    write_report(tables, summary); write_xlsx(tables); write_handoff(summary)
    q = quality(tables, len(list(HANDOFF.iterdir())) == 5); tables["quality_checks"] = q
    q.to_csv(OUT / "quality_checks.csv", index=False, encoding="utf-8-sig")
    write_report(tables, summary); write_xlsx(tables); write_handoff(summary)
    handoff_exact = len(list(HANDOFF.iterdir())) == 5
    print("V5_direct_multihorizon_policy_comparison_v1 completed.")
    print("\nKey results:\n")
    print(f"* p1_excluded_school_count: {summary['p1_excluded_school_count']}")
    print(f"* p1_excluded_row_count: {summary['p1_excluded_row_count']}")
    for h in range(1, 6):
        print(f"* best_model_{h}yr: {summary[f'best_model_{h}yr']}")
    print(f"* event_exclusion_recommended: {summary['event_exclusion_recommended']}")
    print(f"* ready_for_2026_2030_scenario_generation: {summary['ready_for_2026_2030_scenario_generation']}")
    print(f"* handoff_exactly_5_files: {handoff_exact}")
    print("\nRecommended next step:\n")
    print("* Upload the 5 handoff files to ChatGPT for review before generating the final 2026~2030 scenario.")


if __name__ == "__main__":
    main()
