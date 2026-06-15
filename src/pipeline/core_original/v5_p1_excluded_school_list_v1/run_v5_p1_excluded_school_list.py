from __future__ import annotations

import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
INPUT = ROOT / "data" / "v5_direct_multihorizon_policy_comparison_v1"
PATCH = ROOT / "data" / "v5_clean_dataset_patch_v1"
OUT = ROOT / "data" / "v5_p1_excluded_school_list_v1"
REPORT = ROOT / "reports" / "v5_p1_excluded_school_list_v1"
HANDOFF = ROOT / "handoff_for_chatgpt" / "v5_p1_excluded_school_list_v1"


def ensure_dirs() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    REPORT.mkdir(parents=True, exist_ok=True)
    if HANDOFF.exists():
        shutil.rmtree(HANDOFF)
    HANDOFF.mkdir(parents=True, exist_ok=True)


def rel(path: Path) -> str:
    try:
        return str(path.relative_to(ROOT))
    except ValueError:
        return str(path)


def categorize(flags: Any) -> str:
    text = str(flags or "").lower()
    if any(x in text for x in ["coordinate_outlier", "entity_resolution", "meta_change"]):
        return "coordinate_or_entity_issue"
    if any(x in text for x in ["temporary_zero", "one_year_drop", "one_year_spike"]):
        return "temporary_zero_or_data_artifact"
    if any(x in text for x in ["jump_from_zero", "multi_year_zero_then_jump"]):
        return "new_or_reopened_school"
    if any(x in text for x in ["positive_event_jump", "large_positive_jump", "persistent_level_shift"]):
        return "redevelopment_or_large_growth"
    if any(x in text for x in ["negative_event_drop", "large_negative_drop", "drop_to_zero", "critical_student_count_anomaly"]):
        return "rapid_decline_or_closure_candidate"
    return "other_event_school"


def top_reasons(series: pd.Series, n: int = 3) -> str:
    vals: list[str] = []
    for item in series.dropna().astype(str):
        for part in item.split(","):
            part = part.strip()
            if part:
                vals.append(part)
    if not vals:
        return ""
    counts = pd.Series(vals).value_counts().head(n)
    return "; ".join(f"{k}:{v}" for k, v in counts.items())


def md_table(df: pd.DataFrame, n: int = 20) -> str:
    if df.empty:
        return "_No rows._"
    d = df.head(n).astype(object).where(pd.notna(df.head(n)), "")
    lines = [
        "| " + " | ".join(map(str, d.columns)) + " |",
        "| " + " | ".join(["---"] * len(d.columns)) + " |",
    ]
    for _, row in d.iterrows():
        lines.append("| " + " | ".join(str(row[c]).replace("|", "/").replace("\n", " ") for c in d.columns) + " |")
    return "\n".join(lines)


def style_xlsx(path: Path) -> None:
    wb = load_workbook(path)
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            max_len = max(len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, min(ws.max_row, 80) + 1))
            ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 50)
        ws.auto_filter.ref = ws.dimensions
    wb.save(path)


def main() -> None:
    ensure_dirs()
    run_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    audit = pd.read_csv(INPUT / "audit" / "event_exclusion_policy_audit.csv", low_memory=False)
    policy = pd.read_csv(INPUT / "results" / "policy_comparison_summary.csv", low_memory=False)
    panel = pd.read_csv(PATCH / "canonical" / "school_year_panel.csv", low_memory=False)

    latest = panel.sort_values(["school_key", "year"]).groupby("school_key").tail(1)[
        ["school_key", "year", "student_count", "status", "size_bucket"]
    ].rename(columns={
        "year": "last_observed_year",
        "student_count": "latest_student_count",
        "status": "latest_status",
        "size_bucket": "latest_size_bucket",
    })
    first = panel.groupby("school_key", as_index=False)["year"].min().rename(columns={"year": "first_observed_year"})
    total_p0 = panel["school_key"].nunique()

    excluded = audit[audit["exclude_p1_school_level"].fillna(False).astype(bool)].copy()
    excluded = excluded.merge(first, on="school_key", how="left").merge(latest, on="school_key", how="left")
    excluded["event_layer_category"] = excluded["event_flags"].apply(categorize)
    required = [
        "school_key", "school_name", "sido", "sgg", "school_level", "years_observed", "event_flags", "event_years",
        "max_positive_delta", "max_negative_delta", "max_pct_change", "exclusion_reason", "recommended_action",
        "first_observed_year", "last_observed_year", "latest_student_count", "latest_status", "latest_size_bucket",
        "event_layer_category",
    ]
    excluded = excluded[required].sort_values(["sido", "sgg", "school_level", "school_name"])

    p1_excluded_school_count = len(excluded)
    # policy comparison has repeated rows by stage/model; use horizon max to avoid double counting.
    row_share_by_h = {}
    row_counts_by_h = {}
    for h, g in policy.groupby("horizon"):
        p0_rows = int(g["p0_rows"].max())
        p1_rows = int(g["p1_rows"].max())
        ex_rows = p0_rows - p1_rows
        row_counts_by_h[int(h)] = ex_rows
        row_share_by_h[int(h)] = ex_rows / p0_rows if p0_rows else np.nan
    p1_excluded_row_count = int(sum(row_counts_by_h.values()))

    reason_summary = excluded.groupby(["exclusion_reason", "event_layer_category"], dropna=False).agg(
        school_count=("school_key", "nunique"),
        example_schools=("school_name", lambda x: ", ".join(x.dropna().astype(str).head(5))),
    ).reset_index()
    reason_summary["row_count_if_available"] = np.nan
    reason_summary["share_of_excluded_schools"] = reason_summary["school_count"] / p1_excluded_school_count
    reason_summary = reason_summary[[
        "exclusion_reason", "event_layer_category", "school_count", "row_count_if_available",
        "share_of_excluded_schools", "example_schools",
    ]].sort_values("school_count", ascending=False)

    total_by_region_level = panel.groupby(["sido", "school_level"])["school_key"].nunique().reset_index(name="total_p0_school_count")
    excl_region = excluded.groupby(["sido", "school_level"]).agg(
        excluded_school_count=("school_key", "nunique"),
        top_exclusion_reasons=("event_flags", top_reasons),
    ).reset_index()
    by_region = total_by_region_level.merge(excl_region, on=["sido", "school_level"], how="left")
    by_region["excluded_school_count"] = by_region["excluded_school_count"].fillna(0).astype(int)
    by_region["top_exclusion_reasons"] = by_region["top_exclusion_reasons"].fillna("")
    by_region["excluded_share"] = by_region["excluded_school_count"] / by_region["total_p0_school_count"]
    by_region = by_region[["sido", "school_level", "excluded_school_count", "total_p0_school_count", "excluded_share", "top_exclusion_reasons"]].sort_values(["excluded_share", "excluded_school_count"], ascending=False)

    total_by_sgg = panel.groupby(["sido", "sgg"])["school_key"].nunique().reset_index(name="total_p0_school_count")
    excl_sgg = excluded.groupby(["sido", "sgg"]).agg(
        excluded_school_count=("school_key", "nunique"),
        top_exclusion_reasons=("event_flags", top_reasons),
    ).reset_index()
    by_sgg = total_by_sgg.merge(excl_sgg, on=["sido", "sgg"], how="left")
    by_sgg["excluded_school_count"] = by_sgg["excluded_school_count"].fillna(0).astype(int)
    by_sgg["top_exclusion_reasons"] = by_sgg["top_exclusion_reasons"].fillna("")
    by_sgg["excluded_share"] = by_sgg["excluded_school_count"] / by_sgg["total_p0_school_count"]
    by_sgg = by_sgg[["sido", "sgg", "excluded_school_count", "total_p0_school_count", "excluded_share", "top_exclusion_reasons"]].sort_values(["excluded_share", "excluded_school_count"], ascending=False)

    excluded_school_share = p1_excluded_school_count / total_p0 if total_p0 else np.nan
    if excluded_school_share >= 0.15:
        interp = "high warning: P1 excludes at least 15% of schools, so final reporting must be framed as event-excluded existing-school decline pressure."
    elif excluded_school_share >= 0.10:
        interp = "medium warning: P1 excludes at least 10% of schools; acceptable for decline-pressure modeling if event layer is reported separately."
    else:
        interp = "low warning: exclusion share is below 10%."
    scale = pd.DataFrame([{
        "p1_excluded_school_count": p1_excluded_school_count,
        "p1_excluded_row_count": p1_excluded_row_count,
        "p0_unique_school_count": total_p0,
        "p1_unique_school_count": total_p0 - p1_excluded_school_count,
        "excluded_school_share": excluded_school_share,
        "excluded_row_share_by_horizon_1yr": row_share_by_h.get(1, np.nan),
        "excluded_row_share_by_horizon_2yr": row_share_by_h.get(2, np.nan),
        "excluded_row_share_by_horizon_3yr": row_share_by_h.get(3, np.nan),
        "excluded_row_share_by_horizon_4yr": row_share_by_h.get(4, np.nan),
        "excluded_row_share_by_horizon_5yr": row_share_by_h.get(5, np.nan),
        "interpretation": interp,
        "recommended_policy": "Use P1 only as event-excluded existing-school decline-pressure layer; keep excluded schools as separate event layer.",
    }])

    top_reason = reason_summary.iloc[0]["event_layer_category"] if not reason_summary.empty else ""
    summary = pd.DataFrame([{
        "run_time": run_time,
        "p1_excluded_school_count": p1_excluded_school_count,
        "p1_excluded_row_count": p1_excluded_row_count,
        "p0_unique_school_count": total_p0,
        "excluded_school_share": excluded_school_share,
        "top_exclusion_reason": top_reason,
        "report_created": True,
    }])

    excluded.to_csv(OUT / "p1_excluded_schools_2173.csv", index=False, encoding="utf-8-sig")
    reason_summary.to_csv(OUT / "p1_exclusion_reason_summary.csv", index=False, encoding="utf-8-sig")
    by_region.to_csv(OUT / "p1_excluded_by_region_level.csv", index=False, encoding="utf-8-sig")
    by_sgg.to_csv(OUT / "p1_excluded_by_sgg.csv", index=False, encoding="utf-8-sig")
    scale.to_csv(OUT / "p1_exclusion_scale_audit.csv", index=False, encoding="utf-8-sig")

    report_text = "\n".join([
        "# V5 P1 Excluded School List v1",
        "",
        "## Summary",
        md_table(summary),
        "",
        "## P1 Excluded School List",
        f"P1 excludes {p1_excluded_school_count:,} schools. These schools should be managed as a separate event layer, not silently mixed into the decline-pressure model.",
        md_table(excluded, 20),
        "",
        "## Exclusion Reason Summary",
        md_table(reason_summary, 30),
        "",
        "## Region and School Level Summary",
        md_table(by_region, 30),
        "",
        "## Exclusion Scale Audit",
        md_table(scale),
        "",
        "## Interpretation for Decline Pressure Project",
        interp,
        "",
        "## Recommended Scenario Handling",
        "Final scenario reports should explicitly say that P1 represents event-excluded existing-school decline pressure. The excluded schools should be displayed or reviewed as a separate event layer.",
    ])
    (REPORT / "00_COMBINED_REPORT.md").write_text(report_text, encoding="utf-8")

    xlsx = REPORT / "01_KEY_TABLES.xlsx"
    with pd.ExcelWriter(xlsx, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="summary", index=False)
        excluded.to_excel(writer, sheet_name="p1_excluded_schools", index=False)
        reason_summary.to_excel(writer, sheet_name="exclusion_reason_summary", index=False)
        by_region.to_excel(writer, sheet_name="excluded_by_region_level", index=False)
        by_sgg.to_excel(writer, sheet_name="excluded_by_sgg", index=False)
        scale.to_excel(writer, sheet_name="exclusion_scale_audit", index=False)
    style_xlsx(xlsx)

    manifest = HANDOFF / "MANIFEST.md"
    copied = []
    missing = []
    manifest.write_text("\n".join([
        "# V5 P1 Excluded School List v1 Handoff",
        "",
        f"- run_time: {run_time}",
        f"- project_root: {ROOT}",
        f"- data_output_path: {OUT}",
        f"- report_output_path: {REPORT}",
        f"- p1_excluded_school_count: {p1_excluded_school_count}",
        f"- p1_excluded_row_count: {p1_excluded_row_count}",
        f"- excluded_school_share: {excluded_school_share}",
        f"- top_exclusion_reason: {top_reason}",
        "- handoff_file_count: 5",
    ]), encoding="utf-8")
    copied.append({"file": "MANIFEST.md", "source": "generated"})
    for src in [REPORT / "00_COMBINED_REPORT.md", REPORT / "01_KEY_TABLES.xlsx"]:
        if src.exists():
            shutil.copy2(src, HANDOFF / src.name)
            copied.append({"file": src.name, "source": rel(src)})
        else:
            missing.append({"file": src.name, "source": rel(src)})
    pd.DataFrame(copied).to_csv(HANDOFF / "copied_files_manifest.csv", index=False, encoding="utf-8-sig")
    pd.DataFrame(missing, columns=["file", "source"]).to_csv(HANDOFF / "missing_files.csv", index=False, encoding="utf-8-sig")
    handoff_exact = len(list(HANDOFF.iterdir())) == 5

    print("V5_extract_p1_excluded_school_list_v1 completed.")
    print("\nKey results:\n")
    print(f"* p1_excluded_school_count: {p1_excluded_school_count}")
    print(f"* p1_excluded_row_count: {p1_excluded_row_count}")
    print(f"* excluded_school_share: {excluded_school_share:.6f}")
    print(f"* top_exclusion_reason: {top_reason}")
    print("* report_created: True")
    print(f"* handoff_exactly_5_files: {handoff_exact}")
    print("\nRecommended next step:\n")
    print("* Upload the 5 handoff files to ChatGPT to review whether P1 exclusion is too broad before final scenario generation.")


if __name__ == "__main__":
    main()
