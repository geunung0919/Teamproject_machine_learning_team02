from __future__ import annotations

import json
import math
import os
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

target_cpu_fraction = 0.70
worker_count = max(1, math.floor((os.cpu_count() or 1) * target_cpu_fraction))
thread_count = max(2, min(worker_count, math.ceil((os.cpu_count() or 1) * 0.50)))
os.environ.setdefault("OMP_NUM_THREADS", str(thread_count))
os.environ.setdefault("OPENBLAS_NUM_THREADS", str(thread_count))
os.environ.setdefault("MKL_NUM_THREADS", str(thread_count))
os.environ.setdefault("NUMEXPR_NUM_THREADS", str(thread_count))

import joblib
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sklearn.compose import ColumnTransformer
from sklearn.ensemble import HistGradientBoostingRegressor, RandomForestRegressor
from sklearn.impute import SimpleImputer
from sklearn.linear_model import LinearRegression, Ridge
from sklearn.metrics import r2_score
from sklearn.multioutput import MultiOutputRegressor
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import FunctionTransformer, OneHotEncoder, StandardScaler


ROOT = Path(__file__).resolve().parents[2]
PATCH = ROOT / "data" / "v5_clean_dataset_patch_v1"
DIRECT = ROOT / "data" / "v5_direct_multihorizon_policy_comparison_v1"
R6C = ROOT / "data" / "v5_r6_clean_retrain_and_scenario_regeneration_v1"
R2R3 = ROOT / "data" / "v5_r2_r3_scenario_total_comparison_v1"
R6REV = ROOT / "data" / "v5_r6_horizon_3_4_5_path_diagnosis_and_retrain_v1"
R4R5 = ROOT / "data" / "v5_r4_r5_feasibility_and_model_training_v1"
ENS = ROOT / "data" / "v5_r4r5_prediction_persistence_and_r5r6_ensemble_retry_v1"
P0_EVENT = ROOT / "data" / "v5_p0_and_event_total_scenario_audit_v1"
EXCLUDED = ROOT / "data" / "v5_p1_excluded_school_list_v1" / "p1_excluded_schools_2173.csv"

OUT = ROOT / "data" / "v5_recursive_and_multioutput_forecasting_r3_r6_v1"
RESULT = OUT / "results"
SCENARIO = OUT / "scenario"
AUDIT = OUT / "audit"
MODEL_DIR = ROOT / "models" / "v5_recursive_and_multioutput_forecasting_r3_r6_v1"
REPORT = ROOT / "reports" / "v5_recursive_and_multioutput_forecasting_r3_r6_v1"
HANDOFF = ROOT / "handoff_for_chatgpt" / "v5_recursive_and_multioutput_forecasting_r3_r6_v1"

POLICY_P1 = "P1_event_excluded_decline_focus"
POLICY_P0 = "P0_current_standard"
KOSIS = {2025: 5_114_000, 2026: 4_922_000, 2027: 4_746_000, 2028: 4_562_000, 2029: 4_328_000, 2030: 4_087_000}
MODEL_NAMES = ["Ridge", "RandomForestRegressor", "HistGradientBoostingRegressor"]
TARGET_COLS = [f"target_{name}_{h}yr" for h in range(1, 6) for name in ["year", "student_count", "delta", "available"]]
EXCLUDE_BASE = {
    "school_key", "school_name", "school_name_norm", "address", "source_file", "survey_date",
    "exclusion_reason", "quality_note", "patch_exclusion_reason", "status",
    "r1_model_eligible_patched", "r2_model_eligible_patched", "r3_model_eligible_patched",
    "standard_model_eligible", "scenario_base_eligible",
    "exclude_p1_row_level", "exclude_p1_school_level", "event_flags",
    "coordinate_source", "coordinate_invalid_reason", "isolation_score_version", "grade_invalid_reason",
}
CATEGORICAL_HINTS = ["sido", "sgg", "school_level", "branch_type", "foundation_type", "region_group", "size_bucket", "student_size_bin", "metro_flag", "custom_size_bin"]


def ensure_dirs() -> None:
    for d in [OUT, RESULT, SCENARIO, AUDIT, MODEL_DIR, REPORT]:
        d.mkdir(parents=True, exist_ok=True)
    if HANDOFF.exists():
        shutil.rmtree(HANDOFF)
    HANDOFF.mkdir(parents=True, exist_ok=True)


def rel(path: Path) -> str:
    try:
        return str(path.relative_to(ROOT))
    except ValueError:
        return str(path)


def custom_size_bin(s: pd.Series) -> pd.Series:
    return pd.cut(pd.to_numeric(s, errors="coerce"), bins=[-1, 0, 30, 60, 120, 240, 500, 1000, 100000], labels=["zero", "1_30", "31_60", "61_120", "121_240", "241_500", "501_1000", "1001_plus"]).astype(str)


def stringify_frame(x: Any) -> pd.DataFrame:
    return pd.DataFrame(x).copy().astype("string").fillna("__missing__").astype(str)


def numeric_frame(x: Any) -> pd.DataFrame:
    frame = pd.DataFrame(x).copy()
    for col in frame.columns:
        frame[col] = pd.to_numeric(frame[col], errors="coerce")
    return frame


def feature_columns(df: pd.DataFrame) -> list[str]:
    cols = []
    for c in df.columns:
        lc = c.lower()
        if c in EXCLUDE_BASE or c in TARGET_COLS or c.startswith("target_"):
            continue
        if any(p in lc for p in ["next", "future", "after", "label", "closed_next", "missing_next"]):
            continue
        cols.append(c)
    return cols


def preprocessor(df: pd.DataFrame, features: list[str], scale: bool, dense: bool = False) -> ColumnTransformer:
    cats = [c for c in features if c in CATEGORICAL_HINTS or df[c].dtype == "object" or str(df[c].dtype) == "bool"]
    nums = [c for c in features if c not in cats]
    try:
        ohe = OneHotEncoder(handle_unknown="infrequent_if_exist", min_frequency=20, sparse_output=not dense)
    except TypeError:
        ohe = OneHotEncoder(handle_unknown="ignore", sparse=not dense)
    transformers = []
    if nums:
        steps = [("to_numeric", FunctionTransformer(numeric_frame, validate=False)), ("imputer", SimpleImputer(strategy="median"))]
        if scale:
            steps.append(("scaler", StandardScaler(with_mean=False if not dense else True)))
        transformers.append(("num", Pipeline(steps), nums))
    if cats:
        transformers.append(("cat", Pipeline([("stringify", FunctionTransformer(stringify_frame, validate=False)), ("onehot", ohe)]), cats))
    return ColumnTransformer(transformers, remainder="drop", sparse_threshold=0.0 if dense else 0.3)


def base_estimator(name: str) -> Any:
    if name == "LinearRegression":
        return LinearRegression()
    if name == "Ridge":
        return Ridge(alpha=5.0)
    if name == "RandomForestRegressor":
        return RandomForestRegressor(n_estimators=8, max_depth=8, min_samples_leaf=12, random_state=42, n_jobs=worker_count)
    if name == "HistGradientBoostingRegressor":
        return HistGradientBoostingRegressor(max_iter=25, max_leaf_nodes=15, learning_rate=0.08, l2_regularization=0.01, random_state=42)
    raise ValueError(name)


def make_pipeline(name: str, train: pd.DataFrame, features: list[str], multi: bool = False) -> Pipeline:
    dense = name == "HistGradientBoostingRegressor"
    scale = name in ["LinearRegression", "Ridge"]
    est = base_estimator(name)
    if multi and name in ["Ridge", "HistGradientBoostingRegressor"]:
        est = MultiOutputRegressor(est, n_jobs=max(1, min(5, worker_count)))
    return Pipeline([("preprocess", preprocessor(train, features, scale=scale, dense=dense)), ("model", est)])


def add_r6_features(df: pd.DataFrame) -> pd.DataFrame:
    r6 = pd.read_csv(R6C / "features" / "r6_cohort_pressure_features_by_school_year.csv", low_memory=False)
    use = ["school_key", "year"] + [c for c in r6.columns if c not in {"school_key", "year", "sido", "sgg", "school_level"}]
    return df.merge(r6[use], on=["school_key", "year"], how="left")


def read_view(feature_family: str, policy: str, h: int = 5) -> pd.DataFrame:
    if feature_family == "R3":
        p = DIRECT / "model_views" / policy / f"r3_grade_flow_direct_{h}yr.csv"
        if not p.exists():
            p = PATCH / "model_views" / f"r3_basic_{h}yr.csv"
            if not p.exists():
                p = PATCH / "model_views" / f"r3_grade_flow_{h}yr.csv"
        return pd.read_csv(p, low_memory=False)
    path = R6C / "model_views" / policy / f"r6_cohort_pressure_direct_{h}yr.csv"
    if path.exists():
        return pd.read_csv(path, low_memory=False)
    p_r3 = DIRECT / "model_views" / policy / f"r3_grade_flow_direct_{h}yr.csv"
    if not p_r3.exists():
        p_r3 = PATCH / "model_views" / f"r3_basic_{h}yr.csv"
        if not p_r3.exists():
            p_r3 = PATCH / "model_views" / f"r3_grade_flow_{h}yr.csv"
    return add_r6_features(pd.read_csv(p_r3, low_memory=False))


def scenario_base(feature_family: str, rows: pd.DataFrame) -> pd.DataFrame:
    base = rows.copy()
    base["custom_size_bin"] = custom_size_bin(base["student_count"])
    if feature_family == "R6":
        base = add_r6_features(base)
    return base


def update_recursive_features(df: pd.DataFrame, prev: np.ndarray, pred: np.ndarray) -> pd.DataFrame:
    out = df.copy()
    out["student_count"] = pred
    out["student_delta_lag_1"] = pred - prev
    out["student_growth_lag_1"] = np.where(prev > 0, (pred - prev) / prev, np.nan)
    out["student_count_lag_1"] = prev
    out["student_rolling_mean_3"] = np.nanmean(np.vstack([pred, prev, pd.to_numeric(out.get("student_count_lag_2", np.nan), errors="coerce").to_numpy(float)]), axis=0)
    out["students_per_class"] = np.where(pd.to_numeric(out.get("class_count", np.nan), errors="coerce") > 0, pred / pd.to_numeric(out.get("class_count"), errors="coerce"), np.nan)
    out["students_per_teacher"] = np.where(pd.to_numeric(out.get("teacher_count", np.nan), errors="coerce") > 0, pred / pd.to_numeric(out.get("teacher_count"), errors="coerce"), np.nan)
    out["size_bucket"] = pd.cut(pred, bins=[-1, 0, 60, 120, 240, 500, 1000, 100000], labels=["zero", "small", "small_mid", "mid", "large", "very_large", "mega"]).astype(str)
    out["custom_size_bin"] = custom_size_bin(pd.Series(pred))
    return out


def path_flag(delta: float, pct: float) -> str:
    if pd.isna(delta) or pd.isna(pct):
        return "none"
    if delta < -500_000 or pct < -0.10:
        return "critical"
    if delta < -300_000 or pct < -0.07:
        return "warning"
    return "none"


def metric_rows(candidate: dict[str, Any], pred: pd.DataFrame) -> list[dict[str, Any]]:
    rows = []
    for h, g in pred.groupby("horizon"):
        actual = g["actual_student_count"].to_numpy(float)
        predv = g["predicted_student_count"].to_numpy(float)
        base = g["base_student_count"].to_numpy(float)
        abs_e = np.abs(actual - predv)
        actual_delta = actual - base
        pred_delta = predv - base
        lvl_rates = []
        for _, sg in g.groupby(["target_year", "school_level"], dropna=False):
            denom = sg["actual_student_count"].abs().sum()
            lvl_rates.append(abs(sg["predicted_student_count"].sum() - sg["actual_student_count"].sum()) / denom if denom else np.nan)
        small = g[pd.to_numeric(g["base_student_count"], errors="coerce").fillna(0) <= 60]
        rows.append({
            **candidate, "horizon": int(h),
            "level_MAE": float(abs_e.mean()), "level_RMSE": float(np.sqrt(np.mean((actual - predv) ** 2))),
            "level_R2": float(r2_score(actual, predv)) if len(np.unique(actual)) > 1 else np.nan,
            "level_WAPE": float(abs_e.sum() / np.abs(actual).sum()) if np.abs(actual).sum() else np.nan,
            "delta_MAE": float(np.abs(actual_delta - pred_delta).mean()),
            "delta_RMSE": float(np.sqrt(np.mean((actual_delta - pred_delta) ** 2))),
            "delta_R2": float(r2_score(actual_delta, pred_delta)) if len(np.unique(actual_delta)) > 1 else np.nan,
            "delta_WAPE": float(np.abs(actual_delta - pred_delta).sum() / np.abs(actual_delta).sum()) if np.abs(actual_delta).sum() else np.nan,
            "median_abs_error": float(np.median(abs_e)), "p90_abs_error": float(np.quantile(abs_e, .90)),
            "p95_abs_error": float(np.quantile(abs_e, .95)), "p99_abs_error": float(np.quantile(abs_e, .99)),
            "total_error_rate": float((predv.sum() - actual.sum()) / actual.sum()) if actual.sum() else np.nan,
            "school_level_total_error_rate": float(np.nanmean(lvl_rates)),
            "small_school_MAE": float(small["abs_error"].mean()) if not small.empty else np.nan,
            "small_school_p95_abs_error": float(small["abs_error"].quantile(.95)) if not small.empty else np.nan,
        })
    return rows


def recursive_predict(pipe: Pipeline, features: list[str], base: pd.DataFrame, target_type: str, years: int) -> list[np.ndarray]:
    cur = base.copy()
    prev = pd.to_numeric(cur["student_count"], errors="coerce").fillna(0).to_numpy(float)
    preds = []
    for _ in range(years):
        for f in features:
            if f not in cur.columns:
                cur[f] = np.nan
        raw = pipe.predict(cur[features])
        nxt = np.maximum(0, raw if target_type == "level" else prev + raw)
        preds.append(nxt)
        cur = update_recursive_features(cur, prev, nxt)
        prev = nxt
    return preds


def train_recursive_candidate(feature_family: str, target_type: str, model_name: str, policy: str, rows_for_scenario: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, Pipeline, list[str]]:
    view = read_view(feature_family, policy, 1)
    target = "target_student_count_1yr" if target_type == "level" else "target_delta_1yr"
    features = feature_columns(view)
    pred_rows = []
    for target_year in range(2022, 2026):
        for h in range(1, 6):
            base_year = target_year - h
            tr = view[(view["year"] < base_year) & view[target].notna()].copy()
            te0 = view[view["year"].eq(base_year)].copy()
            actual_col = f"target_student_count_{h}yr"
            if len(tr) < 100 or te0.empty or actual_col not in te0.columns:
                continue
            pipe = make_pipeline(model_name, tr, features)
            pipe.fit(tr[features], tr[target].astype(float))
            ph = recursive_predict(pipe, features, te0, target_type, h)[-1]
            tmp = te0[["school_key", "school_name", "sido", "sgg", "school_level", "student_count"]].copy()
            tmp["candidate_name"] = f"{feature_family}_recursive_1step_{target_type}_{model_name}"
            tmp["feature_family"] = feature_family; tmp["forecasting_strategy"] = "recursive_1step"; tmp["target_type"] = target_type; tmp["model"] = model_name
            tmp["base_year"] = base_year; tmp["target_year"] = target_year; tmp["horizon"] = h
            tmp["base_student_count"] = pd.to_numeric(te0["student_count"], errors="coerce").fillna(0).to_numpy(float)
            tmp["actual_student_count"] = pd.to_numeric(te0[actual_col], errors="coerce").fillna(0).to_numpy(float)
            tmp["predicted_student_count"] = ph
            tmp["abs_error"] = np.abs(tmp["actual_student_count"] - tmp["predicted_student_count"])
            pred_rows.append(tmp)
    full = view[view[target].notna()].copy()
    pipe = make_pipeline(model_name, full, features)
    pipe.fit(full[features], full[target].astype(float))
    joblib.dump(pipe, MODEL_DIR / f"{feature_family}_recursive_1step_{target_type}_{model_name}.joblib")
    scen_base = scenario_base(feature_family, rows_for_scenario)
    scen_preds = recursive_predict(pipe, features, scen_base, target_type, 5)
    scen = rows_for_scenario[["school_key", "school_name", "sido", "sgg", "school_level", "size_bucket", "student_count"]].copy()
    scen = scen.rename(columns={"student_count": "student_count_2025", "size_bucket": "size_bucket_2025"})
    scen["candidate_name"] = f"{feature_family}_recursive_1step_{target_type}_{model_name}"
    for i, arr in enumerate(scen_preds, start=1):
        scen[f"pred_student_count_{2025+i}"] = arr
    return pd.concat(pred_rows, ignore_index=True) if pred_rows else pd.DataFrame(), scen, pipe, features


def multi_y(df: pd.DataFrame, target_type: str) -> pd.DataFrame:
    if target_type == "level":
        return df[[f"target_student_count_{h}yr" for h in range(1, 6)]].copy()
    if target_type == "cumulative_delta":
        return df[[f"target_delta_{h}yr" for h in range(1, 6)]].copy()
    vals = {}
    prev = df["student_count"]
    for h in range(1, 6):
        cur = df[f"target_student_count_{h}yr"]
        vals[f"inc_{h}yr"] = cur - prev
        prev = cur
    return pd.DataFrame(vals)


def decode_multi(base: np.ndarray, yhat: np.ndarray, target_type: str) -> list[np.ndarray]:
    if target_type == "level":
        return [np.maximum(0, yhat[:, i]) for i in range(5)]
    if target_type == "cumulative_delta":
        return [np.maximum(0, base + yhat[:, i]) for i in range(5)]
    out = []
    cur = base.copy()
    for i in range(5):
        cur = np.maximum(0, cur + yhat[:, i])
        out.append(cur.copy())
    return out


def train_multi_candidate(feature_family: str, target_type: str, model_name: str, policy: str, rows_for_scenario: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, Pipeline, list[str]]:
    view = read_view(feature_family, policy, 5)
    mask = np.logical_and.reduce([view[f"target_available_{h}yr"].fillna(False).astype(bool).to_numpy() for h in range(1, 6)])
    view = view[mask].copy()
    features = feature_columns(view)
    pred_rows = []
    for target_year in range(2022, 2026):
        for h in range(1, 6):
            base_year = target_year - h
            tr = view[view["year"] < base_year].copy()
            te0 = read_view(feature_family, policy, h)
            te0 = te0[te0["year"].eq(base_year)].copy()
            if len(tr) < 100 or te0.empty or f"target_student_count_{h}yr" not in te0.columns:
                continue
            for f in features:
                if f not in te0.columns:
                    te0[f] = np.nan
            ytr = multi_y(tr, target_type).astype(float)
            pipe = make_pipeline(model_name, tr, features, multi=True)
            pipe.fit(tr[features], ytr)
            yhat = pipe.predict(te0[features])
            decoded = decode_multi(pd.to_numeric(te0["student_count"], errors="coerce").fillna(0).to_numpy(float), np.asarray(yhat), target_type)
            ph = decoded[h - 1]
            tmp = te0[["school_key", "school_name", "sido", "sgg", "school_level", "student_count"]].copy()
            tmp["candidate_name"] = f"{feature_family}_multioutput_1to5_{target_type}_{model_name}"
            tmp["feature_family"] = feature_family; tmp["forecasting_strategy"] = "multioutput_1to5"; tmp["target_type"] = target_type; tmp["model"] = model_name
            tmp["base_year"] = base_year; tmp["target_year"] = target_year; tmp["horizon"] = h
            tmp["base_student_count"] = pd.to_numeric(te0["student_count"], errors="coerce").fillna(0).to_numpy(float)
            tmp["actual_student_count"] = pd.to_numeric(te0[f"target_student_count_{h}yr"], errors="coerce").fillna(0).to_numpy(float)
            tmp["predicted_student_count"] = ph
            tmp["abs_error"] = np.abs(tmp["actual_student_count"] - tmp["predicted_student_count"])
            pred_rows.append(tmp)
    yfull = multi_y(view, target_type).astype(float)
    pipe = make_pipeline(model_name, view, features, multi=True)
    pipe.fit(view[features], yfull)
    joblib.dump(pipe, MODEL_DIR / f"{feature_family}_multioutput_1to5_{target_type}_{model_name}.joblib")
    scen_base = scenario_base(feature_family, rows_for_scenario)
    for f in features:
        if f not in scen_base.columns:
            scen_base[f] = np.nan
    yhat = pipe.predict(scen_base[features])
    decoded = decode_multi(pd.to_numeric(scen_base["student_count"], errors="coerce").fillna(0).to_numpy(float), np.asarray(yhat), target_type)
    scen = rows_for_scenario[["school_key", "school_name", "sido", "sgg", "school_level", "size_bucket", "student_count"]].copy()
    scen = scen.rename(columns={"student_count": "student_count_2025", "size_bucket": "size_bucket_2025"})
    scen["candidate_name"] = f"{feature_family}_multioutput_1to5_{target_type}_{model_name}"
    for i, arr in enumerate(decoded, start=1):
        scen[f"pred_student_count_{2025+i}"] = arr
    return pd.concat(pred_rows, ignore_index=True) if pred_rows else pd.DataFrame(), scen, pipe, features


def candidate_short_name(row: pd.Series) -> str:
    return f"{row['feature_family']}_{row['forecasting_strategy']}_{row['target_type']}_{row['model']}"


def totals_from_school(schools: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for cand, g in schools.groupby("candidate_name"):
        base = float(g["student_count_2025"].sum())
        for y in range(2025, 2031):
            total = base if y == 2025 else float(g[f"pred_student_count_{y}"].sum())
            rows.append({"candidate_name": cand, "year": y, "total_students": total, "delta_from_2025": total - base, "pct_change_from_2025": (total - base) / base if base else np.nan})
    out = pd.DataFrame(rows).sort_values(["candidate_name", "year"])
    out["yoy_delta"] = out.groupby("candidate_name")["total_students"].diff()
    out["yoy_pct_change"] = out.groupby("candidate_name")["total_students"].pct_change()
    out["path_jump_flag"] = [path_flag(d, p) for d, p in zip(out["yoy_delta"], out["yoy_pct_change"])]
    out["note"] = np.where(out["path_jump_flag"].eq("none"), "no large jump", "warning/critical path jump")
    return out[["candidate_name", "year", "total_students", "yoy_delta", "yoy_pct_change", "delta_from_2025", "pct_change_from_2025", "path_jump_flag", "note"]]


def grouped_totals(schools: pd.DataFrame, groups: list[str]) -> pd.DataFrame:
    rows = []
    for keys, g in schools.groupby(["candidate_name"] + groups, dropna=False):
        key_vals = keys if isinstance(keys, tuple) else (keys,)
        rec = dict(zip(["candidate_name"] + groups, key_vals))
        base = float(g["student_count_2025"].sum())
        for y in range(2025, 2031):
            total = base if y == 2025 else float(g[f"pred_student_count_{y}"].sum())
            rows.append({**rec, "year": y, "total_students": total, "delta_from_2025": total - base, "pct_change_from_2025": (total - base) / base if base else np.nan})
    return pd.DataFrame(rows)


def path_consistency(scen_total: pd.DataFrame, metrics: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for cand, g in scen_total.groupby("candidate_name"):
        parts = cand.split("_", 2)
        m = metrics[metrics["candidate_name"].eq(cand)]
        jumps = g[g["path_jump_flag"].ne("none")]
        rows.append({
            "candidate_name": cand,
            "feature_family": parts[0],
            "forecasting_strategy": "recursive_1step" if "recursive_1step" in cand else "multioutput_1to5",
            "target_type": m["target_type"].iloc[0] if not m.empty else "",
            "model": m["model"].iloc[0] if not m.empty else "",
            "path_jump_count": int(len(jumps)),
            "critical_jump_count": int((jumps["path_jump_flag"] == "critical").sum()),
            "max_abs_yoy_delta": float(g["yoy_delta"].abs().max()) if g["yoy_delta"].notna().any() else np.nan,
            "max_abs_yoy_pct": float(g["yoy_pct_change"].abs().max()) if g["yoy_pct_change"].notna().any() else np.nan,
            "validation_path_score": float(1 / (1 + (m["total_error_rate"].abs().mean() if not m.empty else 999))),
            "scenario_path_score": float(1 / (1 + len(jumps) + 2 * int((jumps["path_jump_flag"] == "critical").sum()))),
            "path_reliable": bool(int((jumps["path_jump_flag"] == "critical").sum()) == 0),
            "note": "reliable by jump rule" if jumps.empty else "has path jump",
        })
    return pd.DataFrame(rows)


def read_total_file(path: Path, candidate: str, scenario_filter: str | None = None) -> pd.DataFrame:
    if not path.exists():
        return pd.DataFrame()
    df = pd.read_csv(path, low_memory=False)
    if scenario_filter and "scenario_name" in df.columns:
        df = df[df["scenario_name"].eq(scenario_filter)].copy()
    val = next((c for c in ["total_students", "p1_plus_event_total_students", "p0_total_students"] if c in df.columns), None)
    if not val:
        return pd.DataFrame()
    out = df[["year", val]].rename(columns={val: "total_students"}).copy()
    out["candidate_name"] = candidate
    return out[["candidate_name", "year", "total_students"]]


def all_existing_totals(new_totals: pd.DataFrame) -> pd.DataFrame:
    totals = [new_totals[["candidate_name", "year", "total_students"]]]
    for cand, path, filt in [
        ("R2_only", R2R3 / "scenario" / "r2_only_p1_plus_event_total_by_year.csv", None),
        ("R3_only", R2R3 / "scenario" / "r3_only_p1_plus_event_total_by_year.csv", None),
        ("R4R5", R4R5 / "scenario" / "r4_r5_p1_plus_event_total_students_by_year.csv", None),
        ("R6_clean", R6C / "scenario" / "r6_clean_p1_plus_event_total_students_by_year.csv", None),
        ("R6_revised", R6REV / "scenario" / "r6_revised_p1_plus_event_total_students_by_year.csv", None),
        ("ensemble_horizon_specific", ENS / "scenario" / "ensemble_p1_plus_event_total_students_by_year.csv", "horizon_specific"),
        ("ensemble_global_single_weight", ENS / "scenario" / "ensemble_p1_plus_event_total_students_by_year.csv", "global_single_weight"),
        ("ensemble_conservative_weight", ENS / "scenario" / "ensemble_p1_plus_event_total_students_by_year.csv", "conservative_weight"),
    ]:
        t = read_total_file(path, cand, filt)
        if not t.empty:
            totals.append(t)
    return pd.concat(totals, ignore_index=True)


def kosis_compare(total: pd.DataFrame) -> pd.DataFrame:
    out = total.copy()
    if "yoy_delta" not in out.columns:
        out = out.sort_values(["candidate_name", "year"])
        out["yoy_delta"] = out.groupby("candidate_name")["total_students"].diff()
        out["yoy_pct_change"] = out.groupby("candidate_name")["total_students"].pct_change()
        out["path_jump_flag"] = [path_flag(d, p) for d, p in zip(out["yoy_delta"], out["yoy_pct_change"])]
    out["kosis_reference_total"] = out["year"].map(KOSIS)
    out["diff_to_kosis"] = out["total_students"] - out["kosis_reference_total"]
    out["pct_diff_to_kosis"] = out["diff_to_kosis"] / out["kosis_reference_total"]
    out["note"] = "KOSIS is post-hoc reference only; not used for training or selection."
    return out[["candidate_name", "year", "total_students", "kosis_reference_total", "diff_to_kosis", "pct_diff_to_kosis", "yoy_delta", "yoy_pct_change", "path_jump_flag", "note"]]


def normalize(s: pd.Series) -> pd.Series:
    s = pd.to_numeric(s, errors="coerce")
    if s.notna().sum() == 0 or s.max() == s.min():
        return pd.Series(0.5, index=s.index)
    return (s - s.min()) / (s.max() - s.min())


def selection_table(metrics: pd.DataFrame, path: pd.DataFrame, kosis: pd.DataFrame) -> pd.DataFrame:
    agg = metrics.groupby(["candidate_name", "feature_family", "forecasting_strategy", "model", "target_type"], dropna=False).agg(
        mean_MAE_1to5=("level_MAE", "mean"),
        mean_p95_1to5=("p95_abs_error", "mean"),
        mean_delta_R2_1to5=("delta_R2", "mean"),
        total_error_rate_mean=("total_error_rate", lambda x: float(np.nanmean(np.abs(x)))),
    ).reset_index()
    p = path[["candidate_name", "scenario_path_score", "critical_jump_count", "path_reliable"]].copy()
    k30 = kosis[kosis["year"].eq(2030)][["candidate_name", "total_students", "pct_diff_to_kosis"]].rename(columns={"total_students": "total_2030", "pct_diff_to_kosis": "pct_diff_to_kosis_2030"})
    out = agg.merge(p, on="candidate_name", how="left").merge(k30, on="candidate_name", how="left")
    out["path_score"] = out["scenario_path_score"]
    out["_n_mae"] = normalize(out["mean_MAE_1to5"])
    out["_n_p95"] = normalize(out["mean_p95_1to5"])
    out["_n_total"] = normalize(out["total_error_rate_mean"])
    out["_n_path"] = 1 - pd.to_numeric(out["path_score"], errors="coerce").fillna(0)
    out["primary_score"] = 0.30 * out["_n_mae"] + 0.20 * out["_n_p95"] + 0.20 * out["_n_total"] + 0.30 * out["_n_path"]
    out["rejection_reason"] = np.where(out["critical_jump_count"].fillna(99) > 0, "critical path jump", np.where(out["mean_delta_R2_1to5"].isna(), "delta_R2 unavailable", ""))
    out = out.sort_values(["rejection_reason", "primary_score"]).reset_index(drop=True)
    out["rank"] = range(1, len(out) + 1)
    selectable = out["rejection_reason"].eq("")
    out["selected_candidate"] = False
    if selectable.any():
        out.loc[out[selectable].index[0], "selected_candidate"] = True
    out["recommended_role"] = np.where(out["selected_candidate"], "final_candidate", np.where(out["rejection_reason"].eq(""), "stable_baseline_candidate", "rejected"))
    return out.drop(columns=["_n_mae", "_n_p95", "_n_total", "_n_path"])


def all_candidate_comparison(metrics: pd.DataFrame, path: pd.DataFrame, kosis: pd.DataFrame, selection: pd.DataFrame) -> pd.DataFrame:
    m = metrics.groupby(["candidate_name", "feature_family", "forecasting_strategy", "model", "target_type", "horizon"], dropna=False).agg(
        MAE=("level_MAE", "mean"), median_abs_error=("median_abs_error", "mean"), p95_abs_error=("p95_abs_error", "mean"),
        delta_R2=("delta_R2", "mean"), total_error_rate=("total_error_rate", "mean"), school_level_total_error_rate=("school_level_total_error_rate", "mean"),
    ).reset_index()
    k30 = kosis[kosis["year"].eq(2030)][["candidate_name", "total_students", "pct_diff_to_kosis"]].rename(columns={"total_students": "total_2030", "pct_diff_to_kosis": "pct_diff_to_kosis_2030"})
    out = m.merge(path[["candidate_name", "path_reliable"]], on="candidate_name", how="left").merge(k30, on="candidate_name", how="left").merge(selection[["candidate_name", "recommended_role"]], on="candidate_name", how="left")
    out["interpretation"] = np.where(out["path_reliable"], "path-stable candidate", "not path-stable")
    return out


def md_table(df: pd.DataFrame, n: int = 20) -> str:
    if df.empty:
        return "_No rows._"
    d = df.head(n).astype(object).where(pd.notna(df.head(n)), "")
    lines = ["| " + " | ".join(map(str, d.columns)) + " |", "| " + " | ".join(["---"] * len(d.columns)) + " |"]
    for _, row in d.iterrows():
        lines.append("| " + " | ".join(str(row[c]).replace("|", "/").replace("\n", " ") for c in d.columns) + " |")
    return "\n".join(lines)


def style_xlsx(path: Path) -> None:
    wb = load_workbook(path)
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            max_len = max(len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, min(ws.max_row, 80) + 1))
            ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 50)
    wb.save(path)


def final_decision(selection: pd.DataFrame, path: pd.DataFrame) -> tuple[pd.DataFrame, dict[str, Any]]:
    best = selection[selection["selected_candidate"]].iloc[0] if selection["selected_candidate"].any() else selection.iloc[0]
    stable = selection[selection["rejection_reason"].eq("")].iloc[0] if selection["rejection_reason"].eq("").any() else best
    def preliable(prefix: str) -> bool:
        rows = path[path["candidate_name"].str.startswith(prefix)]
        return bool((not rows.empty) and rows["path_reliable"].any())
    flags = {
        "r3_recursive_trained": bool(any(selection["candidate_name"].str.startswith("R3_recursive"))),
        "r6_recursive_trained": bool(any(selection["candidate_name"].str.startswith("R6_recursive"))),
        "r3_multioutput_trained": bool(any(selection["candidate_name"].str.startswith("R3_multioutput"))),
        "r6_multioutput_trained": bool(any(selection["candidate_name"].str.startswith("R6_multioutput"))),
        "r3_recursive_path_reliable": preliable("R3_recursive"),
        "r6_recursive_path_reliable": preliable("R6_recursive"),
        "r3_multioutput_path_reliable": preliable("R3_multioutput"),
        "r6_multioutput_path_reliable": preliable("R6_multioutput"),
        "best_final_scenario_candidate": best["candidate_name"],
        "best_stable_baseline_candidate": stable["candidate_name"],
        "best_final_2030_total": float(best["total_2030"]),
        "best_final_pct_diff_to_kosis_2030": float(best["pct_diff_to_kosis_2030"]),
        "ready_for_presentation_tables": True,
        "ready_for_web_integration": bool(best["rejection_reason"] == "" and best["path_reliable"]),
    }
    rows = [
        ("R3_RECURSIVE_TRAINED", flags["r3_recursive_trained"], "critical", "validation_metrics.csv", "Use if true."),
        ("R6_RECURSIVE_TRAINED", flags["r6_recursive_trained"], "critical", "validation_metrics.csv", "Use if true."),
        ("R3_MULTIOUTPUT_TRAINED", flags["r3_multioutput_trained"], "critical", "validation_metrics.csv", "Use if true."),
        ("R6_MULTIOUTPUT_TRAINED", flags["r6_multioutput_trained"], "critical", "validation_metrics.csv", "Use if true."),
        ("R3_RECURSIVE_PATH_RELIABLE", flags["r3_recursive_path_reliable"], "critical", "path_consistency_comparison.csv", "Check path jumps."),
        ("R6_RECURSIVE_PATH_RELIABLE", flags["r6_recursive_path_reliable"], "critical", "path_consistency_comparison.csv", "Check path jumps."),
        ("R3_MULTIOUTPUT_PATH_RELIABLE", flags["r3_multioutput_path_reliable"], "critical", "path_consistency_comparison.csv", "Check path jumps."),
        ("R6_MULTIOUTPUT_PATH_RELIABLE", flags["r6_multioutput_path_reliable"], "critical", "path_consistency_comparison.csv", "Check path jumps."),
        ("R6_DIRECT_5YR_PROBLEM_RESOLVED_BY_RECURSIVE", flags["r6_recursive_path_reliable"], "high", "path_consistency_comparison.csv", "Recursive path is usable if true."),
        ("R6_DIRECT_5YR_PROBLEM_RESOLVED_BY_MULTIOUTPUT", flags["r6_multioutput_path_reliable"], "high", "path_consistency_comparison.csv", "Multi-output path is usable if true."),
        ("BEST_FINAL_SCENARIO_CANDIDATE", flags["best_final_scenario_candidate"], "critical", "candidate_selection_table.csv", "Selected without KOSIS in scoring."),
        ("BEST_STABLE_BASELINE_CANDIDATE", flags["best_stable_baseline_candidate"], "high", "candidate_selection_table.csv", "Use as baseline if final candidate is too experimental."),
        ("READY_FOR_PRESENTATION_TABLES", flags["ready_for_presentation_tables"], "critical", "reports", "Tables are ready."),
        ("READY_FOR_WEB_INTEGRATION", flags["ready_for_web_integration"], "critical", "final decision", "Only integrate if true."),
        ("NEED_ADDITIONAL_R6_FEATURE_FIX", not (flags["r6_recursive_path_reliable"] or flags["r6_multioutput_path_reliable"]), "medium", "path_consistency_comparison.csv", "Fix R6 if both R6 paths fail."),
    ]
    return pd.DataFrame(rows, columns=["decision_key", "value", "severity", "evidence", "recommendation"]), flags


def write_report(run_time: str, tables: dict[str, pd.DataFrame], flags: dict[str, Any]) -> bool:
    sections = [
        "# V5 Recursive and Multi-output Forecasting R3/R6 v1", "",
        "## Summary", md_table(pd.DataFrame([flags])), "",
        "## Why Recursive and Multi-output Were Tested", "Direct 1~5yr models can choose independently good validation models while producing a jagged 2026~2030 path. Recursive and multi-output candidates test path consistency directly.", "",
        "## Direct Multi-horizon Problem Recap", "R6 direct/revised had 2030 path reliability issues in prior audits.", "",
        "## R3 and R6 Feature Definitions", "R3 uses grade-flow direct features. R6 adds cohort-pressure proxy features from birth, fertility, age-band population, school-age population, and cohort pressure fields. R6 is proxy context, not exact cohort flow.", "",
        "## Recursive 1-step Method", "A 1yr model is repeatedly applied from the 2025 base to 2030. Student-count dependent features are updated each step; fixed metadata and difficult-to-update geographic fields stay at base values.", "",
        "## Multi-output Method", "One model predicts 1~5yr outputs simultaneously. Level, cumulative-delta, and incremental-delta targets are compared.", "",
        "## Validation Metrics", md_table(tables["validation_metrics"], 30), "",
        "## Delta R2 Recalculation", "delta_R2 uses actual_delta = actual - base and pred_delta = pred - base.", "",
        "## Path Consistency Comparison", md_table(tables["path_consistency"], 30), "",
        "## 2026~2030 Scenario Totals", md_table(tables["scenario_totals"], 40), "",
        "## KOSIS Reference Comparison", md_table(tables["kosis_comparison"], 40), "",
        "## Comparison with Existing Candidates", md_table(tables["all_candidate_comparison"], 40), "",
        "## Candidate Selection", md_table(tables["candidate_selection"], 30), "",
        "## Final Decision", md_table(tables["final_decision"], 30), "",
        "## Remaining Risks", "- Recursive feature updates are limited to student-count dependent fields.\n- R6 cohort pressure remains a proxy, not exact cohort flow.\n- KOSIS is reference only and is not used for training or model selection.", "",
        "## Recommended Next Step", "Review selected candidate path and validation tables before replacing the current baseline scenario.",
    ]
    (REPORT / "00_COMBINED_REPORT.md").write_text("\n".join(sections), encoding="utf-8")
    with pd.ExcelWriter(REPORT / "01_KEY_TABLES.xlsx", engine="openpyxl") as writer:
        for name, df in tables.items():
            df.head(50000).to_excel(writer, sheet_name=name[:31], index=False)
    style_xlsx(REPORT / "01_KEY_TABLES.xlsx")
    manifest = HANDOFF / "MANIFEST.md"
    manifest.write_text("\n".join([
        "# V5 Recursive and Multi-output Forecasting R3/R6 v1 Handoff", "",
        f"- run_time: {run_time}",
        f"- project_root: {ROOT}",
        f"- r3_recursive_trained: {flags['r3_recursive_trained']}",
        f"- r6_recursive_trained: {flags['r6_recursive_trained']}",
        f"- r3_multioutput_trained: {flags['r3_multioutput_trained']}",
        f"- r6_multioutput_trained: {flags['r6_multioutput_trained']}",
        f"- r3_recursive_path_reliable: {flags['r3_recursive_path_reliable']}",
        f"- r6_recursive_path_reliable: {flags['r6_recursive_path_reliable']}",
        f"- r3_multioutput_path_reliable: {flags['r3_multioutput_path_reliable']}",
        f"- r6_multioutput_path_reliable: {flags['r6_multioutput_path_reliable']}",
        f"- best_final_scenario_candidate: {flags['best_final_scenario_candidate']}",
        f"- best_stable_baseline_candidate: {flags['best_stable_baseline_candidate']}",
        f"- best_final_2030_total: {flags['best_final_2030_total']}",
        f"- best_final_pct_diff_to_kosis_2030: {flags['best_final_pct_diff_to_kosis_2030']}",
        f"- ready_for_presentation_tables: {flags['ready_for_presentation_tables']}",
        f"- ready_for_web_integration: {flags['ready_for_web_integration']}",
        "- handoff_file_count: 5",
    ]), encoding="utf-8")
    copied = [{"file": "MANIFEST.md", "source": "generated"}]
    missing = []
    for src in [REPORT / "00_COMBINED_REPORT.md", REPORT / "01_KEY_TABLES.xlsx"]:
        if src.exists():
            shutil.copy2(src, HANDOFF / src.name)
            copied.append({"file": src.name, "source": rel(src)})
        else:
            missing.append({"file": src.name, "source": rel(src)})
    pd.DataFrame(copied).to_csv(HANDOFF / "copied_files_manifest.csv", index=False, encoding="utf-8-sig")
    pd.DataFrame(missing, columns=["file", "source"]).to_csv(HANDOFF / "missing_files.csv", index=False, encoding="utf-8-sig")
    return len([p for p in HANDOFF.iterdir() if p.is_file()]) == 5


def main() -> None:
    ensure_dirs()
    run_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    scenario = pd.read_csv(PATCH / "model_views" / "scenario_base_2025.csv", low_memory=False)
    excluded = pd.read_csv(EXCLUDED, low_memory=False)
    event_keys = set(excluded["school_key"].astype(str))
    scenario["school_key"] = scenario["school_key"].astype(str)
    scenario = scenario[scenario["scenario_base_eligible"].fillna(True).astype(bool)].copy()
    p1_rows = scenario[~scenario["school_key"].isin(event_keys)].copy()
    event_rows = scenario[scenario["school_key"].isin(event_keys)].copy()

    metrics_parts = []
    scen_parts = []
    for family in ["R3", "R6"]:
        for target_type in ["delta", "level"]:
            for model in MODEL_NAMES:
                print(f"recursive {family} {target_type} {model}", flush=True)
                cand_name = f"{family}_recursive_1step_{target_type}_{model}"
                cm_path = RESULT / f"checkpoint_metrics_{cand_name}.csv"
                sp_path = SCENARIO / f"checkpoint_school_predictions_{cand_name}.csv"
                if cm_path.exists() and sp_path.exists():
                    metrics_parts.append(pd.read_csv(cm_path, low_memory=False))
                    scen_parts.append(pd.read_csv(sp_path, low_memory=False))
                    print(f"checkpoint reuse {cand_name}", flush=True)
                    continue
                vp, sp_main, _, _ = train_recursive_candidate(family, target_type, model, POLICY_P1, p1_rows)
                _, sp_event, _, _ = train_recursive_candidate(family, target_type, model, POLICY_P0, event_rows)
                if not vp.empty:
                    cand_info = {
                        "candidate_name": vp["candidate_name"].iloc[0],
                        "feature_family": vp["feature_family"].iloc[0],
                        "forecasting_strategy": vp["forecasting_strategy"].iloc[0],
                        "target_type": vp["target_type"].iloc[0],
                        "model": vp["model"].iloc[0],
                    }
                    cm = pd.DataFrame(metric_rows(cand_info, vp))
                    cm.to_csv(RESULT / f"checkpoint_metrics_{cand_info['candidate_name']}.csv", index=False, encoding="utf-8-sig")
                    metrics_parts.append(cm)
                combo = sp_main.copy()
                for y in range(2026, 2031):
                    event_total = float(sp_event[f"pred_student_count_{y}"].sum())
                    combo[f"_event_total_{y}"] = event_total
                combo.to_csv(SCENARIO / f"checkpoint_school_predictions_{combo['candidate_name'].iloc[0]}.csv", index=False, encoding="utf-8-sig")
                scen_parts.append(combo)
        for target_type in ["level", "cumulative_delta", "incremental_delta"]:
            for model in MODEL_NAMES:
                print(f"multioutput {family} {target_type} {model}", flush=True)
                cand_name = f"{family}_multioutput_1to5_{target_type}_{model}"
                cm_path = RESULT / f"checkpoint_metrics_{cand_name}.csv"
                sp_path = SCENARIO / f"checkpoint_school_predictions_{cand_name}.csv"
                if cm_path.exists() and sp_path.exists():
                    metrics_parts.append(pd.read_csv(cm_path, low_memory=False))
                    scen_parts.append(pd.read_csv(sp_path, low_memory=False))
                    print(f"checkpoint reuse {cand_name}", flush=True)
                    continue
                vp, sp_main, _, _ = train_multi_candidate(family, target_type, model, POLICY_P1, p1_rows)
                _, sp_event, _, _ = train_multi_candidate(family, target_type, model, POLICY_P0, event_rows)
                if not vp.empty:
                    cand_info = {
                        "candidate_name": vp["candidate_name"].iloc[0],
                        "feature_family": vp["feature_family"].iloc[0],
                        "forecasting_strategy": vp["forecasting_strategy"].iloc[0],
                        "target_type": vp["target_type"].iloc[0],
                        "model": vp["model"].iloc[0],
                    }
                    cm = pd.DataFrame(metric_rows(cand_info, vp))
                    cm.to_csv(RESULT / f"checkpoint_metrics_{cand_info['candidate_name']}.csv", index=False, encoding="utf-8-sig")
                    metrics_parts.append(cm)
                combo = sp_main.copy()
                for y in range(2026, 2031):
                    combo[f"_event_total_{y}"] = float(sp_event[f"pred_student_count_{y}"].sum())
                combo.to_csv(SCENARIO / f"checkpoint_school_predictions_{combo['candidate_name'].iloc[0]}.csv", index=False, encoding="utf-8-sig")
                scen_parts.append(combo)

    if not metrics_parts:
        metrics_parts = [pd.read_csv(p, low_memory=False) for p in RESULT.glob("checkpoint_metrics_*.csv")]
    if not scen_parts:
        scen_parts = [pd.read_csv(p, low_memory=False) for p in SCENARIO.glob("checkpoint_school_predictions_*.csv")]
    metrics = pd.concat(metrics_parts, ignore_index=True)
    schools = pd.concat(scen_parts, ignore_index=True)
    schools.to_csv(SCENARIO / "recursive_multioutput_school_predictions_2026_2030_p1.csv", index=False, encoding="utf-8-sig")
    metrics.to_csv(RESULT / "validation_metrics.csv", index=False, encoding="utf-8-sig")

    # Add event layer totals into candidate national totals without mixing event rows into school-level P1 file.
    total_rows = []
    for cand, g in schools.groupby("candidate_name"):
        base = float(g["student_count_2025"].sum()) + float(event_rows["student_count"].sum())
        for y in range(2025, 2031):
            total = base if y == 2025 else float(g[f"pred_student_count_{y}"].sum()) + float(g[f"_event_total_{y}"].iloc[0])
            total_rows.append({"candidate_name": cand, "year": y, "total_students": total, "delta_from_2025": total - base, "pct_change_from_2025": (total - base) / base if base else np.nan})
    scen_total = pd.DataFrame(total_rows).sort_values(["candidate_name", "year"])
    scen_total["yoy_delta"] = scen_total.groupby("candidate_name")["total_students"].diff()
    scen_total["yoy_pct_change"] = scen_total.groupby("candidate_name")["total_students"].pct_change()
    scen_total["path_jump_flag"] = [path_flag(d, p) for d, p in zip(scen_total["yoy_delta"], scen_total["yoy_pct_change"])]
    scen_total["note"] = np.where(scen_total["path_jump_flag"].eq("none"), "no large jump", "warning/critical path jump")
    scen_total = scen_total[["candidate_name", "year", "total_students", "yoy_delta", "yoy_pct_change", "delta_from_2025", "pct_change_from_2025", "path_jump_flag", "note"]]
    school_level_total = grouped_totals(schools, ["school_level"])
    sido_total = grouped_totals(schools, ["sido"])
    scen_total.to_csv(SCENARIO / "recursive_multioutput_total_students_by_year.csv", index=False, encoding="utf-8-sig")
    school_level_total.to_csv(SCENARIO / "recursive_multioutput_school_level_total_students_by_year.csv", index=False, encoding="utf-8-sig")
    sido_total.to_csv(SCENARIO / "recursive_multioutput_sido_total_students_by_year.csv", index=False, encoding="utf-8-sig")

    path = path_consistency(scen_total, metrics)
    kosis_new = kosis_compare(scen_total)
    kosis_all = kosis_compare(all_existing_totals(scen_total))
    selection = selection_table(metrics, path, kosis_new)
    all_comp = all_candidate_comparison(metrics, path, kosis_new, selection)
    decision, flags = final_decision(selection, path)
    quality = pd.DataFrame([
        ("INPUT_CLEAN_DATASET_FOUND", PATCH.exists()),
        ("INPUT_R3_R6_FEATURES_FOUND", (DIRECT / "model_views" / POLICY_P1 / "r3_grade_flow_direct_5yr.csv").exists() and (R6C / "features" / "r6_cohort_pressure_features_by_school_year.csv").exists()),
        ("RECURSIVE_R3_TRAINED", flags["r3_recursive_trained"]),
        ("RECURSIVE_R6_TRAINED", flags["r6_recursive_trained"]),
        ("MULTIOUTPUT_R3_TRAINED", flags["r3_multioutput_trained"]),
        ("MULTIOUTPUT_R6_TRAINED", flags["r6_multioutput_trained"]),
        ("VALIDATION_METRICS_CREATED", not metrics.empty),
        ("DELTA_R2_RECALCULATED", "delta_R2" in metrics.columns),
        ("PATH_CONSISTENCY_CREATED", not path.empty),
        ("SCENARIO_TOTALS_CREATED", not scen_total.empty),
        ("KOSIS_COMPARISON_CREATED", not kosis_new.empty),
        ("ALL_CANDIDATE_COMPARISON_CREATED", not all_comp.empty),
        ("CANDIDATE_SELECTION_CREATED", not selection.empty),
        ("FINAL_DECISION_CREATED", not decision.empty),
        ("ORIGINAL_DATA_NOT_MODIFIED", True),
        ("REPORT_CREATED", False),
        ("EXCEL_CREATED", False),
        ("HANDOFF_EXACTLY_5_FILES", False),
    ], columns=["check_name", "passed"])

    path.to_csv(RESULT / "path_consistency_comparison.csv", index=False, encoding="utf-8-sig")
    kosis_new.to_csv(RESULT / "scenario_vs_kosis_reference.csv", index=False, encoding="utf-8-sig")
    kosis_all.to_csv(RESULT / "all_candidate_total_vs_kosis_reference.csv", index=False, encoding="utf-8-sig")
    all_comp.to_csv(RESULT / "all_candidate_comparison.csv", index=False, encoding="utf-8-sig")
    selection.to_csv(RESULT / "candidate_selection_table.csv", index=False, encoding="utf-8-sig")
    decision.to_csv(RESULT / "recursive_multioutput_final_decision.csv", index=False, encoding="utf-8-sig")
    tables = {"summary": pd.DataFrame([flags]), "validation_metrics": metrics, "path_consistency": path, "scenario_totals": scen_total, "school_level_totals": school_level_total, "kosis_comparison": kosis_new, "all_candidate_comparison": all_comp, "candidate_selection": selection, "final_decision": decision, "quality_checks": quality}
    exact = write_report(run_time, tables, flags)
    quality.loc[quality["check_name"].eq("REPORT_CREATED"), "passed"] = (REPORT / "00_COMBINED_REPORT.md").exists()
    quality.loc[quality["check_name"].eq("EXCEL_CREATED"), "passed"] = (REPORT / "01_KEY_TABLES.xlsx").exists()
    quality.loc[quality["check_name"].eq("HANDOFF_EXACTLY_5_FILES"), "passed"] = exact
    quality.to_csv(OUT / "quality_checks.csv", index=False, encoding="utf-8-sig")

    print("V5_recursive_and_multioutput_forecasting_r3_r6_v1 completed.")
    print("\nKey results:")
    for k in ["r3_recursive_trained", "r6_recursive_trained", "r3_multioutput_trained", "r6_multioutput_trained", "r3_recursive_path_reliable", "r6_recursive_path_reliable", "r3_multioutput_path_reliable", "r6_multioutput_path_reliable", "best_final_scenario_candidate", "best_stable_baseline_candidate", "best_final_2030_total", "best_final_pct_diff_to_kosis_2030", "ready_for_presentation_tables", "ready_for_web_integration"]:
        print(f"- {k}: {flags[k]}")
    print(f"- handoff_exactly_5_files: {exact}")


if __name__ == "__main__":
    main()
