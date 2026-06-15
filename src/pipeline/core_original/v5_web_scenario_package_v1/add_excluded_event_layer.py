from __future__ import annotations

import hashlib
import json
import shutil
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
DATA_OUT = ROOT / "data" / "v5_web_scenario_package_v1"
PUBLIC_OUT = ROOT / "public" / "data" / "scenario_v5_v2"
HANDOFF = ROOT / "handoff_for_chatgpt" / "v5_web_scenario_package_v1"

EXCLUDED_SRC = ROOT / "data" / "v5_p1_excluded_school_list_v1" / "p1_excluded_schools_2173.csv"
BASE_2025 = ROOT / "data" / "v5_clean_dataset_patch_v1" / "model_views" / "scenario_base_2025.csv"
GEO_SRC = ROOT / "data" / "raw" / "school_data_2008_2025_geocoded.csv"
MAIN_WEB = DATA_OUT / "final_scenario_school_web.csv"


def read_csv(path: Path) -> pd.DataFrame:
    if not path.exists():
        return pd.DataFrame()
    return pd.read_csv(path, low_memory=False)


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def backup(path: Path, suffix: str) -> None:
    if path.exists():
        target = path.with_name(path.stem + suffix + path.suffix)
        if not target.exists():
            shutil.copy2(path, target)


def coord_quality(lat: pd.Series, lon: pd.Series) -> tuple[pd.Series, pd.Series]:
    latn = pd.to_numeric(lat, errors="coerce")
    lonn = pd.to_numeric(lon, errors="coerce")
    missing = latn.isna() | lonn.isna()
    out_bounds = (~missing) & (~latn.between(33.0, 39.5) | ~lonn.between(124.0, 132.0))
    flag = pd.Series("valid", index=lat.index, dtype=object)
    flag.loc[missing] = "missing_coordinate"
    flag.loc[out_bounds] = "out_of_korea_bounds"
    return flag.eq("valid"), flag


def to_json_records(df: pd.DataFrame, path: Path) -> None:
    clean = df.replace({np.nan: None})
    path.write_text(
        json.dumps(clean.to_dict(orient="records"), ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )


def normalize_category(row: pd.Series) -> str:
    raw = str(row.get("event_layer_category", "") or "").lower()
    flags = str(row.get("event_flags", "") or "").lower()
    status = str(row.get("latest_status", "") or row.get("school_status_2025", "") or "")
    reason = str(row.get("exclusion_reason", "") or "").lower()
    text = ";".join([raw, flags, reason])

    if "휴교" in status:
        return "suspended_school"
    if "폐" in status:
        return "closed_school"
    if "coordinate" in text:
        return "coordinate_issue"
    if "entity" in text or "identity" in text:
        return "entity_resolution_issue"
    if "temporary_zero" in text or "zero" in text:
        return "zero_to_large_jump"
    if "reopened" in text:
        return "reopened_school"
    if "new" in text:
        return "new_school"
    if "large_growth" in text or "positive" in text or "increase" in text:
        return "rapid_student_increase"
    if "rapid_decline" in text or "negative" in text or "decline" in text:
        return "rapid_student_decrease"
    if "status" in text:
        return "status_event"
    if text.strip(";"):
        return "other_event_or_anomaly"
    return "unknown"


def display_description(category: str) -> str:
    descriptions = {
        "new_school": "신설 또는 새로 관측된 학교로, 안정 기존학교 감소압력 모델의 학습/예측 대상에서 제외되어 별도 event/anomaly layer로 관리됩니다. 이 레이어는 행정 상태 확정 예측이 아닙니다.",
        "closed_school": "원천 데이터의 학교 상태 변화가 확인된 학교로, 안정 기존학교 감소압력 모델에 섞지 않고 별도 event/anomaly layer로 관리됩니다. 이 레이어는 향후 상태를 예측하는 레이어가 아닙니다.",
        "suspended_school": "휴교 등 학교 상태 이벤트가 확인된 학교로, 안정 기존학교 감소압력 모델의 학습/예측 대상에서 제외되었습니다. 이 레이어는 행정 검토 참고용입니다.",
        "reopened_school": "재개교 또는 관측 재등장 가능성이 있는 학교로, 일반 감소 추세와 구분해 별도 event/anomaly layer로 관리됩니다.",
        "rapid_student_increase": "학생 수가 급격히 증가한 이력이 있어 일반적인 감소압력 학습 대상에서 제외되었습니다. 웹에서는 별도 event/anomaly 학교로 표시합니다.",
        "rapid_student_decrease": "학생 수가 급격히 감소한 이력이 있어 안정 기존학교 감소압력 모델의 학습/예측 대상에서 제외되었습니다. 이는 행정 상태 확정 예측이 아닙니다.",
        "zero_to_large_jump": "학생 수 0 또는 급격한 점프 이력이 있어 원천 데이터 확인이 필요한 event/anomaly 학교로 분리했습니다.",
        "coordinate_issue": "좌표 또는 위치 정보 품질 이슈가 있어 지도 해석 시 주의가 필요한 학교입니다. 임의 좌표 보정은 하지 않았습니다.",
        "entity_resolution_issue": "학교 식별자 또는 동일 학교 매칭 검토가 필요한 학교로, 안정학교 예측 레이어와 분리했습니다.",
        "status_event": "학교 상태 이벤트가 확인되어 안정 기존학교 감소압력 모델과 별도로 관리합니다.",
        "other_event_or_anomaly": "신설, 상태 변화, 급격한 학생 수 변화 등 event/anomaly 사유로 안정학교 모델에서 제외된 학교입니다.",
        "unknown": "제외 사유가 명확히 분류되지 않아 별도 event/anomaly layer로 표시합니다.",
    }
    return descriptions.get(category, descriptions["unknown"])


def marker_priority(category: str) -> int:
    priorities = {
        "coordinate_issue": 90,
        "entity_resolution_issue": 85,
        "zero_to_large_jump": 80,
        "rapid_student_decrease": 75,
        "rapid_student_increase": 70,
        "closed_school": 65,
        "suspended_school": 65,
        "new_school": 55,
        "reopened_school": 55,
        "status_event": 60,
        "other_event_or_anomaly": 45,
        "unknown": 30,
    }
    return priorities.get(category, 30)


def make_dictionary(df: pd.DataFrame) -> pd.DataFrame:
    labels = {
        "school_key": ("학교 고유키", "학교를 식별하는 V5 기준 고유키"),
        "school_name": ("학교명", "학교 이름"),
        "sido": ("시도", "시도명"),
        "sgg": ("시군구", "시군구명"),
        "school_level": ("학교급", "초등학교/중학교/고등학교 등 학교급"),
        "school_status_2025": ("2025 학교 상태", "2025년 원천 데이터에서 확인되는 학교 상태"),
        "latitude": ("위도", "지도 표시용 위도. 결측이면 임의 생성하지 않음"),
        "longitude": ("경도", "지도 표시용 경도. 결측이면 임의 생성하지 않음"),
        "coordinate_valid": ("좌표 유효 여부", "대한민국 대략 범위 기준 좌표 유효성"),
        "coordinate_source": ("좌표 출처", "좌표가 조인된 원천 또는 출처 정보"),
        "coordinate_quality_flag": ("좌표 품질 플래그", "valid/missing_coordinate/out_of_korea_bounds"),
        "student_count_2025": ("2025 학생수", "2025년 최신 학생 수"),
        "class_count_2025": ("2025 학급수", "2025년 학급 수"),
        "teacher_count_2025": ("2025 교원수", "2025년 교원 수"),
        "students_per_class_2025": ("2025 학급당 학생수", "학생수/학급수 기반 2025 지표"),
        "students_per_teacher_2025": ("2025 교원 1인당 학생수", "학생수/교원수 기반 2025 지표"),
        "isolation_score": ("고립도 점수", "동일 학교급 주변 학교 접근성을 반영한 고립도 지표"),
        "nearest_same_level_school_km": ("가장 가까운 동일 학교급 거리", "동일 학교급 기준 최근접 학교 거리(km)"),
        "same_level_school_count_5km": ("5km 내 동일 학교급 수", "반경 5km 내 동일 학교급 학교 수"),
        "excluded_from_model": ("학습 대상 제외 여부", "안정학교 모델 학습/예측 대상 제외 여부"),
        "exclusion_reason": ("제외 사유", "P1 정책에서 제외된 원천 사유"),
        "event_layer_category": ("event/anomaly 유형", "웹 표시용 정규화된 이벤트/이상 유형"),
        "modeling_layer": ("모델링 계층", "excluded_event_anomaly 레이어 구분값"),
        "has_prediction": ("예측값 보유 여부", "이 레이어는 학교별 2026~2030 예측값을 갖지 않음"),
        "display_label": ("화면 표시 라벨", "웹 지도/목록에서 보여줄 짧은 라벨"),
        "display_description": ("화면 표시 설명", "상세 패널에 보여줄 한국어 설명"),
        "web_marker_type": ("지도 마커 유형", "웹 지도에서 excluded event/anomaly로 표시하기 위한 유형"),
        "web_marker_priority": ("지도 마커 우선순위", "웹 표시 정렬 또는 강조를 위한 우선순위"),
        "source_file": ("원천 파일", "해당 row가 만들어진 원천 파일"),
        "source_row_id": ("원천 row ID", "원천 파일 내 행 번호"),
    }
    rows = []
    for col in df.columns:
        label, desc = labels.get(col, (col, "excluded/event layer 웹 데이터 컬럼"))
        sample = ""
        if len(df) and col in df:
            non_null = df[col].dropna()
            if len(non_null):
                sample = str(non_null.iloc[0])
        rows.append(
            {
                "column_name": col,
                "korean_label": label,
                "description_ko": desc,
                "dtype": str(df[col].dtype),
                "example": sample,
                "web_usage": "map/detail/filter/audit",
            }
        )
    return pd.DataFrame(rows)


def build_summary(excluded: pd.DataFrame) -> pd.DataFrame:
    out = []

    def add(group_type: str, keys: list[str]) -> None:
        if keys:
            grouped = excluded.groupby(keys, dropna=False)
        else:
            grouped = [("__all__", excluded)]
        for group_key, g in grouped:
            if not isinstance(group_key, tuple):
                group_key = (group_key,)
            rec = {
                "group_type": group_type,
                "sido": "",
                "sgg": "",
                "school_level": "",
                "event_layer_category": "",
                "school_count": g["school_key"].nunique(),
                "valid_coordinate_count": int(g["coordinate_valid"].sum()),
                "invalid_coordinate_count": int((~g["coordinate_valid"]).sum()),
                "has_student_count_2025_count": int(g["student_count_2025"].notna().sum()),
                "mean_student_count_2025": g["student_count_2025"].mean(),
                "median_student_count_2025": g["student_count_2025"].median(),
            }
            for k, v in zip(keys, group_key):
                rec[k] = v
            out.append(rec)

    add("national_total", [])
    add("by_sido", ["sido"])
    add("by_school_level", ["school_level"])
    add("by_sido_school_level", ["sido", "school_level"])
    add("by_event_layer_category", ["event_layer_category"])
    add("by_sido_event_layer_category", ["sido", "event_layer_category"])
    return pd.DataFrame(out)


def style_xlsx(path: Path) -> None:
    wb = load_workbook(path)
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            max_len = max(len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, min(ws.max_row, 100) + 1))
            ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 55)
    wb.save(path)


def append_report(report_path: Path, stats: dict[str, object]) -> None:
    backup(report_path, ".before_excluded_layer_update")
    section = f"""

## 16. 학습 대상 제외학교 / event-anomaly layer

이번 웹 패키지에는 P1 안정 학교 예측 row와 별도로, 학습/예측 대상에서 제외된 event-anomaly 학교 목록을 추가했다.

이 레이어는 행정 상태 확정 예측이 아니다. 신설, 상태 변화, 급격한 학생 수 변화, 좌표/식별자 검토 필요 등으로 안정적인 감소압력 모델 학습 대상에서 분리된 학교를 웹에서 참고 레이어로 확인하기 위한 데이터다.

- excluded_school_rows: {stats["excluded_school_rows"]}
- unique_excluded_school_keys: {stats["unique_excluded_school_keys"]}
- valid_coordinate_count: {stats["valid_coordinate_count"]}
- invalid_coordinate_count: {stats["invalid_coordinate_count"]}
- overlap_with_final_scenario_school_web: {stats["overlap_with_final_scenario_school_web"]}

권장 UI 사용 방식은 기본 지도에는 P1 안정 학교를 표시하고, 사용자가 선택하면 excluded/event 학교를 별도 마커로 켜서 검토하는 것이다.
"""
    old = report_path.read_text(encoding="utf-8") if report_path.exists() else ""
    marker = "## 16. 학습 대상 제외학교 / event-anomaly layer"
    if marker in old:
        old = old.split(marker)[0].rstrip()
    report_path.write_text(old + section, encoding="utf-8")


def append_manifest(manifest_path: Path, stats: dict[str, object]) -> None:
    backup(manifest_path, ".before_excluded_layer_update")
    section = f"""

## Excluded Event/Anomaly Layer

excluded_school_layer_created: True
excluded_school_rows: {stats["excluded_school_rows"]}
excluded_school_valid_coordinate_count: {stats["valid_coordinate_count"]}
excluded_school_invalid_coordinate_count: {stats["invalid_coordinate_count"]}
overlap_with_final_scenario_school_web: {stats["overlap_with_final_scenario_school_web"]}

excluded_school_files:
- excluded_school_web.csv
- excluded_school_web.json
- excluded_school_summary.csv
- excluded_school_data_dictionary_ko.csv
- excluded_school_audit.csv
- excluded_school_year_long.csv
- excluded_school_year_long.json
"""
    old = manifest_path.read_text(encoding="utf-8") if manifest_path.exists() else ""
    marker = "## Excluded Event/Anomaly Layer"
    if marker in old:
        old = old.split(marker)[0].rstrip()
    manifest_path.write_text(old + section, encoding="utf-8")


def update_metadata(path: Path) -> bool:
    backup(path, ".before_excluded_layer_update")
    meta = {}
    if path.exists():
        meta = json.loads(path.read_text(encoding="utf-8"))
    meta["excluded_school_layer"] = {
        "file_csv": "excluded_school_web.csv",
        "file_json": "excluded_school_web.json",
        "meaning": "학습/예측 대상에서 제외되어 별도 event/anomaly layer로 관리되는 학교",
        "default_visible": False,
        "recommended_ui_label": "학습 대상 제외학교",
        "recommended_ui_description": "신설, 상태 변화, 급격한 학생 수 변화, 좌표/식별자 검토 필요 등으로 안정 학교 감소압력 모델의 학습/예측 대상에서 제외된 학교입니다.",
        "warning": "이 레이어는 행정 상태 확정 예측 결과가 아닙니다.",
    }
    path.write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")
    return True


def main() -> None:
    DATA_OUT.mkdir(parents=True, exist_ok=True)
    PUBLIC_OUT.mkdir(parents=True, exist_ok=True)
    HANDOFF.mkdir(parents=True, exist_ok=True)

    missing_rows: list[dict[str, str]] = []
    excluded_src = read_csv(EXCLUDED_SRC)
    base = read_csv(BASE_2025)
    main_web = read_csv(MAIN_WEB)

    if excluded_src.empty:
        raise FileNotFoundError(EXCLUDED_SRC)

    excluded_src = excluded_src.copy()
    excluded_src["source_file"] = str(EXCLUDED_SRC.relative_to(ROOT))
    excluded_src["source_row_id"] = np.arange(len(excluded_src))

    base_cols = {
        "school_key": "school_key",
        "status": "school_status_2025",
        "student_count": "student_count_2025_base",
        "class_count": "class_count_2025",
        "teacher_count": "teacher_count_2025",
        "students_per_class": "students_per_class_2025",
        "students_per_teacher": "students_per_teacher_2025",
        "isolation_score": "isolation_score",
        "nearest_same_level_distance_km": "nearest_same_level_school_km",
        "same_level_school_count_within_5km": "same_level_school_count_5km",
        "coordinate_source": "coordinate_source_base",
    }
    base_use = pd.DataFrame()
    if not base.empty:
        base_use = base[[c for c in base_cols if c in base.columns]].rename(columns=base_cols)

    df = excluded_src.merge(base_use, on="school_key", how="left")
    df["student_count_2025"] = pd.to_numeric(df.get("student_count_2025_base"), errors="coerce")
    df["student_count_2025"] = df["student_count_2025"].fillna(pd.to_numeric(df.get("latest_student_count"), errors="coerce"))
    df["school_status_2025"] = df.get("school_status_2025", pd.Series(index=df.index, dtype=object)).fillna(df.get("latest_status"))

    geo = read_csv(GEO_SRC)
    if not geo.empty:
        geo25 = geo[geo["year"].eq(2025)].copy()
        geo25 = geo25.rename(
            columns={
                "시도": "sido",
                "행정구": "sgg",
                "학교급": "school_level",
                "학교명": "school_name",
                "lttud": "latitude",
                "lgtud": "longitude",
                "nearest_same_level_school_km": "nearest_same_level_school_km_geo",
                "same_level_school_count_5km": "same_level_school_count_5km_geo",
                "school_isolation_score": "isolation_score_geo",
            }
        )
        keep = [
            "sido",
            "sgg",
            "school_level",
            "school_name",
            "latitude",
            "longitude",
            "coordinate_source",
            "nearest_same_level_school_km_geo",
            "same_level_school_count_5km_geo",
            "isolation_score_geo",
        ]
        geo25 = geo25[[c for c in keep if c in geo25.columns]].drop_duplicates(["sido", "sgg", "school_level", "school_name"])
        df = df.merge(geo25, on=["sido", "sgg", "school_level", "school_name"], how="left", suffixes=("", "_geo"))
        df["coordinate_source"] = df.get("coordinate_source_base", pd.Series(index=df.index, dtype=object)).fillna(df.get("coordinate_source"))
        df["nearest_same_level_school_km"] = df.get("nearest_same_level_school_km").fillna(df.get("nearest_same_level_school_km_geo"))
        df["same_level_school_count_5km"] = df.get("same_level_school_count_5km").fillna(df.get("same_level_school_count_5km_geo"))
        df["isolation_score"] = df.get("isolation_score").fillna(df.get("isolation_score_geo"))
    else:
        missing_rows.append(
            {
                "missing_item": "2025 geocoded source",
                "item_type": "file",
                "expected_path_or_column": str(GEO_SRC.relative_to(ROOT)),
                "reason": "source file not found",
                "impact": "coordinates unavailable for excluded layer",
                "required_action": "provide source-backed coordinate file",
            }
        )

    df["coordinate_valid"], df["coordinate_quality_flag"] = coord_quality(df.get("latitude"), df.get("longitude"))
    df["event_layer_category"] = df.apply(normalize_category, axis=1)
    df["excluded_from_model"] = True
    df["modeling_layer"] = "excluded_event_anomaly"
    df["has_prediction"] = False
    df["display_label"] = "학습 대상 제외"
    df["display_description"] = df["event_layer_category"].map(display_description)
    df["web_marker_type"] = "excluded_event_anomaly"
    df["web_marker_priority"] = df["event_layer_category"].map(marker_priority)

    for col in ["class_count_2025", "teacher_count_2025", "students_per_class_2025", "students_per_teacher_2025"]:
        if col not in df.columns:
            df[col] = np.nan
            missing_rows.append(
                {
                    "missing_item": col,
                    "item_type": "column",
                    "expected_path_or_column": col,
                    "reason": "column not available after source joins",
                    "impact": "emitted as null where source-backed value is unavailable",
                    "required_action": "provide audited source if needed",
                }
            )

    required_cols = [
        "school_key",
        "school_name",
        "sido",
        "sgg",
        "school_level",
        "school_status_2025",
        "latitude",
        "longitude",
        "coordinate_valid",
        "coordinate_source",
        "coordinate_quality_flag",
        "student_count_2025",
        "class_count_2025",
        "teacher_count_2025",
        "students_per_class_2025",
        "students_per_teacher_2025",
        "isolation_score",
        "nearest_same_level_school_km",
        "same_level_school_count_5km",
        "excluded_from_model",
        "exclusion_reason",
        "event_layer_category",
        "modeling_layer",
        "has_prediction",
        "display_label",
        "display_description",
        "web_marker_type",
        "web_marker_priority",
        "source_file",
        "source_row_id",
    ]
    for col in required_cols:
        if col not in df.columns:
            df[col] = np.nan
            missing_rows.append(
                {
                    "missing_item": col,
                    "item_type": "column",
                    "expected_path_or_column": col,
                    "reason": "required column not available",
                    "impact": "emitted as null for schema stability",
                    "required_action": "check excluded source/schema",
                }
            )
    excluded_web = df[required_cols].copy()

    excluded_web_path = DATA_OUT / "excluded_school_web.csv"
    excluded_web_json = DATA_OUT / "excluded_school_web.json"
    excluded_web.to_csv(excluded_web_path, index=False, encoding="utf-8-sig")
    to_json_records(excluded_web, excluded_web_json)

    long = excluded_web[
        [
            "school_key",
            "school_name",
            "sido",
            "sgg",
            "school_level",
            "school_status_2025",
            "latitude",
            "longitude",
            "coordinate_valid",
            "student_count_2025",
            "event_layer_category",
            "exclusion_reason",
            "modeling_layer",
            "has_prediction",
        ]
    ].copy()
    long["year"] = 2025
    long["student_count"] = long["student_count_2025"]
    long["is_observed"] = True
    long["is_predicted"] = False
    long_path = DATA_OUT / "excluded_school_year_long.csv"
    long_json = DATA_OUT / "excluded_school_year_long.json"
    long.to_csv(long_path, index=False, encoding="utf-8-sig")
    to_json_records(long, long_json)

    summary = build_summary(excluded_web)
    summary_path = DATA_OUT / "excluded_school_summary.csv"
    summary.to_csv(summary_path, index=False, encoding="utf-8-sig")

    overlap = 0
    if not main_web.empty:
        overlap = int(excluded_web["school_key"].isin(set(main_web["school_key"])).sum())
    category_counts = excluded_web["event_layer_category"].value_counts(dropna=False).to_dict()
    audit_rows = [
        ("excluded_school_rows", len(excluded_web), "Rows in excluded_school_web.csv"),
        ("unique_excluded_school_keys", excluded_web["school_key"].nunique(), "Unique school_key count"),
        ("overlap_with_final_scenario_school_web", overlap, "Overlap is recorded only; rows are not deleted"),
        ("excluded_with_valid_coordinates", int(excluded_web["coordinate_valid"].sum()), "Coordinate valid under Korea bounds check"),
        ("excluded_with_invalid_coordinates", int((~excluded_web["coordinate_valid"]).sum()), "Missing or out-of-bounds coordinates"),
        ("excluded_with_student_count_2025", int(excluded_web["student_count_2025"].notna().sum()), "Rows with latest 2025 student count"),
        ("excluded_without_student_count_2025", int(excluded_web["student_count_2025"].isna().sum()), "Rows missing latest 2025 student count"),
        ("excluded_by_event_layer_category", json.dumps(category_counts, ensure_ascii=False), "Category distribution"),
    ]
    audit = pd.DataFrame(audit_rows, columns=["metric_name", "value", "note"])
    audit_path = DATA_OUT / "excluded_school_audit.csv"
    audit.to_csv(audit_path, index=False, encoding="utf-8-sig")

    dictionary = make_dictionary(excluded_web)
    dictionary_path = DATA_OUT / "excluded_school_data_dictionary_ko.csv"
    dictionary.to_csv(dictionary_path, index=False, encoding="utf-8-sig")

    stats = {
        "excluded_school_rows": len(excluded_web),
        "unique_excluded_school_keys": excluded_web["school_key"].nunique(),
        "valid_coordinate_count": int(excluded_web["coordinate_valid"].sum()),
        "invalid_coordinate_count": int((~excluded_web["coordinate_valid"]).sum()),
        "overlap_with_final_scenario_school_web": overlap,
    }

    metadata_updated = update_metadata(DATA_OUT / "scenario_metadata.json")
    append_report(DATA_OUT / "00_COMBINED_REPORT.md", stats)
    append_manifest(DATA_OUT / "MANIFEST.md", stats)

    coverage_path = DATA_OUT / "web_data_coverage_audit.csv"
    backup(coverage_path, ".before_excluded_layer_update")
    coverage = read_csv(coverage_path)
    extra_coverage = pd.DataFrame(
        [
            {"metric_name": "excluded_school_layer_rows", "value": len(excluded_web), "note": "event/anomaly layer rows; not school-level prediction rows"},
            {"metric_name": "excluded_school_layer_valid_coordinates", "value": stats["valid_coordinate_count"], "note": ""},
            {"metric_name": "excluded_school_layer_invalid_coordinates", "value": stats["invalid_coordinate_count"], "note": ""},
            {"metric_name": "excluded_school_layer_overlap_with_main", "value": overlap, "note": "recorded for audit; no deletion performed"},
        ]
    )
    coverage = coverage[~coverage["metric_name"].astype(str).str.startswith("excluded_school_layer_")] if not coverage.empty else coverage
    pd.concat([coverage, extra_coverage], ignore_index=True).to_csv(coverage_path, index=False, encoding="utf-8-sig")

    main_dict_path = DATA_OUT / "data_dictionary_ko.csv"
    backup(main_dict_path, ".before_excluded_layer_update")
    main_dict_backup = DATA_OUT / "data_dictionary_ko.before_excluded_layer_update.csv"
    main_dict = read_csv(main_dict_backup if main_dict_backup.exists() else main_dict_path)
    add_dict = dictionary.copy()
    add_dict["web_usage"] = add_dict["web_usage"] + "; excluded_layer"
    pd.concat([main_dict, add_dict], ignore_index=True).to_csv(main_dict_path, index=False, encoding="utf-8-sig")

    xlsx_path = DATA_OUT / "01_KEY_TABLES.xlsx"
    backup(xlsx_path, ".before_excluded_layer_update")
    with pd.ExcelWriter(xlsx_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        excluded_web.head(100).to_excel(writer, sheet_name="excluded_school_sample", index=False)
        summary.to_excel(writer, sheet_name="excluded_school_summary", index=False)
        audit.to_excel(writer, sheet_name="excluded_school_audit", index=False)
        dictionary.to_excel(writer, sheet_name="excluded_school_data_dictionary", index=False)
    style_xlsx(xlsx_path)

    missing_path = DATA_OUT / "missing_files.csv"
    missing = pd.DataFrame(
        missing_rows,
        columns=["missing_item", "item_type", "expected_path_or_column", "reason", "impact", "required_action"],
    )
    missing.to_csv(missing_path, index=False, encoding="utf-8-sig")

    copy_files = [
        excluded_web_path,
        excluded_web_json,
        summary_path,
        dictionary_path,
        audit_path,
        long_path,
        long_json,
        DATA_OUT / "scenario_metadata.json",
        DATA_OUT / "web_data_coverage_audit.csv",
        DATA_OUT / "data_dictionary_ko.csv",
        DATA_OUT / "00_COMBINED_REPORT.md",
        DATA_OUT / "01_KEY_TABLES.xlsx",
    ]
    for path in copy_files:
        if path.exists():
            shutil.copy2(path, PUBLIC_OUT / path.name)

    handoff_files = [
        DATA_OUT / "MANIFEST.md",
        DATA_OUT / "00_COMBINED_REPORT.md",
        DATA_OUT / "01_KEY_TABLES.xlsx",
        DATA_OUT / "missing_files.csv",
        DATA_OUT / "excluded_school_web.csv",
        DATA_OUT / "excluded_school_summary.csv",
        DATA_OUT / "excluded_school_audit.csv",
        DATA_OUT / "excluded_school_data_dictionary_ko.csv",
        DATA_OUT / "scenario_metadata.json",
    ]
    manifest_rows = []
    for src in handoff_files:
        dst = HANDOFF / src.name
        if src.exists():
            shutil.copy2(src, dst)
            manifest_rows.append(
                {
                    "source_path": str(src.relative_to(ROOT)),
                    "copied_path": str(dst.relative_to(ROOT)),
                    "file_size_bytes": dst.stat().st_size,
                    "sha256": sha256(dst),
                    "copy_status": "copied",
                }
            )
        else:
            manifest_rows.append(
                {
                    "source_path": str(src.relative_to(ROOT)),
                    "copied_path": str(dst.relative_to(ROOT)),
                    "file_size_bytes": 0,
                    "sha256": "",
                    "copy_status": "missing",
                }
            )
    copied_manifest = pd.DataFrame(manifest_rows)
    copied_manifest_path = HANDOFF / "copied_files_manifest.csv"
    copied_manifest.to_csv(copied_manifest_path, index=False, encoding="utf-8-sig")
    shutil.copy2(copied_manifest_path, DATA_OUT / "copied_files_manifest.csv")
    shutil.copy2(DATA_OUT / "missing_files.csv", HANDOFF / "missing_files.csv")

    # Re-copy handoff manifest after copied_files_manifest exists.
    manifest_rows.append(
        {
            "source_path": str((DATA_OUT / "copied_files_manifest.csv").relative_to(ROOT)),
            "copied_path": str(copied_manifest_path.relative_to(ROOT)),
            "file_size_bytes": copied_manifest_path.stat().st_size,
            "sha256": sha256(copied_manifest_path),
            "copy_status": "copied",
        }
    )
    pd.DataFrame(manifest_rows).to_csv(copied_manifest_path, index=False, encoding="utf-8-sig")
    shutil.copy2(copied_manifest_path, DATA_OUT / "copied_files_manifest.csv")

    results = {
        **stats,
        "metadata_updated": metadata_updated,
        "manifest_updated": (DATA_OUT / "MANIFEST.md").exists(),
        "report_updated": (DATA_OUT / "00_COMBINED_REPORT.md").exists(),
        "key_tables_updated": (DATA_OUT / "01_KEY_TABLES.xlsx").exists(),
        "handoff_file_count": len([p for p in HANDOFF.iterdir() if p.is_file()]),
    }
    print(json.dumps(results, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
