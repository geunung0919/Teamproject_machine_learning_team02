from __future__ import annotations

import hashlib
import json
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
FINAL_MODEL = "R3_multioutput_1to5_incremental_delta_HistGradientBoostingRegressor_hgb_05_deeper_regularized"
EXPECTED_2030_TOTAL = 4213774.284467546
OUT = ROOT / "data" / "v5_web_scenario_package_v1"
PUBLIC_BASE = ROOT / "public" / "data"
HANDOFF = ROOT / "handoff_for_chatgpt" / "v5_web_scenario_package_v1"

RECUR = ROOT / "data" / "v5_recursive_and_multioutput_forecasting_r3_r6_v1"
PATCH = ROOT / "data" / "v5_clean_dataset_patch_v1"
FINAL_AUDIT = ROOT / "data" / "v5_final_audit_report_for_presentation_v1"


def ensure_dirs() -> Path:
    OUT.mkdir(parents=True, exist_ok=True)
    PUBLIC_BASE.mkdir(parents=True, exist_ok=True)
    public = PUBLIC_BASE / "scenario_v5"
    if public.exists():
        i = 2
        while (PUBLIC_BASE / f"scenario_v5_v{i}").exists():
            i += 1
        public = PUBLIC_BASE / f"scenario_v5_v{i}"
    public.mkdir(parents=True, exist_ok=True)
    if HANDOFF.exists():
        shutil.rmtree(HANDOFF)
    HANDOFF.mkdir(parents=True, exist_ok=True)
    return public


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def rel(path: Path) -> str:
    try:
        return str(path.relative_to(ROOT))
    except ValueError:
        return str(path)


def read_csv(path: Path, **kw: Any) -> pd.DataFrame:
    return pd.read_csv(path, low_memory=False, **kw) if path.exists() else pd.DataFrame()


def pct_rank(s: pd.Series, ascending: bool = True) -> pd.Series:
    return s.rank(pct=True, ascending=ascending).fillna(0) * 100


def coord_quality(lat: pd.Series, lon: pd.Series) -> tuple[pd.Series, pd.Series]:
    latn = pd.to_numeric(lat, errors="coerce")
    lonn = pd.to_numeric(lon, errors="coerce")
    missing = latn.isna() | lonn.isna()
    out_bounds = (~missing) & (~latn.between(33.0, 39.5) | ~lonn.between(124.0, 132.0))
    flag = pd.Series("valid", index=lat.index, dtype=object)
    flag.loc[missing] = "missing_coordinate"
    flag.loc[out_bounds] = "out_of_korea_bounds"
    valid = flag.eq("valid")
    return valid, flag


def style_xlsx(path: Path) -> None:
    wb = load_workbook(path)
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            max_len = max(len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, min(ws.max_row, 80) + 1))
            ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 55)
    wb.save(path)


def to_json_records(df: pd.DataFrame, path: Path) -> None:
    clean = df.replace({np.nan: None})
    path.write_text(json.dumps(clean.to_dict(orient="records"), ensure_ascii=False, separators=(",", ":")), encoding="utf-8")


def md_table(df: pd.DataFrame, n: int = 20) -> str:
    if df.empty:
        return "_No rows._"
    d = df.head(n).astype(object).where(pd.notna(df.head(n)), "")
    lines = ["| " + " | ".join(map(str, d.columns)) + " |", "| " + " | ".join(["---"] * len(d.columns)) + " |"]
    for _, row in d.iterrows():
        lines.append("| " + " | ".join(str(row[c]).replace("|", "/").replace("\n", " ") for c in d.columns) + " |")
    return "\n".join(lines)


def build() -> tuple[dict[str, Any], list[Path], pd.DataFrame]:
    public = ensure_dirs()
    missing_rows: list[dict[str, Any]] = []
    created: list[Path] = []
    run_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    scen_all = read_csv(RECUR / "scenario" / "recursive_multioutput_school_predictions_2026_2030_p1.csv")
    scen = scen_all[scen_all["candidate_name"].eq(FINAL_MODEL)].copy()
    base = read_csv(PATCH / "model_views" / "scenario_base_2025.csv")
    totals = read_csv(RECUR / "scenario" / "recursive_multioutput_total_students_by_year.csv")
    total_final = totals[totals["candidate_name"].eq(FINAL_MODEL)].copy()

    if scen.empty:
        missing_rows.append({"missing_item": "final selected school scenario", "item_type": "file/content", "expected_path_or_column": str(RECUR / "scenario"), "reason": "No matching final model rows", "impact": "Cannot create school web data", "required_action": "Regenerate final scenario outputs"})
    if total_final.empty:
        missing_rows.append({"missing_item": "final selected total scenario", "item_type": "file/content", "expected_path_or_column": str(RECUR / "scenario" / "recursive_multioutput_total_students_by_year.csv"), "reason": "No matching total rows", "impact": "Cannot validate national total", "required_action": "Regenerate final totals"})

    base_cols = ["school_key", "status", "class_count", "teacher_count", "students_per_class", "students_per_teacher", "coordinate_valid", "coordinate_source", "coordinate_invalid_reason", "nearest_same_level_distance_km", "same_level_school_count_within_5km", "isolation_score"]
    base_use = base[[c for c in base_cols if c in base.columns]].copy()
    web = scen.merge(base_use, on="school_key", how="left")

    geo_path = ROOT / "data" / "raw" / "school_data_2008_2025_geocoded.csv"
    geo = read_csv(geo_path)
    if not geo.empty:
        geo25 = geo[geo["year"].eq(2025)].copy()
        geo25 = geo25.rename(columns={"시도": "sido", "행정구": "sgg", "학교급": "school_level", "학교명": "school_name", "lttud": "latitude", "lgtud": "longitude"})
        geo_cols = ["sido", "sgg", "school_level", "school_name", "latitude", "longitude", "coordinate_source"]
        geo25 = geo25[[c for c in geo_cols if c in geo25.columns]].drop_duplicates(["sido", "sgg", "school_level", "school_name"])
        web = web.merge(geo25, on=["sido", "sgg", "school_level", "school_name"], how="left", suffixes=("", "_geo"))
        if "coordinate_source_geo" in web.columns:
            web["coordinate_source"] = web["coordinate_source"].fillna(web["coordinate_source_geo"])
    else:
        missing_rows.append({"missing_item": "coordinate source file", "item_type": "file", "expected_path_or_column": str(geo_path), "reason": "File not found", "impact": "latitude/longitude unavailable", "required_action": "Provide source-backed coordinate master"})

    web["coordinate_valid"], web["coordinate_quality_flag"] = coord_quality(web.get("latitude", pd.Series(index=web.index)), web.get("longitude", pd.Series(index=web.index)))
    web["school_status_2025"] = web.get("status", "")
    for y in range(2026, 2031):
        web[f"delta_2025_{y}"] = web[f"pred_student_count_{y}"] - web["student_count_2025"]
        web[f"pct_change_2025_{y}"] = np.where(web["student_count_2025"] > 0, web[f"delta_2025_{y}"] / web["student_count_2025"], np.nan)
    web = web.rename(columns={
        "class_count": "class_count_2025",
        "teacher_count": "teacher_count_2025",
        "students_per_class": "students_per_class_2025",
        "students_per_teacher": "students_per_teacher_2025",
        "nearest_same_level_distance_km": "nearest_same_level_school_km",
        "same_level_school_count_within_5km": "same_level_school_count_5km",
    })
    web["small_school_flag_2025"] = web["student_count_2025"] <= 60
    web["small_school_flag_2030"] = web["pred_student_count_2030"] <= 60
    web["decline_pressure_flag_2030"] = (web["delta_2025_2030"] <= -30) | (web["pct_change_2025_2030"] <= -0.20)
    web["isolated_small_school_flag_2030"] = web["small_school_flag_2030"] & ((pd.to_numeric(web.get("same_level_school_count_5km"), errors="coerce") <= 1) | (pd.to_numeric(web.get("nearest_same_level_school_km"), errors="coerce") >= 5))
    q75 = pd.to_numeric(web.get("isolation_score"), errors="coerce").quantile(.75)
    web["education_gap_risk_flag_2030"] = web["small_school_flag_2030"] & (pd.to_numeric(web.get("isolation_score"), errors="coerce") >= q75)
    web["priority_score_2030"] = (
        0.30 * pct_rank((-web["delta_2025_2030"]).clip(lower=0)) +
        0.25 * pct_rank((-web["pct_change_2025_2030"]).clip(lower=0)) +
        0.25 * pct_rank(pd.to_numeric(web.get("isolation_score"), errors="coerce")) +
        0.10 * np.where(web["small_school_flag_2030"], 100, 0) +
        0.10 * (100 - pct_rank(pd.to_numeric(web.get("same_level_school_count_5km"), errors="coerce")))
    )
    web["priority_rank_national"] = web["priority_score_2030"].rank(method="first", ascending=False).astype(int)
    web["priority_rank_sido"] = web.groupby("sido")["priority_score_2030"].rank(method="first", ascending=False).astype(int)
    web["priority_rank_sgg"] = web.groupby(["sido", "sgg"])["priority_score_2030"].rank(method="first", ascending=False).astype(int)
    web["scenario_model_name"] = FINAL_MODEL
    web["scenario_type"] = "decline_pressure_scenario"
    web["scenario_note"] = "폐교 확정 예측이 아니라 2026~2030 학생 수 감소 압력과 행정 검토 우선순위를 보여주는 시나리오입니다. 학교별 파일은 P1 안정 학교 기준이며 event layer는 총합 검증에 aggregate로 포함됩니다."

    required = ["school_key", "school_name", "sido", "sgg", "school_level", "school_status_2025", "latitude", "longitude", "coordinate_valid", "coordinate_source", "coordinate_quality_flag", "student_count_2025"] + [f"pred_student_count_{y}" for y in range(2026, 2031)] + [f"delta_2025_{y}" for y in range(2026, 2031)] + [f"pct_change_2025_{y}" for y in range(2026, 2031)] + ["class_count_2025", "teacher_count_2025", "students_per_class_2025", "students_per_teacher_2025", "isolation_score", "nearest_same_level_school_km", "same_level_school_count_5km", "small_school_flag_2025", "small_school_flag_2030", "decline_pressure_flag_2030", "isolated_small_school_flag_2030", "education_gap_risk_flag_2030", "priority_score_2030", "priority_rank_national", "priority_rank_sido", "priority_rank_sgg", "scenario_model_name", "scenario_type", "scenario_note"]
    for col in required:
        if col not in web.columns:
            web[col] = np.nan
            missing_rows.append({"missing_item": col, "item_type": "column", "expected_path_or_column": col, "reason": "Column absent from source-backed inputs", "impact": "Column emitted as null for schema stability", "required_action": "Provide audited source if needed"})
    web = web[required].sort_values("priority_rank_national")

    long_rows = []
    for y in range(2025, 2031):
        tmp = web[["school_key", "school_name", "sido", "sgg", "school_level", "latitude", "longitude", "coordinate_valid", "isolation_score", "priority_score_2030", "scenario_model_name"]].copy()
        tmp["year"] = y
        tmp["student_count"] = web["student_count_2025"] if y == 2025 else web[f"pred_student_count_{y}"]
        tmp["is_observed"] = y == 2025
        tmp["is_predicted"] = y > 2025
        tmp["base_year"] = 2025
        tmp["delta_from_2025"] = 0 if y == 2025 else web[f"delta_2025_{y}"]
        tmp["pct_change_from_2025"] = 0 if y == 2025 else web[f"pct_change_2025_{y}"]
        long_rows.append(tmp)
    long = pd.concat(long_rows, ignore_index=True)

    def write_pair(df: pd.DataFrame, name: str) -> None:
        csv = OUT / f"{name}.csv"; js = OUT / f"{name}.json"
        df.to_csv(csv, index=False, encoding="utf-8-sig")
        to_json_records(df, js)
        created.extend([csv, js])

    write_pair(web, "final_scenario_school_web")
    write_pair(long, "final_scenario_school_year_long")

    def summarize(df: pd.DataFrame, groups: list[str], name: str) -> pd.DataFrame:
        rows = []
        for keys, g in df.groupby(groups + ["year"], dropna=False) if groups else df.groupby(["year"]):
            if not isinstance(keys, tuple):
                keys = (keys,)
            rec = dict(zip(groups + ["year"], keys))
            base_total = long.loc[(long["year"].eq(2025)) & np.logical_and.reduce([long[k].eq(rec[k]) for k in groups]) if groups else long["year"].eq(2025), "student_count"].sum()
            total = g["student_count"].sum()
            web2030 = web.set_index("school_key")
            school_keys = set(g["school_key"])
            w = web[web["school_key"].isin(school_keys)]
            rec.update({
                "school_count": g["school_key"].nunique(),
                "valid_coordinate_school_count": int(w["coordinate_valid"].sum()),
                "total_student_count": total,
                "mean_student_count": g["student_count"].mean(),
                "median_student_count": g["student_count"].median(),
                "small_school_count": int((g["student_count"] <= 60).sum()),
                "decline_pressure_school_count": int(w["decline_pressure_flag_2030"].sum()) if rec["year"] == 2030 else np.nan,
                "isolated_small_school_count": int(w["isolated_small_school_flag_2030"].sum()) if rec["year"] == 2030 else np.nan,
                "education_gap_risk_school_count": int(w["education_gap_risk_flag_2030"].sum()) if rec["year"] == 2030 else np.nan,
                "mean_isolation_score": w["isolation_score"].mean(),
                "total_delta_from_2025": total - base_total,
                "pct_change_from_2025": (total - base_total) / base_total if base_total else np.nan,
            })
            rows.append(rec)
        out = pd.DataFrame(rows)
        out.to_csv(OUT / name, index=False, encoding="utf-8-sig")
        created.append(OUT / name)
        return out

    national = summarize(long, [], "summary_national_by_year.csv")
    sido = summarize(long, ["sido"], "summary_sido_by_year.csv")
    sgg = summarize(long, ["sido", "sgg"], "summary_sgg_by_year.csv")
    level = summarize(long, ["school_level"], "summary_school_level_by_year.csv")
    sido_level = summarize(long, ["sido", "school_level"], "summary_sido_school_level_by_year.csv")

    top_cols = ["school_key", "school_name", "sido", "sgg", "school_level", "student_count_2025", "pred_student_count_2030", "delta_2025_2030", "pct_change_2025_2030", "isolation_score", "nearest_same_level_school_km", "same_level_school_count_5km", "priority_score_2030", "latitude", "longitude", "coordinate_valid"]
    def add_reason(df: pd.DataFrame) -> pd.DataFrame:
        out = df.copy()
        out["priority_reason"] = np.where(out["isolated_small_school_flag_2030"], "2030년 소규모 및 고립 조건이 함께 나타나 행정 검토 우선순위가 높음", "2030년 학생 수 감소 압력과 고립도/주변 학교 접근성을 함께 고려한 우선순위")
        out.insert(0, "rank", range(1, len(out) + 1))
        return out[["rank"] + top_cols + ["priority_reason"]]
    tops = {
        "top_priority_national_2030.csv": add_reason(web.sort_values("priority_score_2030", ascending=False).head(100)),
        "top_decline_national_2030.csv": add_reason(web.sort_values("delta_2025_2030").head(100)),
        "top_isolated_small_school_2030.csv": add_reason(web[web["isolated_small_school_flag_2030"]].sort_values("priority_score_2030", ascending=False).head(100)),
    }
    by_sido = web.sort_values(["sido", "priority_score_2030"], ascending=[True, False]).groupby("sido").head(20)
    tops["top_priority_by_sido_2030.csv"] = add_reason(by_sido)
    for name, df in tops.items():
        df.to_csv(OUT / name, index=False, encoding="utf-8-sig"); created.append(OUT / name)

    coord_audit = web[["school_key", "school_name", "sido", "sgg", "school_level", "latitude", "longitude", "coordinate_valid", "coordinate_quality_flag"]].copy()
    coord_audit["reason"] = coord_audit["coordinate_quality_flag"]
    coord_audit.to_csv(OUT / "coordinate_audit.csv", index=False, encoding="utf-8-sig"); created.append(OUT / "coordinate_audit.csv")

    dq = pd.DataFrame([
        ["missing_prediction", int(web[[f"pred_student_count_{y}" for y in range(2026, 2031)]].isna().any(axis=1).sum()), int(web.loc[web[[f"pred_student_count_{y}" for y in range(2026, 2031)]].isna().any(axis=1), "school_key"].nunique()), "Missing prediction in final school web data", "Do not impute; record and exclude from affected visual calculations if needed", "critical"],
        ["missing_base_student_count", int(web["student_count_2025"].isna().sum()), int(web.loc[web["student_count_2025"].isna(), "school_key"].nunique()), "Missing 2025 base student count", "Do not create arbitrary value", "critical"],
        ["negative_prediction", int((web[[f"pred_student_count_{y}" for y in range(2026, 2031)]] < 0).any(axis=1).sum()), 0, "Negative prediction check", "No negative values allowed", "critical"],
        ["coordinate_missing", int(web["coordinate_quality_flag"].eq("missing_coordinate").sum()), int(web.loc[web["coordinate_quality_flag"].eq("missing_coordinate"), "school_key"].nunique()), "Missing coordinates", "Keep row; flag coordinate_valid=False", "medium"],
        ["coordinate_outlier", int(web["coordinate_quality_flag"].eq("out_of_korea_bounds").sum()), int(web.loc[web["coordinate_quality_flag"].eq("out_of_korea_bounds"), "school_key"].nunique()), "Coordinate outside Korea bounds", "Keep row; flag coordinate_valid=False", "high"],
        ["missing_isolation_score", int(web["isolation_score"].isna().sum()), int(web.loc[web["isolation_score"].isna(), "school_key"].nunique()), "Missing isolation score", "Do not impute for risk score except percentile contributes zero", "medium"],
        ["missing_school_name", int(web["school_name"].isna().sum()), int(web.loc[web["school_name"].isna(), "school_key"].nunique()), "Missing school name", "Report as data issue", "high"],
        ["missing_region", int(web[["sido", "sgg"]].isna().any(axis=1).sum()), int(web.loc[web[["sido", "sgg"]].isna().any(axis=1), "school_key"].nunique()), "Missing region", "Report as data issue", "high"],
        ["duplicate_school_key", int(web["school_key"].duplicated().sum()), int(web.loc[web["school_key"].duplicated(), "school_key"].nunique()), "Duplicate school key", "Must resolve before web use", "critical"],
        ["event_layer_school_rows_not_included", 2173, 2173, "Final candidate school-level output contains P1 stable-school rows; event layer exists only as aggregate totals in this package", "Do not distribute event aggregate to schools; disclose in report", "medium"],
    ], columns=["issue_type", "affected_rows", "affected_schools", "description", "handling_policy", "severity"])
    dq.to_csv(OUT / "data_quality_audit.csv", index=False, encoding="utf-8-sig"); created.append(OUT / "data_quality_audit.csv")

    actual_2030_total = float(total_final.loc[total_final["year"].eq(2030), "total_students"].iloc[0]) if not total_final.empty else np.nan
    rel_diff = abs(actual_2030_total - EXPECTED_2030_TOTAL) / EXPECTED_2030_TOTAL
    validation = pd.DataFrame([{"metric": "final_selected_p1_plus_event_total_2030", "expected_value": EXPECTED_2030_TOTAL, "actual_value": actual_2030_total, "absolute_diff": actual_2030_total - EXPECTED_2030_TOTAL, "relative_diff": rel_diff, "pass": bool(rel_diff <= 0.001), "note": "Validated against final aggregate total, not P1-only school row sum."}])
    validation.to_csv(OUT / "scenario_total_validation.csv", index=False, encoding="utf-8-sig"); created.append(OUT / "scenario_total_validation.csv")

    coverage = pd.DataFrame([
        ["total_school_rows", len(web), "P1 stable-school rows in final selected school-level output"],
        ["rows_with_2026_prediction", int(web["pred_student_count_2026"].notna().sum()), ""],
        ["rows_with_2030_prediction", int(web["pred_student_count_2030"].notna().sum()), ""],
        ["rows_with_valid_coordinates", int(web["coordinate_valid"].sum()), ""],
        ["rows_with_invalid_coordinates", int((~web["coordinate_valid"]).sum()), ""],
        ["rows_with_isolation_score", int(web["isolation_score"].notna().sum()), ""],
        ["rows_with_priority_score", int(web["priority_score_2030"].notna().sum()), ""],
        ["national_total_2025", float(total_final.loc[total_final["year"].eq(2025), "total_students"].iloc[0]) if not total_final.empty else np.nan, "P1+event aggregate total"],
        ["national_total_2030", actual_2030_total, "P1+event aggregate total"],
    ], columns=["metric_name", "value", "note"])
    coverage.to_csv(OUT / "web_data_coverage_audit.csv", index=False, encoding="utf-8-sig"); created.append(OUT / "web_data_coverage_audit.csv")

    metadata = {
        "scenario_name": "V5 2026-2030 학생 수 감소 압력 시나리오",
        "base_year": 2025,
        "prediction_years": [2026, 2027, 2028, 2029, 2030],
        "final_model_name": FINAL_MODEL,
        "scenario_type": "decline_pressure_scenario",
        "interpretation": "폐교 확정 예측이 아니라 학생 수 감소 압력과 행정 검토 우선순위를 보여주는 시나리오",
        "created_at": run_time,
        "source_project_root": str(ROOT),
        "web_data_version": "v5_web_scenario_package_v1",
        "school_level_rows_note": "final_scenario_school_web contains P1 stable-school rows; event layer is represented in aggregate total validation, not distributed to school rows.",
        "risk_flag_rules": {"small_school_threshold": 60, "decline_delta_threshold": -30, "decline_pct_threshold": -0.20, "isolation_high_rule": "valid isolation_score top 25%"},
        "priority_score_formula": "0.30*decline_abs_percentile + 0.25*decline_pct_percentile + 0.25*isolation_percentile + 0.10*small_school + 0.10*nearby_lack_percentile",
    }
    (OUT / "scenario_metadata.json").write_text(json.dumps(metadata, ensure_ascii=False, indent=2), encoding="utf-8"); created.append(OUT / "scenario_metadata.json")

    dict_rows = []
    labels = {"pred_student_count_2030": "2030 예측 학생수", "isolation_score": "고립도 점수", "priority_score_2030": "2030 우선순위 점수"}
    for c in web.columns:
        dict_rows.append({"column_name": c, "korean_label": labels.get(c, c), "description_ko": "웹 시각화 및 발표용 최종 시나리오 컬럼", "dtype": str(web[c].dtype), "example": web[c].dropna().iloc[0] if web[c].notna().any() else "", "web_usage": "map/detail/ranking/filter"})
    dd = pd.DataFrame(dict_rows)
    dd.to_csv(OUT / "data_dictionary_ko.csv", index=False, encoding="utf-8-sig"); created.append(OUT / "data_dictionary_ko.csv")

    report = "\n".join([
        "# V5 Web Scenario Data Package Report", "",
        "## 1. 작업 목적", "최종 V5 모델 산출물을 웹 시각화와 발표에 바로 사용할 수 있는 versioned CSV/JSON 패키지로 정리했다.", "",
        "## 2. 사용한 최종 모델", FINAL_MODEL, "",
        "## 3. 입력 파일 검색 결과", f"- final scenario rows: {len(web)}\n- final aggregate total rows: {len(total_final)}\n- coordinate source: {rel(geo_path)}", "",
        "## 4. 생성된 웹용 데이터 목록", "\n".join(f"- {p.name}" for p in created), "",
        "## 5. 학교 단위 데이터 스키마", md_table(dd, 25), "",
        "## 6. long format 데이터 스키마", "2025는 observed/base row, 2026~2030은 predicted row로 구성했다.", "",
        "## 7. 요약 데이터 설명", "전국/시도/시군구/학교급/시도-학교급별 year summary를 생성했다. 학교별 row는 P1 안정 학교 기준이다.", "",
        "## 8. ranking/top-k 데이터 설명", "priority, decline, isolated small-school 기준 Top-K 파일을 생성했다.", "",
        "## 9. risk flag 생성 규칙", json.dumps(metadata["risk_flag_rules"], ensure_ascii=False), "",
        "## 10. priority_score_2030 공식", metadata["priority_score_formula"], "",
        "## 11. 좌표 검증 결과", md_table(coord_audit["coordinate_quality_flag"].value_counts().rename_axis("flag").reset_index(name="count")), "",
        "## 12. coverage audit 결과", md_table(coverage), "",
        "## 13. 2030 총합 검증 결과", md_table(validation), "",
        "## 14. 웹사이트 사용 가이드", "지도에는 coordinate_valid=True row를 우선 사용하고, invalid row는 표/검색 결과에는 유지한다. priority_score_2030은 행정 검토 우선순위 점수다.", "",
        "## 15. 핵심 주의사항", "이 데이터는 폐교 확정 예측 데이터가 아니며, 2026~2030 학생 수 감소 압력과 행정 검토 우선순위를 보여주는 시나리오 데이터입니다.",
    ])
    (OUT / "00_COMBINED_REPORT.md").write_text(report, encoding="utf-8"); created.append(OUT / "00_COMBINED_REPORT.md")

    with pd.ExcelWriter(OUT / "01_KEY_TABLES.xlsx", engine="openpyxl") as writer:
        pd.DataFrame([metadata]).to_excel(writer, sheet_name="README", index=False)
        web.head(1000).to_excel(writer, sheet_name="school_web_sample", index=False)
        national.to_excel(writer, sheet_name="national_summary", index=False)
        sido[sido["year"].eq(2030)].to_excel(writer, sheet_name="sido_summary_2030", index=False)
        level[level["year"].eq(2030)].to_excel(writer, sheet_name="school_level_summary_2030", index=False)
        tops["top_priority_national_2030.csv"].to_excel(writer, sheet_name="top_priority_national_2030", index=False)
        coord_audit["coordinate_quality_flag"].value_counts().rename_axis("flag").reset_index(name="count").to_excel(writer, sheet_name="coordinate_audit_summary", index=False)
        coverage.to_excel(writer, sheet_name="coverage_audit", index=False)
        validation.to_excel(writer, sheet_name="scenario_total_validation", index=False)
        dd.to_excel(writer, sheet_name="data_dictionary_ko", index=False)
        pd.DataFrame({"file": [p.name for p in created]}).to_excel(writer, sheet_name="generated_files", index=False)
        pd.DataFrame(missing_rows).to_excel(writer, sheet_name="missing_files", index=False)
    style_xlsx(OUT / "01_KEY_TABLES.xlsx"); created.append(OUT / "01_KEY_TABLES.xlsx")

    # Copy web files to versioned public folder.
    for p in created:
        if p.suffix.lower() in [".csv", ".json", ".md", ".xlsx"]:
            shutil.copy2(p, public / p.name)

    missing = pd.DataFrame(missing_rows, columns=["missing_item", "item_type", "expected_path_or_column", "reason", "impact", "required_action"])
    missing.to_csv(OUT / "missing_files.csv", index=False, encoding="utf-8-sig")

    manifest_text = "\n".join([
        "# V5 Web Scenario Package Manifest", "",
        f"- run_time: {run_time}",
        f"- project_root: {ROOT}",
        f"- output_folder: {OUT}",
        f"- web_output_folder: {public}",
        f"- final_model_name: {FINAL_MODEL}",
        f"- expected_2030_total: {EXPECTED_2030_TOTAL}",
        f"- actual_2030_total: {actual_2030_total}",
        f"- total_validation_pass: {bool(validation['pass'].iloc[0])}",
        f"- school_row_count: {len(web)}",
        f"- valid_coordinate_count: {int(web['coordinate_valid'].sum())}",
        f"- invalid_coordinate_count: {int((~web['coordinate_valid']).sum())}",
        f"- created_files: {len(created)}",
        f"- copied_files: public copy + handoff",
        f"- missing_files: {len(missing)}",
        "- warnings: school-level web rows are P1 stable-school rows; event layer is included only in aggregate total validation.",
    ])
    (OUT / "MANIFEST.md").write_text(manifest_text, encoding="utf-8"); created.append(OUT / "MANIFEST.md")

    handoff_sources = [OUT / "MANIFEST.md", OUT / "00_COMBINED_REPORT.md", OUT / "01_KEY_TABLES.xlsx"]
    copied_rows = []
    for src in handoff_sources:
        dst = HANDOFF / src.name
        shutil.copy2(src, dst)
        copied_rows.append({"source_path": str(src), "copied_path": str(dst), "file_size_bytes": dst.stat().st_size, "sha256": sha256(dst), "copy_status": "copied"})
    copied = pd.DataFrame(copied_rows)
    copied.to_csv(HANDOFF / "copied_files_manifest.csv", index=False, encoding="utf-8-sig")
    missing.to_csv(HANDOFF / "missing_files.csv", index=False, encoding="utf-8-sig")

    return {
        "public": public,
        "created": created,
        "missing": missing,
        "actual_2030_total": actual_2030_total,
        "total_validation_pass": bool(validation["pass"].iloc[0]),
        "school_row_count": len(web),
        "valid_coordinate_count": int(web["coordinate_valid"].sum()),
        "invalid_coordinate_count": int((~web["coordinate_valid"]).sum()),
    }, created, missing


def main() -> None:
    result, _, _ = build()
    print("V5 web scenario package completed.")
    print("\nOutput:")
    print("- data/v5_web_scenario_package_v1/")
    print(f"- {result['public'].as_posix()}/")
    print("- handoff_for_chatgpt/v5_web_scenario_package_v1/")
    print("\nFinal model:")
    print(f"- {FINAL_MODEL}")
    print("\nValidation:")
    print(f"- expected_2030_total: {EXPECTED_2030_TOTAL}")
    print(f"- actual_2030_total: {result['actual_2030_total']}")
    print(f"- total_validation_pass: {result['total_validation_pass']}")
    print("\nKey files:")
    for name in ["final_scenario_school_web.csv", "final_scenario_school_web.json", "final_scenario_school_year_long.csv", "scenario_metadata.json", "00_COMBINED_REPORT.md", "01_KEY_TABLES.xlsx", "MANIFEST.md"]:
        print(f"- {name}")


if __name__ == "__main__":
    main()
